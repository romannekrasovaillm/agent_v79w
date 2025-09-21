#codex
import asyncio
import json
import re
import difflib
import time
import base64
import uuid
import urllib3
import threading
import hashlib
import math
import random
import string
import logging
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import streamlit as st
from urllib.parse import urlparse, urljoin
import trafilatura
from datetime import datetime, date
from io import BytesIO
from collections import defaultdict, Counter
from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import List, Dict, Any, Optional, Tuple, Union, Callable, Set
from pathlib import Path
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor

# Third-party imports
try:
    import numpy as np
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    logging.warning("scikit-learn не установлен. Семантический поиск будет недоступен.")

try:
    import nltk
    from nltk.tokenize import sent_tokenize, word_tokenize
    NLTK_AVAILABLE = True
except ImportError:
    NLTK_AVAILABLE = False
    logging.warning("NLTK не установлен. Обработка текста будет ограничена.")

try:
    from bs4 import BeautifulSoup, NavigableString, Tag
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False
    logging.warning("BeautifulSoup не установлен. Парсинг HTML будет ограничен.")

try:
    import html2text
    HTML2TEXT_AVAILABLE = True
except ImportError:
    HTML2TEXT_AVAILABLE = False
    logging.warning("html2text не установлен. Конвертация HTML в markdown недоступна.")

try:
    from ddgs import DDGS
    DDGS_AVAILABLE = True
except ImportError:
    DDGS_AVAILABLE = False
    logging.warning("duckduckgo-search не установлен. Веб-поиск будет недоступен.")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    logging.warning("Playwright не установлен. Браузер будет недоступен.")

# Excel support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    try:
        import xlsxwriter
        EXCEL_AVAILABLE = True
        EXCEL_ENGINE = 'xlsxwriter'
    except ImportError:
        EXCEL_AVAILABLE = False
        logging.warning("Поддержка Excel недоступна. Установите: pip install openpyxl или xlsxwriter")

if EXCEL_AVAILABLE and 'openpyxl' in locals():
    EXCEL_ENGINE = 'openpyxl'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# CURRENT DATE PLACEHOLDER - Update this with the actual current date
CURRENT_DATE = datetime.now()
CURRENT_DATE_STR = CURRENT_DATE.strftime("%Y-%m-%d")
CURRENT_DATE_FORMATTED = CURRENT_DATE.strftime("%A, %B %d, %Y")

# Download NLTK data if available
if NLTK_AVAILABLE:
    try:
        nltk.data.find('tokenizers/punkt')
    except LookupError:
        try:
            nltk.download('punkt')
        except:
            NLTK_AVAILABLE = False


@dataclass
class TaskContext:
    """Контекст задачи с информацией о намерениях пользователя."""
    query: str
    intent: str
    requires_search: bool = False
    requires_browser: bool = False
    requires_computation: bool = False
    requires_excel: bool = False
    complexity: str = "simple"  # simple, medium, complex
    domain: str = "general"
    keywords: List[str] = field(default_factory=list)
    timestamp: datetime = field(default_factory=datetime.now)
    urgency: str = "normal"  # low, normal, high, critical
    expected_sources: int = 3
    temporal_context: str = "current"  # historical, current, future
    user_goal: str = "information"  # information, action, analysis, export
    confidence_score: float = 0.8
    meta_analysis: Dict[str, Any] = field(default_factory=dict)
    original_query: Optional[str] = None
    rewrite_strategy: str = "no_rewrite"
    rewrite_features: Dict[str, Any] = field(default_factory=dict)
    

@dataclass
class ExecutionPlan:
    """План выполнения задачи."""
    steps: List[Dict[str, Any]] = field(default_factory=list)
    estimated_time: float = 0.0
    confidence: float = 0.8
    fallback_plan: Optional[List[Dict[str, Any]]] = None
    reasoning: str = ""
    risk_assessment: Dict[str, str] = field(default_factory=dict)
    success_criteria: List[str] = field(default_factory=list)
    adaptability_level: str = "medium"  # low, medium, high
    current_step_index: int = 0
    completed_steps: int = 0
    progress_notes: List[str] = field(default_factory=list)
    

@dataclass
class ToolResult:
    """Результат выполнения инструмента."""
    tool_name: str
    success: bool
    data: Any
    metadata: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None
    execution_time: float = 0.0
    confidence: float = 0.8


@dataclass
class QueryFeatures:
    """Лингвистические признаки запроса для адаптивной переформулировки."""

    anaphora: bool = False
    subordination: bool = False
    mismatch: bool = False
    presupposition: bool = False
    pragmatics: bool = False
    rarity: bool = False
    negation: bool = False
    superlative: bool = False
    polysemy: bool = False
    answerability: bool = True
    excessive: bool = False
    subjectivity: bool = False
    ambiguity: bool = False
    grounding: bool = False
    constraints: bool = False
    entities: bool = False
    specialization: bool = False
    length: int = 0
    clarity_score: float = 1.0
    confidence: float = 0.5

    def to_dict(self) -> Dict[str, Any]:
        """Преобразует признаки в словарь для логирования и метаданных."""
        return asdict(self)

    def flagged_features(self) -> List[str]:
        """Возвращает список активных признаков."""
        return [
            key for key, value in self.to_dict().items()
            if isinstance(value, bool) and value
        ]


class QueryFeatureExtractor:
    """Извлекает лингвистические признаки запроса согласно спецификации."""

    ANAPHORA_PRONOUNS = {
        'он', 'она', 'они', 'оно', 'его', 'ее', 'их', 'это', 'этот',
        'эта', 'эти', 'там', 'тут', 'такой', 'такое', 'подобное'
    }
    SUBORDINATE_MARKERS = {
        'что', 'котор', 'если', 'когда', 'пока', 'поскольку', 'потому что',
        'чтобы', 'как только', 'когда бы'
    }
    PRAGMATIC_MARKERS = {
        'пожалуйста', 'не могли бы', 'будьте добры', 'нужно', 'важно',
        'желательно', 'прошу', 'давай'
    }
    PRESUPPOSITION_MARKERS = {
        'снова', 'еще', 'опять', 'как обычно', 'по-прежнему'
    }
    SUBJECTIVE_MARKERS = {
        'лучший', 'худший', 'нравится', 'не нравится', 'кажется',
        'интересный', 'увлекательный', 'ужасный', 'потрясающий'
    }
    CONSTRAINT_MARKERS = {
        'только', 'исключительно', 'не более', 'не менее', 'ровно',
        'строго', 'в течение', 'до ', 'после '
    }
    GROUNDING_MARKERS = {
        'согласно', 'по данным', 'по информации', 'источник', 'подтвердите',
        'с опорой', 'сошлитесь'
    }
    POLYSEMY_CANDIDATES = {
        'банк', 'мышь', 'лук', 'ключ', 'мир', 'коса', 'ручка', 'класс',
        'депозит', 'ставка', 'счет'
    }
    GENERIC_NOUNS = {
        'вещь', 'тема', 'ситуация', 'проблема', 'вопрос', 'дело', 'информация'
    }
    QUESTION_WORDS = {
        'кто', 'что', 'где', 'когда', 'почему', 'зачем', 'как', 'сколько'
    }
    DIRECTIVE_VERBS = {
        'объясни', 'расскажи', 'покажи', 'поясни', 'составь', 'проанализируй',
        'найди', 'подбери', 'создай', 'подскажи'
    }
    COMMON_WORDS = {
        'что', 'как', 'когда', 'почему', 'где', 'кто', 'можно', 'нужно',
        'сделай', 'прошу', 'расскажи', 'объясни', 'из', 'для', 'или',
        'и', 'о', 'про', 'это', 'какой', 'какая', 'сколько', 'дай',
        'на', 'в', 'с', 'по', 'а'
    }
    DOMAIN_TERMS = {
        'rl', 'reinforcement', 'api', 'sdk', 'kpi', 'roi', 'ml', 'ai',
        'devops', 'sql', 'saas', 'crm', 'erp', 'финтех', 'биоинформатика'
    }

    def extract(self, query: str) -> QueryFeatures:
        """Определяет признаки исходного запроса."""
        features = QueryFeatures()
        normalized = query.strip()
        lower_query = normalized.lower()
        tokens = re.findall(r"[\wёЁ]+", lower_query)

        features.length = len(tokens)
        features.anaphora = any(token in self.ANAPHORA_PRONOUNS for token in tokens)
        features.subordination = any(
            re.search(rf",\s*{marker}", lower_query)
            for marker in self.SUBORDINATE_MARKERS
        )
        features.pragmatics = any(marker in lower_query for marker in self.PRAGMATIC_MARKERS)
        features.presupposition = any(marker in lower_query for marker in self.PRESUPPOSITION_MARKERS)
        features.negation = ' не ' in f" {lower_query} " or 'нет' in tokens
        features.superlative = any(word in lower_query for word in ['самый', 'наиболее', 'наименьший', 'крайне'])
        features.polysemy = any(token in self.POLYSEMY_CANDIDATES for token in tokens)
        features.subjectivity = any(marker in lower_query for marker in self.SUBJECTIVE_MARKERS)
        features.grounding = any(marker in lower_query for marker in self.GROUNDING_MARKERS)
        features.constraints = bool(re.search(r"\d", lower_query)) or any(
            marker in lower_query for marker in self.CONSTRAINT_MARKERS
        )
        features.entities = bool(re.search(r"[A-ZА-ЯЁ]{2,}", query)) or bool(
            re.search(r"\b[А-ЯЁA-Z][а-яёa-z]+", query)
        )
        features.specialization = any(token in self.DOMAIN_TERMS for token in tokens) or bool(
            re.search(r"\b[A-Z]{2,}\b", query)
        )

        rare_tokens = [token for token in tokens if len(token) > 10 or token not in self.COMMON_WORDS]
        features.rarity = bool(tokens) and len(rare_tokens) / max(len(tokens), 1) > 0.35

        features.excessive = features.length > 28 or len(normalized) > 200
        features.answerability = (
            '?' in normalized
            or any(token in self.QUESTION_WORDS for token in tokens[:3])
            or any(lower_query.startswith(verb) for verb in self.DIRECTIVE_VERBS)
        )

        generic_opening = bool(tokens) and tokens[0] in self.GENERIC_NOUNS
        features.ambiguity = features.anaphora or features.polysemy or generic_opening

        question_like = any(token in self.QUESTION_WORDS for token in tokens)
        imperative_like = any(verb in lower_query for verb in self.DIRECTIVE_VERBS)
        features.mismatch = question_like and imperative_like and '?' not in normalized

        penalties = [
            features.anaphora, features.subordination, features.mismatch,
            features.presupposition, features.rarity, features.negation,
            features.superlative, features.polysemy, features.subjectivity,
            features.ambiguity, features.constraints
        ]
        penalty = sum(0.08 for flag in penalties if flag)
        features.clarity_score = max(0.0, min(1.0, 1.0 - penalty))
        features.confidence = 0.5 + 0.5 * features.clarity_score

        return features

    def summarize(self, features: QueryFeatures) -> str:
        """Готовит краткое текстовое описание ключевых признаков."""
        descriptions = {
            'anaphora': 'есть анафорические ссылки',
            'subordination': 'сложноподчинённые конструкции',
            'mismatch': 'смешение вопросительной и повелительной формы',
            'presupposition': 'скрытые предпосылки',
            'pragmatics': 'прагматические маркеры',
            'rarity': 'редкая лексика',
            'negation': 'отрицания',
            'superlative': 'превосходная степень',
            'polysemy': 'многозначные термины',
            'excessive': 'избыточный объём',
            'subjectivity': 'субъективные формулировки',
            'ambiguity': 'общие или двусмысленные слова',
            'grounding': 'запрос на обоснование источниками',
            'constraints': 'много ограничений',
            'entities': 'именованные сущности',
            'specialization': 'узкоспециализированная терминология'
        }

        flagged = [descriptions[key] for key in descriptions if getattr(features, key, False)]
        if not flagged:
            return "Запрос ясен, критичных признаков не обнаружено."

        summary = "; ".join(flagged)
        summary += f". Индекс ясности: {features.clarity_score:.2f}"
        return summary


class RewriteStrategy(Enum):
    """Перечень стратегий переформулировки запросов."""

    NO_REWRITE = "no_rewrite"
    PARAPHRASE = "paraphrase"
    SIMPLIFY = "simplify"
    DISAMBIGUATE = "disambiguate"
    EXPAND = "expand"
    CLARIFY = "clarify"


@dataclass
class QueryOptimizationResult:
    """Результат адаптивной переформулировки запроса."""

    original_query: str
    optimized_query: str
    strategy: RewriteStrategy
    features: QueryFeatures
    feature_summary: str
    notes: List[str] = field(default_factory=list)
    duration: float = 0.0

    @property
    def changed(self) -> bool:
        return self._normalize(self.optimized_query) != self._normalize(self.original_query)

    def to_metadata(self) -> Dict[str, Any]:
        return {
            "original_query": self.original_query,
            "optimized_query": self.optimized_query,
            "strategy": self.strategy.value,
            "changed": self.changed,
            "feature_summary": self.feature_summary,
            "features": self.features.to_dict(),
            "notes": self.notes,
            "duration": self.duration
        }

    @staticmethod
    def _normalize(text: str) -> str:
        return re.sub(r"\s+", " ", text or "").strip().lower()


class RewriteStrategySelector:
    """Определяет подходящую стратегию переформулировки по признакам."""

    def __init__(self, clarity_threshold: float = 0.68):
        self.clarity_threshold = clarity_threshold

    def select(self, features: QueryFeatures) -> RewriteStrategy:
        if (
            features.clarity_score >= self.clarity_threshold
            and not features.ambiguity
            and not features.subordination
            and not features.rarity
            and not features.constraints
            and not features.polysemy
        ):
            return RewriteStrategy.NO_REWRITE

        scores: Dict[RewriteStrategy, float] = {
            RewriteStrategy.PARAPHRASE: 0.0,
            RewriteStrategy.SIMPLIFY: 0.0,
            RewriteStrategy.DISAMBIGUATE: 0.0,
            RewriteStrategy.EXPAND: 0.0,
            RewriteStrategy.CLARIFY: 0.0
        }

        if features.subordination:
            scores[RewriteStrategy.DISAMBIGUATE] += 1.0
            scores[RewriteStrategy.SIMPLIFY] += 0.4

        if features.pragmatics:
            scores[RewriteStrategy.SIMPLIFY] += 1.0

        if features.constraints:
            scores[RewriteStrategy.EXPAND] += 0.8
            scores[RewriteStrategy.CLARIFY] += 0.2

        if features.specialization:
            scores[RewriteStrategy.EXPAND] += 0.7
            scores[RewriteStrategy.CLARIFY] += 0.5

        if features.rarity:
            scores[RewriteStrategy.CLARIFY] += 1.0

        if features.polysemy or features.anaphora or features.ambiguity:
            scores[RewriteStrategy.DISAMBIGUATE] += 1.1
            scores[RewriteStrategy.CLARIFY] += 0.4

        if features.answerability:
            scores[RewriteStrategy.PARAPHRASE] += 0.9

        if features.excessive:
            scores[RewriteStrategy.SIMPLIFY] += 0.6

        if features.grounding:
            scores[RewriteStrategy.EXPAND] += 0.3

        if features.subjectivity:
            scores[RewriteStrategy.CLARIFY] += 0.3

        if features.mismatch or features.presupposition:
            scores[RewriteStrategy.DISAMBIGUATE] += 0.5

        best_strategy = max(scores, key=scores.get)
        best_score = scores[best_strategy]

        if best_score <= 0.0:
            return RewriteStrategy.NO_REWRITE if features.clarity_score >= 0.5 else RewriteStrategy.SIMPLIFY

        return best_strategy


class BaseQueryRewriter:
    """Базовый класс для стратегий переформулировки."""

    def __init__(self, gigachat_client: 'GigaChatClient', strategy: RewriteStrategy):
        self.client = gigachat_client
        self.strategy = strategy

    def rewrite(self, query: str, feature_summary: str = "") -> str:
        raise NotImplementedError


class LLMQueryRewriter(BaseQueryRewriter):
    """Универсальный LLM-реализатор стратегий переформулировки."""

    def __init__(self, gigachat_client: 'GigaChatClient', strategy: RewriteStrategy, prompt_template: str):
        super().__init__(gigachat_client, strategy)
        self.prompt_template = prompt_template

    def rewrite(self, query: str, feature_summary: str = "") -> str:
        prompt = self.prompt_template.format(query=query)
        system_prompt = (
            "Ты переписываешь пользовательские запросы так, чтобы они были понятны поисковым системам и LLM. "
            "Сохраняй смысл, но устраняй двусмысленности и сложные конструкции."
        )

        if feature_summary:
            prompt += f"\n\nКонтекст признаков запроса: {feature_summary}"

        try:
            response = self.client.chat(
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.2,
                max_tokens=300
            )

            if response and 'choices' in response:
                content = response['choices'][0]['message']['content'].strip()
                return self._cleanup_response(content) or query

        except Exception as exc:
            logger.warning("Не удалось переписать запрос стратегией %s: %s", self.strategy.value, exc)

        return query

    @staticmethod
    def _cleanup_response(text: str) -> str:
        if not text:
            return ""

        cleaned = text.strip()
        if cleaned.startswith('```'):
            cleaned = re.sub(r"```[a-zA-Z]*", "", cleaned)
            cleaned = cleaned.replace('```', '')
        cleaned = cleaned.strip()
        if '\n' in cleaned:
            cleaned = cleaned.split('\n')[0].strip()
        if cleaned.startswith('"') and cleaned.endswith('"'):
            cleaned = cleaned[1:-1].strip()
        return cleaned


class QueryOptimizer:
    """Оркестратор адаптивной переформулировки запросов."""

    def __init__(
        self,
        gigachat_client: 'GigaChatClient',
        feature_extractor: Optional[QueryFeatureExtractor] = None,
        strategy_selector: Optional[RewriteStrategySelector] = None
    ):
        self.client = gigachat_client
        self.feature_extractor = feature_extractor or QueryFeatureExtractor()
        self.strategy_selector = strategy_selector or RewriteStrategySelector()
        self.rewriters: Dict[RewriteStrategy, LLMQueryRewriter] = {
            RewriteStrategy.PARAPHRASE: LLMQueryRewriter(
                gigachat_client,
                RewriteStrategy.PARAPHRASE,
                "Перефразируй следующий вопрос, сохранив его смысл, но используя другие слова и структуру: {query}"
            ),
            RewriteStrategy.SIMPLIFY: LLMQueryRewriter(
                gigachat_client,
                RewriteStrategy.SIMPLIFY,
                "Упрости следующий вопрос, убрав сложные конструкции и вложенные предложения: {query}"
            ),
            RewriteStrategy.DISAMBIGUATE: LLMQueryRewriter(
                gigachat_client,
                RewriteStrategy.DISAMBIGUATE,
                "Уточни все неоднозначные термины и местоимения в следующем вопросе: {query}"
            ),
            RewriteStrategy.EXPAND: LLMQueryRewriter(
                gigachat_client,
                RewriteStrategy.EXPAND,
                "Расширь следующий вопрос, добавив релевантный контекст и детали: {query}"
            ),
            RewriteStrategy.CLARIFY: LLMQueryRewriter(
                gigachat_client,
                RewriteStrategy.CLARIFY,
                "Определи и разъясни специализированные или редкие термины в следующем вопросе: {query}"
            )
        }

    def optimize_query(self, original_query: str) -> QueryOptimizationResult:
        start_time = time.time()
        features = self.feature_extractor.extract(original_query)
        feature_summary = self.feature_extractor.summarize(features)
        strategy = self.strategy_selector.select(features)
        optimized_query = original_query
        notes: List[str] = []

        if strategy == RewriteStrategy.NO_REWRITE:
            notes.append("Переформулировка не потребовалась")
        else:
            rewriter = self.rewriters.get(strategy)
            if rewriter is None:
                notes.append("Стратегия не поддерживается, оставлена исходная формулировка")
                strategy = RewriteStrategy.NO_REWRITE
            else:
                candidate = rewriter.rewrite(original_query, feature_summary)
                candidate_normalized = self._normalize(candidate)
                original_normalized = self._normalize(original_query)
                if candidate_normalized and candidate_normalized != original_normalized:
                    optimized_query = candidate
                    notes.append(f"Применена стратегия {strategy.value}")
                else:
                    strategy = RewriteStrategy.NO_REWRITE
                    optimized_query = original_query
                    notes.append("Переписанный запрос не дал улучшений, используется исходный текст")

        duration = time.time() - start_time

        return QueryOptimizationResult(
            original_query=original_query,
            optimized_query=optimized_query,
            strategy=strategy,
            features=features,
            feature_summary=feature_summary,
            notes=notes,
            duration=duration
        )

    @staticmethod
    def _normalize(text: str) -> str:
        return re.sub(r"\s+", " ", text or "").strip().lower()


class ResponseEvaluator:
    """Оценивает качество ответа на основе комбинированной награды."""

    def __init__(
        self,
        gigachat_client: Optional['GigaChatClient'] = None,
        weights: Tuple[float, float, float] = (0.6, 0.3, 0.1)
    ):
        self.client = gigachat_client
        self.weights = weights

    def evaluate(self, response: str, reference: Optional[str]) -> Optional[Dict[str, Any]]:
        if not response or not reference:
            return None

        reward, components = self.calculate_reward(response, reference)
        components['reward'] = reward
        components['weights'] = list(self.weights)
        return components

    def calculate_reward(
        self,
        response: str,
        reference: str,
        weights: Optional[Tuple[float, float, float]] = None
    ) -> Tuple[float, Dict[str, float]]:
        weights = weights or self.weights
        alpha, beta, gamma = weights

        s_llm = self._llm_judge_score(response, reference)
        s_fuzz = self._fuzzy_match_score(response, reference)
        s_bleu = self._bleu1(response, reference)

        reward = alpha * s_llm + beta * s_fuzz + gamma * s_bleu
        return reward, {'llm': s_llm, 'fuzzy': s_fuzz, 'bleu1': s_bleu}

    def _llm_judge_score(self, response: str, reference: str) -> float:
        if not response or not reference:
            return 0.0

        if not self.client:
            return self._heuristic_semantic_score(response, reference)

        prompt = (
            "Оцени, насколько ответ ассистента соответствует референсу по смыслу и фактам. "
            "Верни JSON вида {\"score\": число от 0 до 1}."
            f"\n\nРЕФЕРЕНС:\n{reference}\n\nОТВЕТ:\n{response}"
        )

        try:
            evaluation = self.client.chat(
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Ты строгий судья качества ответов. Оцени точность и полноту, не добавляй лишний текст."
                        )
                    },
                    {"role": "user", "content": prompt}
                ],
                temperature=0.0,
                max_tokens=200
            )
            if evaluation and 'choices' in evaluation:
                content = evaluation['choices'][0]['message']['content']
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    try:
                        data = json.loads(json_match.group())
                        score = float(data.get('score', 0.0))
                        return max(0.0, min(1.0, score))
                    except (ValueError, TypeError, json.JSONDecodeError):
                        pass
                number_match = re.search(r"0?\.\d+|1\.0", content)
                if number_match:
                    try:
                        score = float(number_match.group())
                        return max(0.0, min(1.0, score))
                    except ValueError:
                        pass
        except Exception:
            logger.debug("LLM-оенка не удалась, используем эвристику", exc_info=True)

        return self._heuristic_semantic_score(response, reference)

    def _heuristic_semantic_score(self, response: str, reference: str) -> float:
        response_tokens = set(self._normalize_tokens(response))
        reference_tokens = set(self._normalize_tokens(reference))
        if not reference_tokens:
            return 0.0
        overlap = len(response_tokens & reference_tokens)
        return overlap / max(len(reference_tokens), 1)

    def _fuzzy_match_score(self, response: str, reference: str) -> float:
        if not response or not reference:
            return 0.0
        return difflib.SequenceMatcher(None, response, reference).ratio()

    def _bleu1(self, response: str, reference: str) -> float:
        response_tokens = self._normalize_tokens(response)
        reference_tokens = self._normalize_tokens(reference)
        if not response_tokens or not reference_tokens:
            return 0.0

        response_counts = Counter(response_tokens)
        reference_counts = Counter(reference_tokens)
        overlap = sum(min(response_counts[token], reference_counts[token]) for token in response_counts)
        precision = overlap / len(response_tokens)

        if len(response_tokens) < len(reference_tokens) and len(response_tokens) > 0:
            brevity_penalty = math.exp(1 - len(reference_tokens) / max(len(response_tokens), 1))
        else:
            brevity_penalty = 1.0

        return precision * brevity_penalty

    @staticmethod
    def _normalize_tokens(text: str) -> List[str]:
        return re.findall(r"[\wёЁ]+", (text or '').lower())

class AdvancedIntentAnalyzer:
    """Упрощенный анализатор намерений пользователя."""

    def __init__(self, gigachat_client):
        # Клиент сохраняем для совместимости, но для анализа он не требуется
        self.client = gigachat_client
        self.search_markers = {
            'найди', 'поищи', 'что такое', 'кто такой', 'узнай', 'данные',
            'информация', 'новости', 'обзор'
        }
        self.browser_markers = {
            'сайт', 'страница', 'url', 'ссылка', 'перейди', 'открой', 'заполни'
        }
        self.computation_markers = {
            'посчитай', 'вычисли', 'рассчитай', 'сколько', 'проанализируй',
            'анализ', 'формула', 'график'
        }
        self.excel_markers = {
            'excel', 'таблица', 'экспорт', 'выгрузи', 'xlsx', 'csv', 'отчет'
        }
        self.freshness_markers = {'сегодня', 'сейчас', 'актуальный', 'текущий'}

    def analyze_with_llm(self, query: str, original_query: Optional[str] = None) -> TaskContext:
        """Выполняет детерминированный анализ запроса без обращения к LLM."""
        return self._rule_based_analysis(query, original_query)

    def _rule_based_analysis(self, query: str, original_query: Optional[str] = None) -> TaskContext:
        normalized = query.strip()
        lower_query = normalized.lower()

        requires_browser = any(marker in lower_query for marker in self.browser_markers)
        requires_excel = any(marker in lower_query for marker in self.excel_markers)
        requires_computation = any(marker in lower_query for marker in self.computation_markers)

        question_starters = {'кто', 'что', 'где', 'когда', 'почему', 'зачем', 'как', 'сколько'}
        starts_with_question = any(lower_query.startswith(word + ' ') for word in question_starters)
        has_question_mark = '?' in normalized
        contains_search_marker = any(marker in lower_query for marker in self.search_markers)

        requires_search = bool(
            contains_search_marker
            or has_question_mark
            or starts_with_question
        ) and not requires_browser

        intent = 'general'
        if requires_browser:
            intent = 'web_interaction'
        elif requires_computation:
            intent = 'computation'
        elif requires_excel:
            intent = 'excel_export'
        elif requires_search:
            intent = 'search'

        if intent == 'general' and (contains_search_marker or has_question_mark or starts_with_question):
            requires_search = True
            intent = 'search'

        complexity = 'simple'
        if len(normalized) > 120 or sum([
            requires_browser,
            requires_computation,
            requires_excel,
            requires_search
        ]) > 1:
            complexity = 'medium'
        if len(normalized) > 240:
            complexity = 'complex'

        temporal_context = 'current'
        if any(marker in lower_query for marker in {'завтра', 'будет', 'план', 'будущ'}):
            temporal_context = 'future'
        elif any(marker in lower_query for marker in {'прошлый', 'ранее', 'история', 'ретроспектива'}):
            temporal_context = 'historical'
        elif any(marker in lower_query for marker in self.freshness_markers):
            temporal_context = 'current'

        user_goal = 'information'
        if intent == 'computation':
            user_goal = 'analysis'
        elif intent == 'excel_export':
            user_goal = 'export'
        elif intent == 'web_interaction':
            user_goal = 'action'

        keywords = self._extract_keywords(query)
        domain = self._detect_domain(lower_query, keywords)

        reasoning_parts = []
        if requires_browser:
            reasoning_parts.append('Обнаружены глаголы взаимодействия с веб-сайтом.')
        if requires_search:
            reasoning_parts.append('Запрос похож на вопрос и требует поиска информации.')
        if requires_computation:
            reasoning_parts.append('Есть указания на необходимость расчетов или анализа.')
        if requires_excel:
            reasoning_parts.append('В тексте упомянут экспорт или таблицы.')
        if not reasoning_parts:
            reasoning_parts.append('Явные требования к инструментам не обнаружены, выбран общий режим.')

        confidence_score = 0.9
        if complexity == 'medium':
            confidence_score = 0.8
        if complexity == 'complex':
            confidence_score = 0.7

        meta_analysis = {
            'llm_analysis': False,
            'reasoning': ' '.join(reasoning_parts)
        }

        if requires_search:
            expected_sources = 3
        elif requires_browser or requires_computation:
            expected_sources = 2
        else:
            expected_sources = 1

        return TaskContext(
            query=query,
            intent=intent,
            user_goal=user_goal,
            requires_search=requires_search,
            requires_browser=requires_browser,
            requires_computation=requires_computation,
            requires_excel=requires_excel,
            complexity=complexity,
            domain=domain,
            keywords=keywords,
            timestamp=CURRENT_DATE,
            urgency='normal',
            temporal_context=temporal_context,
            expected_sources=expected_sources,
            confidence_score=confidence_score,
            meta_analysis=meta_analysis,
            original_query=original_query or query
        )

    def _extract_keywords(self, text: str) -> List[str]:
        """Извлекает ключевые слова из текста."""
        words = re.findall(r'\b[а-яёa-z]{3,}\b', text.lower())
        stop_words = {
            'что', 'как', 'где', 'когда', 'почему', 'который', 'какой',
            'это', 'для', 'или', 'при', 'под', 'над', 'без', 'про'
        }
        keywords = [word for word in words if word not in stop_words]
        return list(set(keywords))[:10]

    def _detect_domain(self, query: str, keywords: List[str]) -> str:
        """Определяет предметную область запроса."""
        domains = {
            'technology': ['технология', 'компьютер', 'программирование', 'ai', 'ии'],
            'science': ['наука', 'исследование', 'эксперимент', 'теория'],
            'business': ['бизнес', 'компания', 'экономика', 'финансы'],
            'health': ['здоровье', 'медицина', 'лечение', 'болезнь'],
            'education': ['образование', 'учеба', 'университет', 'курс'],
            'finance': ['банк', 'ставка', 'процент', 'кредит', 'цб', 'рефинансирование']
        }

        for domain, domain_keywords in domains.items():
            if any(kw in query or kw in keywords for kw in domain_keywords):
                return domain

        return 'general'


class AdvancedTaskPlanner:
    """Простой планировщик задач на базе детерминированных правил."""

    def __init__(self, gigachat_client):
        self.client = gigachat_client

    def create_smart_plan(self, context: TaskContext) -> ExecutionPlan:
        """Создает понятный план выполнения задачи без обращения к LLM."""
        steps = self._build_rule_based_steps(context)
        estimated_time = max(2.0, len(steps) * 3.0)
        confidence = 0.85 if steps else 0.9

        reasoning = (
            "План сформирован на основе набора простых правил: анализируем запрос, "
            "подбираем инструменты и выстраиваем последовательность действий."
        )
        success_criteria = [
            'Каждый шаг плана выполнен без ошибок',
            'Пользователь получает итоговый ответ'
        ]
        risk_assessment = {
            'general': 'Планирование основано на четких правилах, поэтому основные риски связаны только с неверно определенной категорией задачи.'
        }

        return ExecutionPlan(
            steps=steps,
            estimated_time=estimated_time,
            confidence=confidence,
            fallback_plan=self._create_fallback_plan(context),
            reasoning=reasoning,
            risk_assessment=risk_assessment,
            success_criteria=success_criteria,
            adaptability_level='medium',
            current_step_index=0,
            completed_steps=0,
            progress_notes=[]
        )

    def _build_rule_based_steps(self, context: TaskContext) -> List[Dict[str, Any]]:
        """Формирует последовательность шагов с учетом требуемых инструментов."""
        steps: List[Dict[str, Any]] = []
        priority = 1

        def add_step(tool: str, description: str, **extra: Any) -> None:
            nonlocal priority
            step = {'tool': tool, 'description': description, 'priority': priority}
            step.update(extra)
            steps.append(step)
            priority += 1

        if context.requires_search:
            add_step('web_search', 'Найти актуальную информацию по запросу', query=context.query)
            add_step('web_parse', 'Структурировать и проанализировать найденные материалы')

        if context.requires_browser:
            add_step('browser_navigate', 'Открыть целевую страницу в браузере')
            add_step('browser_extract', 'Собрать необходимые данные со страницы')

        if context.requires_computation:
            add_step('code_execute', 'Провести вычисления или анализ данных')

        if context.requires_excel:
            add_step('excel_export', 'Подготовить выгрузку или отчет в Excel')

        return steps

    def _create_fallback_plan(self, context: TaskContext) -> List[Dict[str, Any]]:
        """Создает резервный план на случай сбоев основного сценария."""
        fallback_steps: List[Dict[str, Any]] = []
        priority = 1

        if context.requires_browser:
            fallback_steps.append({
                'tool': 'browser_navigate',
                'priority': priority,
                'description': 'Открыть нужный сайт'
            })
            priority += 1
            fallback_steps.append({
                'tool': 'browser_extract',
                'priority': priority,
                'description': 'Получить данные со страницы'
            })
            priority += 1
        else:
            fallback_steps.append({
                'tool': 'web_search',
                'priority': priority,
                'description': 'Собрать базовую информацию',
                'query': context.query
            })
            priority += 1
            fallback_steps.append({
                'tool': 'web_parse',
                'priority': priority,
                'description': 'Обработать результаты поиска'
            })
            priority += 1

        if context.requires_computation and all(step['tool'] != 'code_execute' for step in fallback_steps):
            fallback_steps.append({
                'tool': 'code_execute',
                'priority': priority,
                'description': 'Выполнить расчеты по результатам'
            })
            priority += 1

        if context.requires_excel and all(step['tool'] != 'excel_export' for step in fallback_steps):
            fallback_steps.append({
                'tool': 'excel_export',
                'priority': priority,
                'description': 'Экспортировать итоговые данные'
            })
            priority += 1

        if not fallback_steps:
            fallback_steps.append({
                'tool': 'web_search',
                'priority': 1,
                'description': 'Собрать исходную информацию',
                'query': context.query
            })

        return fallback_steps
class GigaChatClient:
    """Клиент для работы с GigaChat API."""

    def __init__(self, client_id: str, client_secret: str, verify_ssl: bool = False, 
                 model: str = "GigaChat-2-Max"):
        self.client_id = client_id
        self.client_secret = client_secret
        self.verify_ssl = verify_ssl
        self.model = model
        self.access_token = None
        self.token_expires_at = 0
        self.base_url = "https://gigachat.devices.sberbank.ru/api/v1"
        self.auth_url = "https://ngw.devices.sberbank.ru:9443/api/v2/oauth"
        self._lock = threading.Lock()
        self.session = self._create_session()

    def _create_session(self) -> requests.Session:
        """Создает requests сессию с retry стратегией."""
        session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        return session

    def _get_token(self):
        """Получает access token для GigaChat API."""
        with self._lock:
            if self.access_token and time.time() < self.token_expires_at - 60:
                return self.access_token

            credentials = base64.b64encode(f"{self.client_id}:{self.client_secret}".encode()).decode()
            headers = {
                'Authorization': f'Basic {credentials}',
                'RqUID': str(uuid.uuid4()),
                'Content-Type': 'application/x-www-form-urlencoded'
            }
            data = {'scope': 'GIGACHAT_API_CORP'}

            try:
                response = self.session.post(self.auth_url, headers=headers, data=data, 
                                           verify=self.verify_ssl, timeout=20)
                response.raise_for_status()
                token_data = response.json()
                self.access_token = token_data['access_token']
                if 'expires_at' in token_data:
                    self.token_expires_at = int(token_data['expires_at']) // 1000
                else:
                    self.token_expires_at = int(time.time() + int(token_data.get('expires_in', 1800)))
                logger.info("Получен новый GigaChat access token")
                return self.access_token
            except requests.exceptions.RequestException as e:
                logger.error(f"Ошибка получения GigaChat токена: {e}")
                raise

    def chat(self, messages: List[Dict], functions: Optional[List[Dict]] = None, 
             temperature: float = 0.3, max_tokens: int = 4096, 
             function_call: Union[str, Dict] = "auto"):
        """Отправляет запрос в GigaChat API."""
        token = self._get_token()
        url = f"{self.base_url}/chat/completions"
        headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
        data = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": max_tokens,
            "stream": False
        }
        
        if functions:
            data["functions"] = functions
            data["function_call"] = function_call

        try:
            response = self.session.post(url, headers=headers, json=data, 
                                       verify=self.verify_ssl, timeout=120)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка GigaChat запроса: {e}")
            if hasattr(e.response, 'text'):
                logger.error(f"Содержимое ответа: {e.response.text}")
            raise


class WebSearchTool:
    """Инструмент веб-поиска."""
    
    def __init__(self):
        self.available = DDGS_AVAILABLE
        if not self.available:
            logger.warning("DuckDuckGo поиск недоступен")
    
    def search(self, query: str, max_results: int = 5, region: str = 'wt-wt') -> ToolResult:
        """Выполняет веб-поиск."""
        start_time = time.time()
        
        if not self.available:
            return ToolResult(
                tool_name="web_search",
                success=False,
                data=None,
                error="DuckDuckGo search недоступен. Установите: pip install duckduckgo-search"
            )
        
        try:
            # Добавляем контекст текущей даты для временных запросов
            if any(time_word in query.lower() for time_word in ['сегодня', 'вчера', 'на этой неделе', 'текущий', 'актуальный', 'today', 'yesterday', 'current']):
                query = f"{query} {CURRENT_DATE_STR}"
                logger.info(f"Добавлена текущая дата к запросу: {CURRENT_DATE_STR}")
            
            logger.info(f"Выполняется поиск: '{query}'")
            
            with DDGS() as ddgs:
                results = list(ddgs.text(
                    query, 
                    region=region, 
                    safesearch='off', 
                    timelimit='y', 
                    max_results=max_results
                ))
            
            if not results:
                return ToolResult(
                    tool_name="web_search",
                    success=False,
                    data=None,
                    error="Результаты поиска не найдены"
                )
            
            # Форматируем результаты
            formatted_results = []
            for i, result in enumerate(results, 1):
                formatted_result = {
                    "rank": i,
                    "title": result.get("title", ""),
                    "link": result.get("href", ""),
                    "snippet": result.get("body", ""),
                    "relevance_score": 1.0 - (i * 0.1),
                    "search_date": CURRENT_DATE_STR
                }
                formatted_results.append(formatted_result)
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="web_search",
                success=True,
                data=formatted_results,
                metadata={
                    'query': query,
                    'results_count': len(formatted_results),
                    'execution_time': execution_time,
                    'source': 'DuckDuckGo',
                    'search_date': CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=0.9 if len(formatted_results) >= 3 else 0.6
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка веб-поиска: {e}")
            return ToolResult(
                tool_name="web_search",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )


class WebParsingTool:
    """Инструмент парсинга веб-страниц."""
    
    def __init__(self):
        self.available = trafilatura is not None
        if not self.available:
            logger.warning("Trafilatura недоступен")
    
    def parse(self, url: str, extract_focus: str = None) -> ToolResult:
        """Парсит содержимое веб-страницы."""
        start_time = time.time()
        
        if not self.available:
            return ToolResult(
                tool_name="web_parse",
                success=False,
                data=None,
                error="Trafilatura недоступен. Установите: pip install trafilatura"
            )
        
        try:
            logger.info(f"Парсинг URL: {url}")
            
            # Загружаем страницу
            downloaded = trafilatura.fetch_url(url)
            if not downloaded:
                return ToolResult(
                    tool_name="web_parse",
                    success=False,
                    data=None,
                    error=f"Не удалось загрузить контент с {url}"
                )
            
            # Извлекаем содержимое
            content = trafilatura.extract(
                downloaded, 
                include_links=True, 
                include_tables=True,
                include_comments=False,
                include_formatting=True,
                deduplicate=True
            )
            
            if not content:
                # Пробуем BeautifulSoup как fallback
                if BS4_AVAILABLE:
                    soup = BeautifulSoup(downloaded, 'html.parser')
                    content = soup.get_text(separator='\n', strip=True)
                else:
                    return ToolResult(
                        tool_name="web_parse",
                        success=False,
                        data=None,
                        error="Не удалось извлечь контент"
                    )
            
            # Анализируем контент
            content_analysis = {
                'word_count': len(content.split()),
                'has_tables': '|' in content or 'table' in downloaded.lower(),
                'has_lists': '•' in content or '1.' in content or '- ' in content,
                'estimated_read_time': len(content.split()) / 200,  # minutes
                'parsed_date': CURRENT_DATE_STR
            }
            
            # Если указан фокус извлечения, выделяем релевантные части
            if extract_focus and NLTK_AVAILABLE:
                try:
                    sentences = sent_tokenize(content)
                    relevant_sentences = []
                    focus_words = set(extract_focus.lower().split())
                    
                    for sentence in sentences:
                        sentence_words = set(sentence.lower().split())
                        overlap = len(focus_words & sentence_words)
                        if overlap > 0:
                            relevant_sentences.append((overlap, sentence))
                    
                    if relevant_sentences:
                        relevant_sentences.sort(key=lambda x: x[0], reverse=True)
                        top_sentences = [s[1] for s in relevant_sentences[:5]]
                        focused_content = "\n".join(top_sentences)
                        content = f"[РЕЛЕВАНТНЫЕ ФРАГМЕНТЫ]\n{focused_content}\n\n[ПОЛНЫЙ КОНТЕНТ]\n{content[:1000]}..."
                except:
                    pass  # Если не удалось, используем весь контент
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="web_parse",
                success=True,
                data=content,
                metadata={
                    'url': url,
                    'content_length': len(content),
                    'content_analysis': content_analysis,
                    'execution_time': execution_time,
                    'extract_focus': extract_focus
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка парсинга: {e}")
            return ToolResult(
                tool_name="web_parse",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )


class BrowserTool:
    """Улучшенный инструмент для работы с браузером."""
    
    def __init__(self):
        self.available = PLAYWRIGHT_AVAILABLE
        self.browser = None
        self.context = None
        self.page = None
        self.current_url = None
        
        if not self.available:
            logger.warning("Playwright недоступен")
    
    def start_session(self, headless: bool = True) -> ToolResult:
        """Запускает сессию браузера."""
        if not self.available:
            return ToolResult(
                tool_name="browser_start",
                success=False,
                data=None,
                error="Playwright недоступен. Установите: pip install playwright && playwright install chromium"
            )
        
        try:
            if self.browser:
                self.close_session()
            
            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.launch(
                headless=headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-web-security',
                    '--disable-features=VizDisplayCompositor'
                ]
            )
            
            self.context = self.browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                java_script_enabled=True,
                extra_http_headers={
                    'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8'
                }
            )
            
            self.page = self.context.new_page()
            self.page.set_default_timeout(30000)
            
            logger.info("Браузер запущен")
            return ToolResult(
                tool_name="browser_start",
                success=True,
                data="Браузер успешно запущен",
                confidence=0.9
            )
            
        except Exception as e:
            logger.error(f"Ошибка запуска браузера: {e}")
            return ToolResult(
                tool_name="browser_start",
                success=False,
                data=None,
                error=str(e)
            )
    
    def navigate(self, url: str) -> ToolResult:
        """Переходит на указанный URL с улучшенной обработкой динамического контента."""
        start_time = time.time()
        
        if not self.page:
            start_result = self.start_session()
            if not start_result.success:
                return start_result
        
        try:
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
            
            logger.info(f"Переход на: {url}")
            
            response = self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
            
            # Ждем загрузки динамического контента
            try:
                self.page.wait_for_load_state("networkidle", timeout=10000)
            except:
                # Если networkidle не сработал, ждем немного
                time.sleep(3)
            
            # Дополнительное ожидание для JavaScript
            try:
                self.page.wait_for_timeout(2000)
                
                # Пытаемся дождаться появления основного контента
                common_selectors = [
                    'main', 'article', '.content', '#content', 
                    '.main', '[role="main"]', 'body'
                ]
                
                for selector in common_selectors:
                    try:
                        self.page.wait_for_selector(selector, timeout=5000)
                        break
                    except:
                        continue
                        
            except Exception as e:
                logger.debug(f"Дополнительное ожидание не удалось: {e}")
            
            self.current_url = self.page.url
            title = self.page.title() or "Без названия"
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="browser_navigate",
                success=True,
                data=f"Успешно перешли на {url}",
                metadata={
                    "url": self.current_url,
                    "title": title,
                    "status": response.status if response else None,
                    "execution_time": execution_time,
                    "navigation_date": CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=0.9
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка навигации: {e}")
            return ToolResult(
                tool_name="browser_navigate",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )
    
    def extract_content(self, selector: str = None, wait_for_element: bool = True) -> ToolResult:
        """Извлекает контент со страницы с адаптацией под структуру сайта."""
        start_time = time.time()

        if not self.page:
            return ToolResult(
                tool_name="browser_extract",
                success=False,
                data=None,
                error="Браузер не запущен"
            )

        domain = ""
        if self.current_url:
            try:
                domain = urlparse(self.current_url).netloc.lower()
            except Exception:
                domain = ""

        extraction_metadata = {
            'requested_selector': selector,
            'wait_for_element': wait_for_element,
            'url': self.current_url,
            'domain': domain,
            'attempts': []
        }

        try:
            # Подготавливаем страницу: убираем скрытые элементы и мусор
            self._prepare_page_for_extraction()

            if selector and wait_for_element:
                # Ждем появления элемента, но не прерываем процесс при ошибке
                try:
                    self.page.wait_for_selector(selector, timeout=10000)
                except Exception as wait_error:
                    logger.debug(f"Селектор {selector} не появился за отведенное время: {wait_error}")
                    extraction_metadata['selector_wait_timeout'] = True

            content = None
            used_selector_info: Dict[str, Any] = {}

            if selector:
                selector_candidates = self._prepare_selector_candidates(selector, domain)

                for candidate in selector_candidates:
                    candidate_start = time.time()
                    extracted, attempt_info = self._extract_with_candidate(candidate)
                    attempt_info['duration'] = round(time.time() - candidate_start, 3)
                    extraction_metadata['attempts'].append(attempt_info)

                    if extracted:
                        content = extracted
                        used_selector_info = {
                            'used_selector': attempt_info.get('selector'),
                            'selector_type': attempt_info.get('type'),
                            'matched_elements': attempt_info.get('matches'),
                            'origin': attempt_info.get('origin')
                        }
                        break

            fallback_info = None
            fallback_used = False

            if not content:
                fallback_used = True
                fallback_info = self._extract_main_content(domain)
                content = fallback_info.get('text') or "Контент не найден"

            content = self._clean_extracted_text(content)

            # Собираем структурированные данные для повышения полезности результата
            structured_data = self._extract_structured_data()

            execution_time = time.time() - start_time

            if len(extraction_metadata['attempts']) > 20:
                extraction_metadata['attempts'] = extraction_metadata['attempts'][:20]

            fallback_metadata = None
            if fallback_info:
                fallback_metadata = fallback_info.copy()
                if 'text' in fallback_metadata:
                    fallback_metadata['text_preview'] = fallback_metadata['text'][:200]
                    fallback_metadata.pop('text', None)

            extraction_metadata.update({
                'used_selector': used_selector_info.get('used_selector'),
                'used_selector_type': used_selector_info.get('selector_type'),
                'used_selector_origin': used_selector_info.get('origin'),
                'matched_elements': used_selector_info.get('matched_elements'),
                'fallback_used': fallback_used,
                'fallback_details': fallback_metadata,
                'content_length': len(content) if content else 0
            })

            # Уровень уверенности зависит от того, насколько далеко пришлось отклониться от исходного селектора
            confidence = 0.8
            if used_selector_info.get('origin') and used_selector_info['origin'] not in {
                'user', 'user-prefixed', 'user-text'
            }:
                confidence = 0.7
            if fallback_used:
                confidence = 0.6 if content and content != "Контент не найден" else 0.4

            return ToolResult(
                tool_name="browser_extract",
                success=True,
                data=content,
                metadata={
                    'extraction': extraction_metadata,
                    'execution_time': execution_time,
                    'structured_data': structured_data,
                    'extraction_date': CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=confidence
            )

        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка извлечения контента: {e}")
            return ToolResult(
                tool_name="browser_extract",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )
    
    def _prepare_page_for_extraction(self) -> None:
        """Удаляет скрытые элементы и подготовляет страницу к извлечению."""
        try:
            self.page.evaluate("""
                () => {
                    const selectors = [
                        '[style*="display: none" i]',
                        '[style*="visibility: hidden" i]',
                        '[hidden]',
                        'script',
                        'style',
                        'noscript',
                        'template'
                    ];
                    const elements = document.querySelectorAll(selectors.join(','));
                    elements.forEach(el => {
                        try {
                            el.remove();
                        } catch (err) {
                            /* ignore */
                        }
                    });
                }
            """)
        except Exception as e:
            logger.debug(f"Не удалось подготовить страницу к извлечению: {e}")

    def _clean_extracted_text(self, text: Optional[str]) -> str:
        """Очищает текст от лишних пробелов и переносов."""
        if not text:
            return ""

        cleaned = text.replace('\r\n', '\n')
        cleaned = re.sub(r'\n{3,}', '\n\n', cleaned)
        cleaned = re.sub(r'[ \t]{2,}', ' ', cleaned)
        return cleaned.strip()

    def _prepare_selector_candidates(self, selector: str, domain: str) -> List[Dict[str, Any]]:
        """Создает набор кандидатов селекторов для повышения устойчивости извлечения."""
        if not selector:
            return []

        raw_candidates = [s.strip() for s in re.split(r'\s*\|\|\s*|\n+', selector) if s.strip()]
        if not raw_candidates:
            raw_candidates = [selector.strip()]

        candidates: List[Dict[str, Any]] = []
        seen: Set[Tuple[str, str]] = set()

        def add_candidate(value: str, selector_type: str, origin: str) -> None:
            if not value:
                return
            key = (selector_type, value)
            if key in seen:
                return
            seen.add(key)
            candidates.append({'value': value, 'type': selector_type, 'origin': origin})

        for raw in raw_candidates:
            normalized = raw.strip()
            if not normalized:
                continue

            lower = normalized.lower()
            if lower.startswith('css='):
                add_candidate(normalized[4:].strip() or normalized[4:], 'css', 'user-prefixed')
                continue

            if lower.startswith('xpath='):
                add_candidate(normalized[6:].strip() or normalized[6:], 'xpath', 'user-prefixed')
                continue

            if lower.startswith('text='):
                value = normalized[5:].strip()
                add_candidate(value, 'text', 'user-text')
                if re.fullmatch(r'[\w\-]+', value):
                    add_candidate(f'#{value}', 'css', 'text-as-id')
                    add_candidate(f'.{value}', 'css', 'text-as-class')
                continue

            if lower.startswith('role='):
                role_value = normalized.split('=', 1)[1].strip().strip("'\"")
                add_candidate(f'[role="{role_value}"]', 'css', 'role-attribute')
                continue

            if normalized.startswith('//') or normalized.startswith('('):
                add_candidate(normalized, 'xpath', 'user-xpath')
                continue

            add_candidate(normalized, 'css', 'user')

            if re.fullmatch(r'[\w\-]+', normalized):
                add_candidate(f'#{normalized}', 'css', 'auto-id')
                add_candidate(f'.{normalized}', 'css', 'auto-class')

            if re.fullmatch(r'[\w\s\-\.]+', normalized) and ' ' in normalized:
                add_candidate(normalized, 'text', 'user-text-heuristic')

        for domain_selector in self._get_domain_specific_selectors(domain):
            add_candidate(domain_selector, 'css', 'domain-heuristic')

        common_candidates = [
            'main', 'article', '[role="main"]', '#main', '#content',
            '.content', '.main-content', '.article', '.article-body',
            '.post', '.post-content', '.entry-content', '.news-content'
        ]
        for common_selector in common_candidates:
            add_candidate(common_selector, 'css', 'common-heuristic')

        return candidates[:20]

    def _extract_with_candidate(self, candidate: Dict[str, Any]) -> Tuple[Optional[str], Dict[str, Any]]:
        """Пробует извлечь контент с использованием конкретного кандидата."""
        attempt_info = {
            'selector': candidate.get('value'),
            'type': candidate.get('type'),
            'origin': candidate.get('origin')
        }

        try:
            selector_type = candidate.get('type')
            selector_value = candidate.get('value')

            texts: List[str] = []

            if selector_type == 'css':
                texts = self.page.eval_on_selector_all(
                    selector_value,
                    "elements => elements.map(el => (el.innerText || '').trim()).filter(Boolean)"
                )
            elif selector_type == 'xpath':
                texts = self.page.evaluate(
                    """
                    (expression) => {
                        try {
                            const iterator = document.evaluate(
                                expression,
                                document,
                                null,
                                XPathResult.ORDERED_NODE_SNAPSHOT_TYPE,
                                null
                            );
                            const results = [];
                            for (let i = 0; i < iterator.snapshotLength; i++) {
                                const node = iterator.snapshotItem(i);
                                if (!node) continue;
                                const text = (node.innerText || node.textContent || '').trim();
                                if (text) {
                                    results.push(text);
                                }
                            }
                            return results;
                        } catch (err) {
                            return [];
                        }
                    }
                    """,
                    selector_value
                )
            elif selector_type == 'text':
                texts = self.page.evaluate(
                    """
                    (text) => {
                        const needle = text ? text.trim().toLowerCase() : '';
                        if (!needle) {
                            return [];
                        }
                        const nodes = Array.from(
                            document.querySelectorAll('p, div, span, h1, h2, h3, h4, h5, h6, li, a')
                        );
                        const matches = [];
                        for (const node of nodes) {
                            const value = (node.innerText || '').trim();
                            if (!value) continue;
                            if (value.toLowerCase().includes(needle)) {
                                matches.push(value);
                            }
                        }
                        return matches;
                    }
                    """,
                    selector_value
                )
            else:
                attempt_info['error'] = f"Неизвестный тип селектора: {selector_type}"
                return None, attempt_info

            unique_texts: List[str] = []
            seen_texts: Set[str] = set()
            for item in texts or []:
                if not item:
                    continue
                normalized = item.strip()
                if not normalized:
                    continue
                if normalized in seen_texts:
                    continue
                seen_texts.add(normalized)
                unique_texts.append(normalized)

            attempt_info['matches'] = len(unique_texts)

            if not unique_texts:
                return None, attempt_info

            combined = "\n\n".join(unique_texts[:5])
            attempt_info['success'] = True
            return combined, attempt_info

        except Exception as e:
            attempt_info['error'] = str(e)[:200]
            attempt_info.setdefault('matches', 0)
            return None, attempt_info

    def _get_domain_specific_selectors(self, domain: str) -> List[str]:
        """Возвращает список селекторов, характерных для конкретных доменов."""
        if not domain:
            return []

        domain = domain.lower()
        domain_map = {
            'wikipedia.org': ['#mw-content-text', '.mw-parser-output'],
            'medium.com': ['article', '.pw-post-body-paragraph'],
            'habr.com': ['.tm-article-presenter__body', '.article-formatted-body'],
            'vc.ru': ['.content', '.article__body'],
            'ria.ru': ['.article__body', '.article__text'],
            'tass.ru': ['.news-content', '.article__text'],
            'bbc.com': ['main', '.ssrcss-uf6wea-RichTextComponentWrapper'],
            'forbes.ru': ['.article__content', '.c-article__body'],
            'rbc.ru': ['.article__text', '.js-article__content'],
            'lenta.ru': ['.topic-body__content', '.js-topic__content'],
            'theguardian.com': ['main', '.article-body-commercial-selector']
        }

        selectors: List[str] = []
        for domain_key, domain_selectors in domain_map.items():
            if domain_key in domain:
                selectors.extend(domain_selectors)

        return selectors

    def _extract_main_content(self, domain: str) -> Dict[str, Any]:
        """Извлекает основной контент страницы с учетом общих и доменных шаблонов."""
        candidate_selectors = self._get_domain_specific_selectors(domain)
        common_candidates = [
            'main', 'article', '[role="main"]', '#main', '#content',
            '.content', '.main-content', '.article', '.article-body',
            '.post', '.post-content', '.entry-content', '.news-content'
        ]

        for selector in common_candidates:
            if selector not in candidate_selectors:
                candidate_selectors.append(selector)

        fallback_result = {}

        try:
            fallback_result = self.page.evaluate(
                r"""
                (candidates) => {
                    if (!Array.isArray(candidates)) {
                        candidates = [];
                    }

                    for (const candidate of candidates) {
                        try {
                            const element = document.querySelector(candidate);
                            if (!element) continue;
                            const style = window.getComputedStyle(element);
                            if (style && (style.display === 'none' || style.visibility === 'hidden')) {
                                continue;
                            }
                            const text = (element.innerText || '').trim();
                            if (text && text.length > 160) {
                                return {
                                    text,
                                    selector: candidate,
                                    strategy: 'candidate_selector'
                                };
                            }
                        } catch (err) {
                            continue;
                        }
                    }

                    const blocks = Array.from(document.querySelectorAll('main, article, section, div'));
                    let best = null;

                    for (const element of blocks) {
                        if (!element) continue;
                        const style = window.getComputedStyle(element);
                        if (style && (style.display === 'none' || style.visibility === 'hidden')) {
                            continue;
                        }
                        const text = (element.innerText || '').trim();
                        if (!text || text.length < 200) {
                            continue;
                        }

                        if (!best || text.length > best.text.length) {
                            let descriptor = element.tagName.toLowerCase();
                            if (element.id) {
                                descriptor += '#' + element.id;
                            } else if (element.className) {
                                const className = element.className.toString().trim().split(/\s+/).filter(Boolean).slice(0, 2).join('.');
                                if (className) {
                                    descriptor += '.' + className;
                                }
                            }
                            best = {
                                text,
                                selector: descriptor,
                                strategy: 'largest_visible_block'
                            };
                        }
                    }

                    if (best) {
                        return best;
                    }

                    const bodyText = (document.body ? document.body.innerText : '').trim();
                    return {
                        text: bodyText,
                        selector: 'body',
                        strategy: 'full_body'
                    };
                }
                """,
                candidate_selectors
            ) or {}
        except Exception as e:
            logger.debug(f"Не удалось извлечь основной контент по кандидатам: {e}")

        text = fallback_result.get('text', '') if isinstance(fallback_result, dict) else ''
        selector_used = fallback_result.get('selector') if isinstance(fallback_result, dict) else None
        strategy = fallback_result.get('strategy') if isinstance(fallback_result, dict) else None

        if not text:
            try:
                text = self.page.evaluate("() => document.body ? document.body.innerText : ''")
            except Exception:
                text = self.page.text_content('body') or ''

            text = text or ''
            if not selector_used:
                selector_used = 'body'
            if not strategy:
                strategy = 'full_body'

        return {
            'text': text,
            'selector': selector_used,
            'strategy': strategy,
            'candidate_pool_size': len(candidate_selectors),
            'candidate_preview': candidate_selectors[:10]
        }

    def _extract_structured_data(self) -> Dict[str, Any]:
        """Извлекает структурированные данные со страницы."""
        try:
            # Извлекаем таблицы
            tables = self.page.evaluate("""
                () => {
                    const tables = Array.from(document.querySelectorAll('table'));
                    return tables.map(table => {
                        const rows = Array.from(table.querySelectorAll('tr'));
                        return rows.map(row => 
                            Array.from(row.querySelectorAll('td, th')).map(cell => cell.innerText.trim())
                        );
                    });
                }
            """)
            
            # Извлекаем списки
            lists = self.page.evaluate("""
                () => {
                    const lists = Array.from(document.querySelectorAll('ul, ol'));
                    return lists.map(list => 
                        Array.from(list.querySelectorAll('li')).map(item => item.innerText.trim())
                    );
                }
            """)
            
            return {
                'tables': tables,
                'lists': lists,
                'has_tables': len(tables) > 0,
                'has_lists': len(lists) > 0
            }
        except:
            return {}
    
    def wait_for_dynamic_content(self, timeout: int = 10000) -> ToolResult:
        """Ждет загрузки динамического контента."""
        try:
            # Комбинация методов ожидания
            strategies = [
                lambda: self.page.wait_for_load_state("networkidle", timeout=timeout),
                lambda: self.page.wait_for_timeout(3000),
                lambda: self.page.wait_for_function("() => document.readyState === 'complete'", timeout=timeout)
            ]
            
            for strategy in strategies:
                try:
                    strategy()
                    break
                except:
                    continue
            
            return ToolResult(
                tool_name="wait_dynamic_content",
                success=True,
                data="Динамический контент загружен",
                confidence=0.8
            )
            
        except Exception as e:
            return ToolResult(
                tool_name="wait_dynamic_content",
                success=False,
                data=None,
                error=str(e)
            )
    
    def click_element(self, selector: str) -> ToolResult:
        """Кликает по элементу."""
        start_time = time.time()
        
        if not self.page:
            return ToolResult(
                tool_name="browser_click",
                success=False,
                data=None,
                error="Браузер не запущен"
            )
        
        try:
            logger.info(f"Клик по элементу: {selector}")
            
            # Пробуем разные стратегии поиска элемента
            strategies = [
                lambda: self.page.click(selector, timeout=5000),
                lambda: self.page.click(f"#{selector}", timeout=5000),
                lambda: self.page.click(f".{selector}", timeout=5000),
                lambda: self.page.click(f"text={selector}", timeout=5000),
            ]
            
            success = False
            for strategy in strategies:
                try:
                    strategy()
                    success = True
                    break
                except:
                    continue
            
            if not success:
                return ToolResult(
                    tool_name="browser_click",
                    success=False,
                    data=None,
                    error=f"Не удалось найти элемент: {selector}"
                )
            
            # Ждем изменений на странице
            self.page.wait_for_load_state("networkidle", timeout=3000)
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="browser_click",
                success=True,
                data=f"Успешно кликнули по {selector}",
                metadata={
                    'selector': selector,
                    'execution_time': execution_time,
                    'new_url': self.page.url
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка клика: {e}")
            return ToolResult(
                tool_name="browser_click",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )
    
    def close_session(self) -> ToolResult:
        """Закрывает сессию браузера."""
        try:
            if self.page:
                self.page.close()
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if hasattr(self, 'playwright'):
                self.playwright.stop()
            
            self.page = None
            self.context = None
            self.browser = None
            self.current_url = None
            
            logger.info("Браузер закрыт")
            return ToolResult(
                tool_name="browser_close",
                success=True,
                data="Браузер закрыт",
                confidence=0.9
            )
            
        except Exception as e:
            logger.error(f"Ошибка закрытия браузера: {e}")
            return ToolResult(
                tool_name="browser_close",
                success=False,
                data=None,
                error=str(e)
            )


class ExcelExporter:
    """Класс для экспорта данных в Excel."""

    def __init__(self):
        self.available = EXCEL_AVAILABLE
    
    def export_to_excel(self, data: Any, filename: str = None, sheet_name: str = "Данные") -> ToolResult:
        """Экспортирует данные в Excel файл."""
        if not self.available:
            return ToolResult(
                tool_name="excel_export",
                success=False,
                data=None,
                error="Excel поддержка недоступна. Установите: pip install openpyxl"
            )
        
        try:
            if filename is None:
                filename = f"export_{CURRENT_DATE_STR}_{int(time.time())}.xlsx"
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            if EXCEL_ENGINE == 'openpyxl':
                return self._export_with_openpyxl(data, filename, sheet_name)
            else:
                return self._export_with_xlsxwriter(data, filename, sheet_name)
                
        except Exception as e:
            return ToolResult(
                tool_name="excel_export",
                success=False,
                data=None,
                error=str(e)
            )
    
    def _export_with_openpyxl(self, data: Any, filename: str, sheet_name: str) -> ToolResult:
        """Экспорт с использованием openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Настройка стилей
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        if isinstance(data, dict):
            # Экспорт словаря
            row = 1
            for key, value in data.items():
                ws.cell(row=row, column=1, value=str(key))
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            # Заголовки
            ws.cell(row=1, column=1, value="Ключ")
            ws.cell(row=1, column=2, value="Значение")
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2).font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            
        elif isinstance(data, list):
            # Экспорт списка
            for row_idx, item in enumerate(data, 1):
                if isinstance(item, dict):
                    # Если это список словарей
                    if row_idx == 1:
                        # Создаем заголовки
                        for col_idx, key in enumerate(item.keys(), 1):
                            cell = ws.cell(row=1, column=col_idx, value=str(key))
                            cell.font = header_font
                            cell.fill = header_fill
                        row_idx = 2
                    
                    for col_idx, value in enumerate(item.values(), 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value))
                else:
                    # Простой список
                    ws.cell(row=row_idx, column=1, value=str(item))
        
        else:
            # Простое значение
            ws.cell(row=1, column=1, value="Данные")
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=2, column=1, value=str(data))
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем информацию о создании
        info_sheet = wb.create_sheet("Информация")
        info_sheet.cell(row=1, column=1, value="Дата создания")
        info_sheet.cell(row=1, column=2, value=CURRENT_DATE_FORMATTED)
        info_sheet.cell(row=2, column=1, value="Время создания")
        info_sheet.cell(row=2, column=2, value=datetime.now().strftime("%H:%M:%S"))
        
        wb.save(filename)
        
        return ToolResult(
            tool_name="excel_export",
            success=True,
            data=f"Данные экспортированы в {filename}",
            metadata={
                'filename': filename,
                'sheet_name': sheet_name,
                'rows_exported': len(data) if isinstance(data, (list, dict)) else 1,
                'export_date': CURRENT_DATE_STR
            },
            confidence=0.9
        )
    
    def _export_with_xlsxwriter(self, data: Any, filename: str, sheet_name: str) -> ToolResult:
        """Экспорт с использованием xlsxwriter."""
        import xlsxwriter
        
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet(sheet_name)
        
        # Стили
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092'
        })
        
        if isinstance(data, dict):
            worksheet.write(0, 0, "Ключ", header_format)
            worksheet.write(0, 1, "Значение", header_format)
            
            for row, (key, value) in enumerate(data.items(), 1):
                worksheet.write(row, 0, str(key))
                worksheet.write(row, 1, str(value))
                
        elif isinstance(data, list):
            for row, item in enumerate(data):
                if isinstance(item, dict):
                    if row == 0:
                        for col, key in enumerate(item.keys()):
                            worksheet.write(0, col, str(key), header_format)
                    for col, value in enumerate(item.values()):
                        worksheet.write(row + 1, col, str(value))
                else:
                    worksheet.write(row, 0, str(item))
        else:
            worksheet.write(0, 0, "Данные", header_format)
            worksheet.write(1, 0, str(data))
        
        workbook.close()
        
        return ToolResult(
            tool_name="excel_export",
            success=True,
            data=f"Данные экспортированы в {filename}",
            metadata={
                'filename': filename,
                'sheet_name': sheet_name,
                'export_date': CURRENT_DATE_STR
            },
            confidence=0.9
        )


class CodeExecutor:
    """Улучшенный исполнитель Python кода с поддержкой Excel."""
    
    def __init__(self):
        self.globals = {
            '__builtins__': __builtins__,
            'json': json,
            'math': math,
            'random': random,
            're': re,
            'time': time,
            'datetime': datetime,
            'CURRENT_DATE': CURRENT_DATE,
            'CURRENT_DATE_STR': CURRENT_DATE_STR
        }
        
        # Добавляем numpy если доступен
        if SKLEARN_AVAILABLE:
            import numpy as np
            self.globals['np'] = np
        
        # Добавляем Excel поддержку
        if EXCEL_AVAILABLE:
            self.excel_exporter = ExcelExporter()
            self.globals['excel_export'] = self.excel_exporter.export_to_excel
    
    def execute(self, code: str) -> ToolResult:
        """Выполняет Python код с поддержкой Excel операций."""
        start_time = time.time()
        
        try:
            # Создаем изолированное пространство имен
            local_namespace = {}
            
            # Добавляем специальные функции
            local_namespace['save_to_excel'] = self._save_to_excel
            local_namespace['create_excel_report'] = self._create_excel_report
            
            # Перехватываем вывод
            import io
            from contextlib import redirect_stdout, redirect_stderr
            
            stdout_buffer = io.StringIO()
            stderr_buffer = io.StringIO()
            
            with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
                exec(code, self.globals, local_namespace)
            
            stdout = stdout_buffer.getvalue()
            stderr = stderr_buffer.getvalue()
            
            # Извлекаем результат
            result_data = local_namespace.get('result', stdout if stdout else "Код выполнен успешно")
            
            # Проверяем, были ли созданы Excel файлы
            excel_files = local_namespace.get('excel_files', [])
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="code_execute",
                success=True,
                data=result_data,
                metadata={
                    'execution_time': execution_time,
                    'code_length': len(code),
                    'output': stdout,
                    'errors': stderr,
                    'excel_files': excel_files,
                    'execution_date': CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка выполнения кода: {e}")
            
            return ToolResult(
                tool_name="code_execute",
                success=False,
                data=None,
                error=str(e),
                metadata={
                    'execution_time': execution_time,
                    'code_length': len(code)
                },
                execution_time=execution_time
            )
    
    def _save_to_excel(self, data, filename=None, sheet_name="Данные"):
        """Вспомогательная функция для сохранения в Excel."""
        if EXCEL_AVAILABLE:
            result = self.excel_exporter.export_to_excel(data, filename, sheet_name)
            return result.data if result.success else f"Ошибка: {result.error}"
        else:
            return "Excel поддержка недоступна"
    
    def _create_excel_report(self, title, data_dict, filename=None):
        """Создает отчет в Excel с заданным форматом."""
        if not EXCEL_AVAILABLE:
            return "Excel поддержка недоступна"
        
        try:
            if filename is None:
                filename = f"report_{CURRENT_DATE_STR}.xlsx"
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"
            
            # Заголовок отчета
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(size=16, bold=True)
            
            # Дата создания
            ws.cell(row=2, column=1, value=f"Дата создания: {CURRENT_DATE_FORMATTED}")
            
            # Данные
            row = 4
            for section, data in data_dict.items():
                ws.cell(row=row, column=1, value=section)
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                if isinstance(data, dict):
                    for key, value in data.items():
                        ws.cell(row=row, column=1, value=f"  {key}")
                        ws.cell(row=row, column=2, value=str(value))
                        row += 1
                elif isinstance(data, list):
                    for item in data:
                        ws.cell(row=row, column=1, value=f"  • {item}")
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=f"  {data}")
                    row += 1
                
                row += 1  # Пустая строка между секциями
            
            wb.save(filename)
            return f"Отчет сохранен в {filename}"

        except Exception as e:
            return f"Ошибка создания отчета: {e}"


class FileSystemTools:
    """Реальные инструменты работы с файловой системой для метапамяти агента."""

    def __init__(self, base_dir: Optional[Union[str, Path]] = None, encoding: str = "utf-8"):
        self.base_dir = Path(base_dir) if base_dir else Path.cwd() / "agent_workspace"
        self.base_dir.mkdir(parents=True, exist_ok=True)
        self.encoding = encoding

    def _resolve_path(self, file_path: Optional[Union[str, Path]]) -> Path:
        if file_path in (None, "", "."):
            return self.base_dir

        path = Path(file_path).expanduser()
        if not path.is_absolute():
            path = self.base_dir / path

        resolved = path.resolve()
        base_resolved = self.base_dir.resolve()
        if not resolved.is_relative_to(base_resolved):
            raise ValueError("Доступ к файлам вне рабочей директории запрещен")

        return resolved

    def _relative_path(self, path: Path) -> str:
        try:
            relative = path.relative_to(self.base_dir)
            return "." if str(relative) == "." else str(relative)
        except ValueError:
            return str(path)

    def _describe_path(self, path: Path) -> Dict[str, Any]:
        info = {
            "name": path.name,
            "path": self._relative_path(path),
            "type": "directory" if path.is_dir() else "file"
        }
        try:
            stats = path.stat()
            info["size"] = stats.st_size
            info["modified"] = datetime.fromtimestamp(stats.st_mtime).isoformat()
        except OSError:
            info["size"] = None
            info["modified"] = None
        return info

    def list_files(self, path: Optional[str] = None, recursive: bool = False,
                   include_hidden: bool = False) -> ToolResult:
        """Возвращает список файлов и директорий в рабочем пространстве."""
        try:
            target = self._resolve_path(path)

            if target.exists() and target.is_file():
                description = self._describe_path(target)
                return ToolResult(
                    tool_name="ls",
                    success=True,
                    data=[description["path"]],
                    metadata={
                        "base_directory": str(self.base_dir),
                        "entries": [description],
                        "target": description["path"],
                        "recursive": False
                    }
                )

            if not target.exists():
                return ToolResult(
                    tool_name="ls",
                    success=False,
                    data=None,
                    error="Указанный каталог не найден"
                )

            if not target.is_dir():
                return ToolResult(
                    tool_name="ls",
                    success=False,
                    data=None,
                    error="Путь не является директорией"
                )

            if recursive:
                entries_iter = target.rglob("*")
            else:
                entries_iter = target.iterdir()

            entries = []
            simple_listing = []
            for entry in sorted(entries_iter, key=lambda p: str(self._relative_path(p)).lower()):
                relative_path = self._relative_path(entry)
                if not include_hidden:
                    parts = Path(relative_path).parts if relative_path not in ("", ".") else ()
                    if any(part.startswith('.') for part in parts):
                        continue
                description = self._describe_path(entry)
                entries.append(description)
                simple_listing.append(description["path"])

            return ToolResult(
                tool_name="ls",
                success=True,
                data=simple_listing,
                metadata={
                    "base_directory": str(self.base_dir),
                    "entries": entries,
                    "target": self._relative_path(target),
                    "recursive": recursive
                }
            )

        except Exception as e:
            return ToolResult(
                tool_name="ls",
                success=False,
                data=None,
                error=str(e)
            )

    def read_file(self, file_path: str, offset: int = 0, limit: int = 2000) -> ToolResult:
        """Читает файл и возвращает его содержимое с номерами строк."""
        try:
            path = self._resolve_path(file_path)
            if not path.exists() or not path.is_file():
                return ToolResult(
                    tool_name="read_file",
                    success=False,
                    data=None,
                    error="Файл не найден"
                )

            with path.open("r", encoding=self.encoding) as f:
                lines = f.readlines()

            total_lines = len(lines)
            start = max(offset, 0)
            end = total_lines if limit is None else min(total_lines, start + max(limit, 0))

            sliced_lines = lines[start:end]
            if not sliced_lines and total_lines == 0:
                content = "Файл пуст"
            else:
                numbered = []
                for idx, line in enumerate(sliced_lines, start=start + 1):
                    stripped_line = line.rstrip("\n")
                    numbered.append(f"{idx:>6}⟶{stripped_line}")
                content = "\n".join(numbered)

            return ToolResult(
                tool_name="read_file",
                success=True,
                data=content,
                metadata={
                    "path": self._relative_path(path),
                    "total_lines": total_lines,
                    "offset": start,
                    "limit": limit,
                    "lines_returned": len(sliced_lines)
                }
            )

        except Exception as e:
            return ToolResult(
                tool_name="read_file",
                success=False,
                data=None,
                error=str(e)
            )

    def write_file(self, file_path: str, content: str, overwrite: bool = False) -> ToolResult:
        """Создает или перезаписывает файл."""
        try:
            path = self._resolve_path(file_path)
            if path.exists() and not overwrite:
                return ToolResult(
                    tool_name="write_file",
                    success=False,
                    data=None,
                    error="Файл уже существует. Укажите overwrite=True для перезаписи"
                )

            path.parent.mkdir(parents=True, exist_ok=True)
            text_content = "" if content is None else str(content)
            path.write_text(text_content, encoding=self.encoding)

            return ToolResult(
                tool_name="write_file",
                success=True,
                data=f"Файл сохранен: {self._relative_path(path)}",
                metadata={
                    "path": self._relative_path(path),
                    "bytes_written": len(text_content.encode(self.encoding)),
                    "overwrite": overwrite
                }
            )

        except Exception as e:
            return ToolResult(
                tool_name="write_file",
                success=False,
                data=None,
                error=str(e)
            )

    def edit_file(self, file_path: str, old_string: str, new_string: str,
                  replace_all: bool = False) -> ToolResult:
        """Изменяет существующий файл, заменяя указанную строку."""
        try:
            if not old_string:
                return ToolResult(
                    tool_name="edit_file",
                    success=False,
                    data=None,
                    error="Параметр old_string не может быть пустым"
                )

            path = self._resolve_path(file_path)
            if not path.exists() or not path.is_file():
                return ToolResult(
                    tool_name="edit_file",
                    success=False,
                    data=None,
                    error="Файл не найден"
                )

            text = path.read_text(encoding=self.encoding)
            occurrences = text.count(old_string)

            if occurrences == 0:
                return ToolResult(
                    tool_name="edit_file",
                    success=False,
                    data=None,
                    error="Строка для замены не найдена"
                )

            if occurrences > 1 and not replace_all:
                return ToolResult(
                    tool_name="edit_file",
                    success=False,
                    data=None,
                    error="Строка встречается несколько раз. Уточните контекст или используйте replace_all=True"
                )

            if replace_all:
                new_text = text.replace(old_string, new_string)
                replacements = occurrences
            else:
                new_text = text.replace(old_string, new_string, 1)
                replacements = 1

            path.write_text(new_text, encoding=self.encoding)

            return ToolResult(
                tool_name="edit_file",
                success=True,
                data=f"Заменено {replacements} фрагмент(ов) в {self._relative_path(path)}",
                metadata={
                    "path": self._relative_path(path),
                    "replacements": replacements,
                    "replace_all": replace_all
                }
            )

        except Exception as e:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error=str(e)
            )

    def ensure_directory(self, directory: str) -> Path:
        """Гарантирует существование директории в рабочем пространстве."""
        path = self._resolve_path(directory)
        path.mkdir(parents=True, exist_ok=True)
        return path

    def resolve_path(self, file_path: str) -> Path:
        """Публичная обертка для получения абсолютного пути внутри рабочей директории."""
        return self._resolve_path(file_path)


class MetacognitionManager:
    """Управляет метапамятью агента через файловую систему."""

    def __init__(self, fs_tools: FileSystemTools):
        self.fs_tools = fs_tools
        self.session_dir: Optional[str] = None

    def _safe_session_name(self, query: str) -> str:
        sanitized = re.sub(r"[^\wа-яА-ЯёЁ-]+", "_", query, flags=re.UNICODE)
        sanitized = re.sub(r"_+", "_", sanitized).strip("_")
        if not sanitized:
            sanitized = "task"
        return sanitized[:60]

    def start_session(self, query: str, context: TaskContext, plan: ExecutionPlan) -> Optional[str]:
        """Создает файловую структуру для хранения метаданных текущей задачи."""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            session_dir = f"metacognition/{timestamp}_{self._safe_session_name(query)}"
            self.fs_tools.ensure_directory(session_dir)
            self.session_dir = session_dir

            metadata = {
                "query": query,
                "created_at": datetime.now().isoformat(),
                "workspace": str(self.fs_tools.base_dir.resolve()),
                "session_directory": session_dir
            }

            context_result = self.fs_tools.write_file(
                f"{session_dir}/context.json",
                json.dumps(asdict(context), ensure_ascii=False, indent=2, default=str),
                overwrite=True
            )
            if not context_result.success:
                logger.warning(f"Не удалось сохранить контекст: {context_result.error}")

            plan_data = {
                "steps": plan.steps,
                "estimated_time": plan.estimated_time,
                "confidence": plan.confidence,
                "fallback_plan": plan.fallback_plan,
                "reasoning": plan.reasoning,
                "risk_assessment": plan.risk_assessment,
                "success_criteria": plan.success_criteria
            }

            plan_result = self.fs_tools.write_file(
                f"{session_dir}/plan.json",
                json.dumps(plan_data, ensure_ascii=False, indent=2, default=str),
                overwrite=True
            )
            if not plan_result.success:
                logger.warning(f"Не удалось сохранить план: {plan_result.error}")

            plan_lines = [
                "# План выполнения",
                "",
                f"Создан: {datetime.now().isoformat()}",
                "",
            ]
            for idx, step in enumerate(plan.steps, 1):
                description = step.get("description") or step.get("tool", "Шаг без описания")
                priority = step.get("priority")
                note = f" (приоритет: {priority})" if priority is not None else ""
                plan_lines.append(f"{idx}. **{step.get('tool', 'tool')}** — {description}{note}")
            if not plan.steps:
                plan_lines.append("План пуст")

            plan_md_result = self.fs_tools.write_file(
                f"{session_dir}/plan.md",
                "\n".join(plan_lines),
                overwrite=True
            )
            if not plan_md_result.success:
                logger.warning(f"Не удалось сохранить план в Markdown: {plan_md_result.error}")

            metadata_result = self.fs_tools.write_file(
                f"{session_dir}/metadata.json",
                json.dumps(metadata, ensure_ascii=False, indent=2, default=str),
                overwrite=True
            )
            if not metadata_result.success:
                logger.warning(f"Не удалось сохранить метаданные: {metadata_result.error}")

            log_init = self.fs_tools.write_file(
                f"{session_dir}/execution_log.jsonl",
                "",
                overwrite=True
            )
            if not log_init.success:
                logger.warning(f"Не удалось создать файл журнала: {log_init.error}")

            return session_dir

        except Exception as e:
            logger.warning(f"Не удалось инициализировать файловую сессию: {e}")
            self.session_dir = None
            return None

    def record_system_prompt(self, prompt: str) -> None:
        if not self.session_dir:
            return
        try:
            self.fs_tools.write_file(
                f"{self.session_dir}/system_prompt.txt",
                prompt,
                overwrite=True
            )
        except Exception as e:
            logger.warning(f"Не удалось сохранить системный промпт: {e}")

    def update_progress(self, progress: Dict[str, Any]) -> None:
        if not self.session_dir:
            return
        try:
            self.fs_tools.write_file(
                f"{self.session_dir}/progress.json",
                json.dumps(progress, ensure_ascii=False, indent=2, default=str),
                overwrite=True
            )
        except Exception as e:
            logger.warning(f"Не удалось обновить прогресс: {e}")

    def log_tool_execution(self, tool_name: str, arguments: Dict[str, Any], result: ToolResult,
                           plan_progress: Dict[str, Any], note: Optional[str] = None) -> None:
        if not self.session_dir:
            return
        try:
            entry = {
                "timestamp": datetime.now().isoformat(),
                "tool": tool_name,
                "arguments": arguments,
                "success": result.success,
                "error": result.error,
                "data_preview": str(result.data)[:500] if result.data else None,
                "metadata": result.metadata,
                "confidence": result.confidence,
                "execution_time": result.execution_time,
                "note": note
            }

            log_path = self.fs_tools.resolve_path(f"{self.session_dir}/execution_log.jsonl")
            log_path.parent.mkdir(parents=True, exist_ok=True)
            with log_path.open("a", encoding=self.fs_tools.encoding) as log_file:
                log_file.write(json.dumps(entry, ensure_ascii=False, default=str) + "\n")

            self.update_progress(plan_progress)

        except Exception as e:
            logger.warning(f"Не удалось записать шаг в журнал: {e}")

    def finalize(self, final_answer: Optional[str], summary: Dict[str, Any]) -> None:
        if not self.session_dir:
            return
        try:
            summary_payload = summary.copy()
            summary_payload["finalized_at"] = datetime.now().isoformat()
            self.fs_tools.write_file(
                f"{self.session_dir}/final_summary.json",
                json.dumps(summary_payload, ensure_ascii=False, indent=2, default=str),
                overwrite=True
            )

            if final_answer is not None:
                self.fs_tools.write_file(
                    f"{self.session_dir}/final_answer.md",
                    final_answer,
                    overwrite=True
                )

        except Exception as e:
            logger.warning(f"Не удалось сохранить финальную информацию: {e}")


class SmartAgent:
    """Умный агент для решения задач пользователя с улучшенным анализом и планированием."""

    def __init__(self, gigachat_client: GigaChatClient):
        self.client = gigachat_client
        self.query_optimizer = QueryOptimizer(gigachat_client)
        self.intent_analyzer = AdvancedIntentAnalyzer(gigachat_client)
        self.task_planner = AdvancedTaskPlanner(gigachat_client)
        self.response_evaluator = ResponseEvaluator(gigachat_client)

        # Инициализируем инструменты
        self.web_search = WebSearchTool()
        self.web_parser = WebParsingTool()
        self.browser = BrowserTool()
        self.code_executor = CodeExecutor()
        self.excel_exporter = ExcelExporter()
        self.file_system = FileSystemTools()
        self.metacognition = MetacognitionManager(self.file_system)

        # История выполнения
        self.execution_history = []
        
    def get_available_functions(self) -> List[Dict]:
        """Возвращает список доступных функций."""
        functions = []
        
        if self.web_search.available:
            functions.append({
                "name": "web_search",
                "description": "Выполняет поиск в интернете по запросу",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "Поисковый запрос"
                        },
                        "max_results": {
                            "type": "integer",
                            "description": "Максимальное количество результатов",
                            "default": 5
                        }
                    },
                    "required": ["query"]
                }
            })
        
        if self.web_parser.available:
            functions.append({
                "name": "web_parse",
                "description": "Извлекает содержимое веб-страницы",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {
                            "type": "string",
                            "description": "URL страницы для парсинга"
                        },
                        "extract_focus": {
                            "type": "string",
                            "description": "Фокус для извлечения (ключевые слова)"
                        }
                    },
                    "required": ["url"]
                }
            })
        
        if self.browser.available:
            functions.extend([
                {
                    "name": "browser_navigate",
                    "description": "Переходит на веб-страницу в браузере с поддержкой динамического контента",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "url": {
                                "type": "string",
                                "description": "URL для перехода"
                            }
                        },
                        "required": ["url"]
                    }
                },
                {
                    "name": "browser_extract",
                    "description": "Извлекает контент со страницы в браузере с ожиданием динамического контента",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "selector": {
                                "type": "string",
                                "description": "CSS селектор для извлечения (опционально)"
                            },
                            "wait_for_element": {
                                "type": "boolean",
                                "description": "Ждать появления элемента",
                                "default": True
                            }
                        }
                    }
                },
                {
                    "name": "browser_click",
                    "description": "Кликает по элементу на странице",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "selector": {
                                "type": "string",
                                "description": "Селектор элемента для клика"
                            }
                        },
                        "required": ["selector"]
                    }
                },
                {
                    "name": "wait_dynamic_content",
                    "description": "Ждет загрузки динамического контента на странице",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "timeout": {
                                "type": "integer",
                                "description": "Таймаут ожидания в миллисекундах",
                                "default": 10000
                            }
                        }
                    }
                }
            ])
        
        functions.append({
            "name": "code_execute",
            "description": "Выполняет Python код для вычислений, анализа и экспорта в Excel",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Python код для выполнения. Доступны функции save_to_excel() и create_excel_report()"
                    }
                },
                "required": ["code"]
            }
        })
        
        if self.excel_exporter.available:
            functions.append({
                "name": "excel_export",
                "description": "Экспортирует данные в Excel файл",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "data": {
                            "type": "string",
                            "description": "Данные для экспорта в JSON формате"
                        },
                        "filename": {
                            "type": "string",
                            "description": "Имя файла (опционально)"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Название листа",
                            "default": "Данные"
                        }
                    },
                    "required": ["data"]
                }
            })

        # Реальные инструменты файловой системы для метакогнитивной памяти
        functions.extend([
            {
                "name": "ls",
                "description": "Показывает содержимое рабочей директории агента",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Путь относительно рабочей директории (опционально)"
                        },
                        "recursive": {
                            "type": "boolean",
                            "description": "Показывать содержимое рекурсивно",
                            "default": False
                        },
                        "include_hidden": {
                            "type": "boolean",
                            "description": "Включать скрытые файлы",
                            "default": False
                        }
                    }
                }
            },
            {
                "name": "read_file",
                "description": "Читает файл с номерами строк из рабочей директории",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Путь к файлу относительно рабочей директории"
                        },
                        "offset": {
                            "type": "integer",
                            "description": "С какой строки начать чтение",
                            "default": 0
                        },
                        "limit": {
                            "type": "integer",
                            "description": "Сколько строк прочитать",
                            "default": 2000
                        }
                    },
                    "required": ["path"]
                }
            },
            {
                "name": "write_file",
                "description": "Создает или перезаписывает файл в рабочей директории",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Путь к файлу относительно рабочей директории"
                        },
                        "content": {
                            "type": "string",
                            "description": "Содержимое файла"
                        },
                        "overwrite": {
                            "type": "boolean",
                            "description": "Перезаписать файл, если он существует",
                            "default": False
                        }
                    },
                    "required": ["path", "content"]
                }
            },
            {
                "name": "edit_file",
                "description": "Заменяет текст в существующем файле",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {
                            "type": "string",
                            "description": "Путь к файлу относительно рабочей директории"
                        },
                        "old_string": {
                            "type": "string",
                            "description": "Исходный текст для замены"
                        },
                        "new_string": {
                            "type": "string",
                            "description": "Новый текст"
                        },
                        "replace_all": {
                            "type": "boolean",
                            "description": "Заменить все вхождения",
                            "default": False
                        }
                    },
                    "required": ["path", "old_string", "new_string"]
                }
            }
        ])

        functions.append({
            "name": "finish_task",
            "description": "Завершает выполнение задачи с финальным ответом",
            "parameters": {
                "type": "object",
                "properties": {
                    "answer": {
                        "type": "string",
                        "description": "Финальный ответ пользователю"
                    }
                },
                "required": ["answer"]
            }
        })
        
        return functions
    
    def execute_function(self, function_name: str, arguments: Dict) -> ToolResult:
        """Выполняет функцию."""
        try:
            if function_name == "web_search":
                return self.web_search.search(
                    query=arguments.get("query"),
                    max_results=arguments.get("max_results", 5)
                )
            
            elif function_name == "web_parse":
                return self.web_parser.parse(
                    url=arguments.get("url"),
                    extract_focus=arguments.get("extract_focus")
                )
            
            elif function_name == "browser_navigate":
                return self.browser.navigate(url=arguments.get("url"))
            
            elif function_name == "browser_extract":
                return self.browser.extract_content(
                    selector=arguments.get("selector"),
                    wait_for_element=arguments.get("wait_for_element", True)
                )
            
            elif function_name == "browser_click":
                return self.browser.click_element(selector=arguments.get("selector"))
            
            elif function_name == "wait_dynamic_content":
                return self.browser.wait_for_dynamic_content(
                    timeout=arguments.get("timeout", 10000)
                )
            
            elif function_name == "code_execute":
                return self.code_executor.execute(code=arguments.get("code"))
            
            elif function_name == "excel_export":
                try:
                    data = json.loads(arguments.get("data"))
                except:
                    data = arguments.get("data")

                return self.excel_exporter.export_to_excel(
                    data=data,
                    filename=arguments.get("filename"),
                    sheet_name=arguments.get("sheet_name", "Данные")
                )

            elif function_name == "ls":
                return self.file_system.list_files(
                    path=arguments.get("path"),
                    recursive=arguments.get("recursive", False),
                    include_hidden=arguments.get("include_hidden", False)
                )

            elif function_name == "read_file":
                return self.file_system.read_file(
                    file_path=arguments.get("path"),
                    offset=arguments.get("offset", 0),
                    limit=arguments.get("limit", 2000)
                )

            elif function_name == "write_file":
                return self.file_system.write_file(
                    file_path=arguments.get("path"),
                    content=arguments.get("content", ""),
                    overwrite=arguments.get("overwrite", False)
                )

            elif function_name == "edit_file":
                return self.file_system.edit_file(
                    file_path=arguments.get("path"),
                    old_string=arguments.get("old_string", ""),
                    new_string=arguments.get("new_string", ""),
                    replace_all=arguments.get("replace_all", False)
                )

            elif function_name == "finish_task":
                return ToolResult(
                    tool_name="finish_task",
                    success=True,
                    data=arguments.get("answer"),
                    confidence=1.0
                )
            
            else:
                return ToolResult(
                    tool_name=function_name,
                    success=False,
                    data=None,
                    error=f"Неизвестная функция: {function_name}"
                )

        except Exception as e:
            return ToolResult(
                tool_name=function_name,
                success=False,
                data=None,
                error=str(e)
            )

    def _initialize_plan_tracking(self, plan: ExecutionPlan) -> None:
        """Подготавливает план к отслеживанию выполнения длинных цепочек."""
        plan.current_step_index = 0
        plan.completed_steps = 0
        plan.progress_notes = []
        plan.progress = {
            'total_steps': len(plan.steps),
            'completed': 0,
            'failed': 0,
            'unplanned_calls': 0,
            'history': [],
            'current_step_order': 0
        }

        for order, step in enumerate(plan.steps):
            step['order'] = order
            step['status'] = 'pending'
            step['attempts'] = 0
            step['history'] = []
            step['last_result_summary'] = None

    def _get_next_pending_step(self, plan: ExecutionPlan) -> Optional[Dict[str, Any]]:
        """Возвращает следующий незавершенный шаг плана."""
        for step in plan.steps:
            if step.get('status') == 'pending':
                return step
        return None

    def _summarize_result(self, result: ToolResult, max_length: int = 200) -> str:
        """Создает краткую сводку по результату инструмента."""
        status = "успешно" if result.success else f"ошибка: {result.error}"

        data_preview = ""
        if result.data:
            data_str = str(result.data).strip()
            if len(data_str) > max_length:
                data_str = data_str[:max_length] + "..."
            data_preview = f" | данные: {data_str}"

        metadata_preview = ""
        if result.metadata:
            items = list(result.metadata.items())[:3]
            if items:
                metadata_pairs = ", ".join(f"{key}={value}" for key, value in items)
                metadata_preview = f" | метаданные: {metadata_pairs}"

        timing_info = ""
        if result.execution_time:
            timing_info = f" | время: {result.execution_time:.2f}с"

        confidence_info = ""
        if result.confidence:
            confidence_info = f" | уверенность: {result.confidence:.0%}"

        summary = f"{status}{data_preview}{metadata_preview}{timing_info}{confidence_info}".strip()
        return summary

    def _update_plan_progress(self, plan: ExecutionPlan, tool_name: str, result: ToolResult) -> Dict[str, Any]:
        """Обновляет прогресс плана после очередного шага."""
        if not hasattr(plan, 'progress'):
            self._initialize_plan_tracking(plan)

        progress = plan.progress
        summary = self._summarize_result(result)
        planned = False
        repeated_execution = False
        step_order = None

        for step in plan.steps:
            if step.get('tool') == tool_name and step.get('status') == 'pending':
                planned = True
                step_order = step.get('order')
                step['attempts'] = step.get('attempts', 0) + 1
                step_history = step.setdefault('history', [])
                step_history.append({
                    'success': result.success,
                    'summary': summary,
                    'timestamp': datetime.now().isoformat()
                })
                if len(step_history) > 10:
                    step['history'] = step_history[-10:]

                step['last_result_summary'] = summary

                if result.success:
                    step['status'] = 'completed'
                    progress['completed'] = progress.get('completed', 0) + 1
                    plan.completed_steps = progress['completed']
                else:
                    step['status'] = 'pending'
                    step['last_error'] = result.error
                    progress['failed'] = progress.get('failed', 0) + 1
                break
        else:
            for step in plan.steps:
                if step.get('tool') == tool_name:
                    planned = True
                    repeated_execution = True
                    step_order = step.get('order')
                    step_history = step.setdefault('history', [])
                    step_history.append({
                        'success': result.success,
                        'summary': summary,
                        'timestamp': datetime.now().isoformat(),
                        'note': 'повторное выполнение'
                    })
                    if len(step_history) > 10:
                        step['history'] = step_history[-10:]
                    break

        status_label = 'success' if result.success else 'failed'
        progress_history = progress.setdefault('history', [])
        progress_history.append({
            'tool': tool_name,
            'status': status_label if planned else 'unplanned',
            'summary': summary,
            'timestamp': datetime.now().isoformat(),
            'step_order': step_order
        })
        if len(progress_history) > 30:
            progress['history'] = progress_history[-30:]

        note = None
        if planned:
            if repeated_execution:
                note = f"Инструмент {tool_name} выполнен повторно после шага {step_order + 1 if step_order is not None else '?'}"
            elif result.success:
                note = f"Шаг {step_order + 1 if step_order is not None else '?'} ({tool_name}) выполнен успешно"
            else:
                note = f"Шаг {step_order + 1 if step_order is not None else '?'} ({tool_name}) завершился ошибкой"
        else:
            progress['unplanned_calls'] = progress.get('unplanned_calls', 0) + 1
            note = f"Выполнен внеплановый инструмент {tool_name}"

        if note:
            plan.progress_notes.append(note)
            if len(plan.progress_notes) > 10:
                plan.progress_notes = plan.progress_notes[-10:]

        next_pending = self._get_next_pending_step(plan)
        if next_pending:
            plan.current_step_index = next_pending.get('order', plan.current_step_index)
            progress['current_step_order'] = plan.current_step_index
        else:
            plan.current_step_index = len(plan.steps)
            progress['current_step_order'] = plan.current_step_index

        plan.progress = progress

        return {
            'planned': planned,
            'step_order': step_order,
            'summary': summary,
            'note': note,
            'status': status_label if planned else 'unplanned'
        }

    def _build_plan_progress_payload(self, plan: ExecutionPlan) -> Dict[str, Any]:
        """Формирует краткое описание прогресса плана."""
        if not hasattr(plan, 'progress'):
            self._initialize_plan_tracking(plan)

        progress = plan.progress
        pending_steps = []
        for step in plan.steps:
            if step.get('status') == 'pending':
                pending_steps.append({
                    'order': step.get('order'),
                    'tool': step.get('tool'),
                    'description': step.get('description'),
                    'expected_outcome': step.get('expected_outcome')
                })

        next_step = pending_steps[0] if pending_steps else None

        return {
            'total_steps': progress.get('total_steps', len(plan.steps)),
            'completed_steps': progress.get('completed', plan.completed_steps),
            'failed_attempts': progress.get('failed', 0),
            'unplanned_calls': progress.get('unplanned_calls', 0),
            'current_step_order': progress.get('current_step_order', plan.current_step_index),
            'next_step': next_step,
            'pending_steps': pending_steps[:5],
            'recent_history': progress.get('history', [])[-5:],
            'recent_notes': plan.progress_notes[-5:] if plan.progress_notes else []
        }

    def _format_plan_steps(self, steps):
        """Форматирует шаги плана в читаемый вид."""
        formatted_steps = []
        for i, step in enumerate(steps, 1):
            description = step.get('description', step['tool'])
            priority = step.get('priority', 'не указан')
            formatted_steps.append(f"{i}. {description} (приоритет: {priority})")
        return "\n".join(formatted_steps)
    
    def build_enhanced_system_prompt(self, context: TaskContext, plan: ExecutionPlan) -> str:
        """Создает улучшенный системный промпт с метаролью и я-конструкциями."""
        
        excel_info = ""
        if EXCEL_AVAILABLE:
            excel_info = """
- Я могу экспортировать данные в Excel используя excel_export или code_execute с save_to_excel()
- У меня есть функция save_to_excel(data, filename, sheet_name) в code_execute
- Я могу создать форматированный отчет через create_excel_report(title, data_dict, filename)"""

        tools_status = f"""
МОИ ДОСТУПНЫЕ ИНСТРУМЕНТЫ:
🔍 Веб-поиск: {'✅ Работает' if self.web_search.available else '❌ Недоступен'}
📄 Парсинг страниц: {'✅ Работает' if self.web_parser.available else '❌ Недоступен'}
🌐 Браузер: {'✅ Работает' if self.browser.available else '❌ Недоступен'}
💻 Выполнение кода: ✅ Работает
📊 Excel экспорт: {'✅ Работает' if EXCEL_AVAILABLE else '❌ Недоступен'}
🗂️ Файловая память: ✅ Работает (каталог: {self.file_system.base_dir})"""

        plan_reasoning = plan.reasoning if plan.reasoning else "План создан на основе базовых шаблонов"

        file_system_guidelines = f"""
МОЯ ФАЙЛОВАЯ ПАМЯТЬ:
- Рабочая директория: {self.file_system.base_dir}
- Всегда придерживаюсь последовательности ls → read_file перед write_file или edit_file
- Храню заметки и журналы задач в каталоге metacognition/ для устойчивой метапамяти"""
        
        # Исправляем строку с форматированием
        plan_steps_formatted = '\n'.join([f'{i}. {step.get("description", step["tool"])} (приоритет: {step.get("priority", "не указан")})' for i, step in enumerate(plan.steps, 1)])

        rewrite_section = ""
        optimization_meta = context.meta_analysis.get('query_optimization') if context.meta_analysis else None
        if optimization_meta and optimization_meta.get('changed'):
            original_q = optimization_meta.get('original_query', context.original_query or '')
            working_q = optimization_meta.get('optimized_query', context.query)
            details = [
                f"Исходная формулировка: \"{original_q}\"",
                f"Рабочая формулировка: \"{working_q}\"",
                f"Стратегия: {optimization_meta.get('strategy', context.rewrite_strategy)}"
            ]
            feature_summary = optimization_meta.get('feature_summary')
            if feature_summary:
                details.append(f"Наблюдения: {feature_summary}")
            rewrite_section = "МОЯ ПОДГОТОВКА ЗАПРОСА:\n- " + "\n- ".join(details)
        elif context.original_query and context.original_query != context.query:
            rewrite_section = (
                "МОЯ ПОДГОТОВКА ЗАПРОСА:\n"
                f"- Исходная формулировка: \"{context.original_query}\"\n"
                f"- Рабочая формулировка: \"{context.query}\""
            )

        return f"""Я - продвинутый интеллектуальный агент X-Master v77 Enhanced. Моя роль - эффективно решать задачи пользователей, используя доступные инструменты и глубокий анализ.

МОЯ ТЕКУЩАЯ СИТУАЦИЯ:
📅 Сегодня: {CURRENT_DATE_FORMATTED}
🎯 Задача пользователя: "{context.query}"
🧠 Мой анализ намерений: {context.intent} (уверенность: {context.confidence_score:.1%})
📋 Цель пользователя: {context.user_goal}
⚡ Сложность: {context.complexity}
🌍 Область: {context.domain}
⏰ Срочность: {context.urgency}
🕐 Временной контекст: {context.temporal_context}

{rewrite_section if rewrite_section else ''}

{tools_status}

{file_system_guidelines}

МОЙ ПЛАН ДЕЙСТВИЙ:
{plan_reasoning}

Конкретные шаги:
{plan_steps_formatted}

Критерии успеха: {', '.join(plan.success_criteria) if plan.success_criteria else 'Полный и точный ответ пользователю'}

МОИ ПРИНЦИПЫ РАБОТЫ:
1. Я всегда начинаю с глубокого анализа поставленной задачи
2. Я использую инструменты последовательно и обдуманно
3. Я проверяю результаты каждого шага перед переходом к следующему
4. Для динамических сайтов я обязательно жду загрузки контента
5. Я синтезирую информацию из разных источников для полного ответа
6. Я учитываю текущую дату при поиске актуальной информации
7. Для финансовых данных я использую браузер для работы с официальными сайтами{excel_info}
8. Я ОБЯЗАТЕЛЬНО завершаю каждую задачу вызовом finish_task с исчерпывающим ответом
9. Я отслеживаю прогресс длинных цепочек инструментов и последовательно выполняю шаги плана

МОИ МЕТАКОГНИТИВНЫЕ СПОСОБНОСТИ:
- Я анализирую свои действия и корректирую план при необходимости
- Я понимаю ограничения своих инструментов и адаптируюсь
- Я оцениваю качество получаемых данных и ищу дополнительные источники
- Я помню контекст всего диалога и использую накопленную информацию
- Я использую файловую систему для фиксации анализа, планов и результатов

МОЙ МОНИТОРИНГ ПРОГРЕССА:
- Я фиксирую статус каждого шага плана и отмечаю завершенные инструменты
- После каждого шага я напоминаю себе, какой инструмент идет следующим
- Для длинных последовательностей я регулярно сверяюсь с планом и избегаю пропусков

СТРАТЕГИЯ РАБОТЫ С ВРЕМЕННЫМИ ДАННЫМИ:
Поскольку сегодня {CURRENT_DATE_FORMATTED}, я:
- Добавляю текущую дату к поисковым запросам об актуальных событиях
- Использую браузер для получения свежих данных с официальных сайтов
- Указываю дату получения информации в своих ответах
- Проверяю актуальность найденной информации

Теперь я приступаю к выполнению задачи с полной концентрацией и профессионализмом!"""
    
    def process_query(self, query: str, max_iterations: int = 15) -> Dict[str, Any]:
        """Обрабатывает запрос пользователя с улучшенным анализом."""
        logger.info(f"Начинаю обработку запроса: {query}")
        logger.info(f"Текущая дата для контекста: {CURRENT_DATE_FORMATTED}")
        
        optimization = self.query_optimizer.optimize_query(query)
        optimized_query = optimization.optimized_query
        logger.info(
            "Оптимизация запроса завершена: стратегия=%s, изменён=%s",
            optimization.strategy.value,
            optimization.changed
        )
        if optimization.notes:
            for note in optimization.notes:
                logger.info("Оптимизация: %s", note)

        # Глубокий анализ намерений с учетом оптимизированной формулировки
        context = self.intent_analyzer.analyze_with_llm(optimized_query, original_query=query)
        context.query = optimized_query
        context.original_query = query
        context.rewrite_strategy = optimization.strategy.value
        context.rewrite_features = optimization.features.to_dict()
        if not context.meta_analysis:
            context.meta_analysis = {}
        context.meta_analysis['query_optimization'] = optimization.to_metadata()
        logger.info(f"Результат анализа намерений: {context}")
        
        # Создаем умный план выполнения
        plan = self.task_planner.create_smart_plan(context)
        logger.info(f"Создан план: {[step['tool'] for step in plan.steps]}")
        logger.info(f"Обоснование плана: {plan.reasoning}")

        # Подготавливаем план к отслеживанию длительных цепочек
        self._initialize_plan_tracking(plan)

        # Запускаем файловую метапамять
        session_dir = self.metacognition.start_session(query, context, plan)
        plan_progress_payload = self._build_plan_progress_payload(plan)
        if session_dir:
            self.metacognition.update_progress(plan_progress_payload)

        # Инициализируем диалог с улучшенным промптом
        system_prompt = self.build_enhanced_system_prompt(context, plan)
        if optimization.changed:
            user_content = (
                f"Исходный запрос пользователя: {query}\n"
                f"Уточнённая формулировка для работы: {optimized_query}"
            )
        else:
            user_content = query

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ]
        self.metacognition.record_system_prompt(system_prompt)

        execution_log = []
        final_answer = None

        iteration_limit = max(max_iterations, len(plan.steps) * 2 if plan.steps else max_iterations)
        if iteration_limit != max_iterations:
            logger.info(f"Увеличен лимит итераций до {iteration_limit} для длинной цепочки инструментов")

        for iteration in range(iteration_limit):
            try:
                logger.info(f"Итерация {iteration + 1}/{iteration_limit}")
                
                # Отправляем запрос в GigaChat
                response = self.client.chat(
                    messages=messages,
                    functions=self.get_available_functions(),
                    temperature=0.2,  # Снижаем температуру для более точного следования плану
                    function_call="auto"
                )
                
                if not response or 'choices' not in response:
                    logger.error("Некорректный ответ от GigaChat")
                    break
                
                choice = response['choices'][0]
                message = choice['message']
                
                # Добавляем сообщение в диалог
                messages.append({
                    "role": message['role'],
                    "content": message.get('content', ''),
                    "function_call": message.get('function_call')
                })
                
                # Обрабатываем вызов функции
                if 'function_call' in message:
                    func_call = message['function_call']
                    func_name = func_call.get('name')
                    
                    # Парсим аргументы
                    try:
                        if isinstance(func_call.get('arguments'), str):
                            func_args = json.loads(func_call['arguments'])
                        else:
                            func_args = func_call.get('arguments', {})
                    except json.JSONDecodeError:
                        func_args = {}
                    
                    logger.info(f"Выполняется функция: {func_name} с аргументами: {func_args}")
                    
                    # Выполняем функцию и обновляем прогресс плана
                    result = self.execute_function(func_name, func_args)
                    progress_info = self._update_plan_progress(plan, func_name, result)
                    plan_progress_payload = self._build_plan_progress_payload(plan)

                    execution_log.append({
                        'function': func_name,
                        'arguments': func_args,
                        'result': result,
                        'iteration': iteration + 1,
                        'timestamp': datetime.now().isoformat(),
                        'planned': progress_info.get('planned', False),
                        'plan_step_order': progress_info.get('step_order'),
                        'plan_status': progress_info.get('status'),
                        'plan_note': progress_info.get('note'),
                        'plan_summary': progress_info.get('summary'),
                        'plan_state': {
                            'completed_steps': plan_progress_payload.get('completed_steps'),
                            'pending_steps': len(plan_progress_payload.get('pending_steps', [])),
                            'next_tool': plan_progress_payload.get('next_step', {}).get('tool') if plan_progress_payload.get('next_step') else None
                        }
                    })

                    self.metacognition.log_tool_execution(
                        tool_name=func_name,
                        arguments=func_args,
                        result=result,
                        plan_progress=plan_progress_payload,
                        note=progress_info.get('note')
                    )

                    # Проверяем на завершение задачи
                    if func_name == "finish_task" and result.success:
                        final_answer = result.data
                        break
                    
                    # Добавляем результат в диалог с дополнительным контекстом
                    function_response = {
                        "role": "function",
                        "name": func_name,
                        "content": json.dumps({
                            "success": result.success,
                            "data": str(result.data)[:2000] if result.data else None,  # Увеличиваем лимит
                            "error": result.error,
                            "metadata": result.metadata,
                            "confidence": result.confidence,
                            "execution_time": result.execution_time,
                            "plan_progress": plan_progress_payload,
                            "plan_note": progress_info.get('note'),
                            "plan_summary": progress_info.get('summary')
                        }, ensure_ascii=False)
                    }
                    messages.append(function_response)
                
                else:
                    # Модель ответила без вызова функции
                    content = message.get('content', '')
                    if content and not final_answer:
                        # Принудительно вызываем finish_task
                        result = self.execute_function("finish_task", {"answer": content})
                        plan_progress_payload = self._build_plan_progress_payload(plan)
                        execution_log.append({
                            'function': 'finish_task',
                            'arguments': {"answer": content},
                            'result': result,
                            'iteration': iteration + 1,
                            'timestamp': datetime.now().isoformat(),
                            'planned': True,
                            'plan_status': 'completion',
                            'plan_note': 'Завершение задачи финальным ответом',
                            'plan_summary': 'finish_task',
                            'plan_state': {
                                'completed_steps': plan_progress_payload.get('completed_steps'),
                                'pending_steps': len(plan_progress_payload.get('pending_steps', [])),
                                'next_tool': None
                            }
                        })
                        self.metacognition.log_tool_execution(
                            tool_name='finish_task',
                            arguments={"answer": content},
                            result=result,
                            plan_progress=plan_progress_payload,
                            note='Завершение задачи финальным ответом'
                        )
                        plan.progress_notes.append('Задача закрыта вызовом finish_task')
                        if len(plan.progress_notes) > 10:
                            plan.progress_notes = plan.progress_notes[-10:]
                        final_answer = content
                        break
                
            except Exception as e:
                logger.error(f"Ошибка в итерации {iteration + 1}: {e}")
                break
        
        # Закрываем браузер если был открыт
        try:
            if self.browser.page:
                self.browser.close_session()
        except:
            pass
        
        # Анализируем выполнение
        successful_tools = sum(1 for log in execution_log if log['result'].success)
        total_tools = len(execution_log)
        
        # Получаем списки инструментов
        planned_tools = [step['tool'] for step in plan.steps]
        executed_tools = [log['function'] for log in execution_log]

        # Анализируем следование плану
        plan_adherence = sum(1 for log in execution_log if log.get('planned', False))
        plan_adherence_percent = (plan_adherence / max(total_tools, 1)) * 100
        plan_completion_percent = (plan.completed_steps / len(plan.steps) * 100) if plan.steps else 100.0
        plan_progress_payload = self._build_plan_progress_payload(plan)

        # Сохраняем историю выполнения
        self.execution_history = execution_log

        # Количество итераций, которые фактически были использованы
        iterations_used = (iteration + 1) if 'iteration' in locals() else 0

        # Оцениваем качество выполнения
        quality_score = 0.0
        if final_answer:
            quality_score += 0.4  # Есть финальный ответ
        if successful_tools > 0:
            quality_score += 0.3 * (successful_tools / max(total_tools, 1))  # Успешность инструментов
        if plan.steps:
            completion_ratio = min(1.0, plan.completed_steps / max(len(plan.steps), 1))
            quality_score += 0.3 * completion_ratio
        elif plan_adherence_percent > 50:
            quality_score += 0.3  # Следование плану

        final_summary_payload = {
            'success': final_answer is not None,
            'final_answer_present': final_answer is not None,
            'quality_score': quality_score,
            'tools_used': total_tools,
            'successful_tools': successful_tools,
            'plan_completion_percent': plan_completion_percent,
            'plan_adherence_percent': plan_adherence_percent,
            'iterations_used': iterations_used,
            'plan_progress': plan_progress_payload,
            'query_optimization': optimization.to_metadata()
        }

        evaluation_payload = None
        reference_answer = None
        if context.meta_analysis:
            reference_answer = context.meta_analysis.get('reference_answer') or context.meta_analysis.get('expected_answer')
        if final_answer and reference_answer:
            evaluation_payload = self.response_evaluator.evaluate(final_answer, reference_answer)
            if evaluation_payload:
                final_summary_payload['evaluation'] = evaluation_payload

        self.metacognition.finalize(final_answer, final_summary_payload)

        return {
            'success': final_answer is not None,
            'final_answer': final_answer or "Не удалось получить ответ",
            'context': context,
            'plan': plan,
            'execution_log': execution_log,
            'total_iterations': iterations_used,
            'tools_used': total_tools,
            'successful_tools': successful_tools,
            'confidence': quality_score,
            'query_date': CURRENT_DATE_STR,
            'planned_tools': planned_tools,
            'executed_tools': executed_tools,
            'plan_adherence_percent': plan_adherence_percent,
            'plan_completion_percent': plan_completion_percent,
            'plan_progress': plan_progress_payload,
            'plan_followed': plan_completion_percent >= 80 and plan_adherence_percent >= 60,
            'excel_support': EXCEL_AVAILABLE,
            'llm_analysis_used': context.meta_analysis.get('llm_analysis', False),
            'analysis_reasoning': context.meta_analysis.get('reasoning', ''),
            'plan_reasoning': plan.reasoning,
            'quality_score': quality_score,
            'risk_assessment': plan.risk_assessment,
            'iteration_limit': iteration_limit,
            'query_optimization': optimization.to_metadata(),
            'optimized_query': optimized_query,
            'evaluation': evaluation_payload
        }


def main():
    """Основное Streamlit приложение."""
    st.set_page_config(
        page_title="Умный X-Master Agent v77 Enhanced+",
        page_icon="🤖",
        layout="wide"
    )

    st.title("🤖 Умный X-Master Agent v77 Enhanced+")
    st.markdown(f"**Интеллектуальный агент с продвинутым анализом намерений и умным планированием задач**")
    st.caption(f"📅 Текущая дата: {CURRENT_DATE_FORMATTED}")

    # Боковая панель конфигурации
    with st.sidebar:
        st.header("⚙️ Настройки")
        
        # Отображаем текущую дату в сайдбаре
        st.info(f"📅 **Сегодня:** {CURRENT_DATE_FORMATTED}")

        client_id = st.text_input("GigaChat Client ID", type="password")
        client_secret = st.text_input("GigaChat Client Secret", type="password")

        st.markdown("---")

        model_name = st.selectbox(
            "Модель GigaChat",
            ["GigaChat-2-Max", "GigaChat", "GigaChat-2-Pro"],
            index=0
        )
        
        max_iterations = st.slider(
            "Максимум итераций",
            min_value=5,
            max_value=20,
            value=15,
            help="Максимальное количество шагов для решения задачи"
        )
        
        st.markdown("---")
        st.subheader("🛠️ Доступные инструменты")
        
        # Показываем статус инструментов
        tools_status = [
            ("🔍 Веб-поиск", DDGS_AVAILABLE),
            ("📄 Парсинг веб-страниц", trafilatura is not None),
            ("🌐 Браузер с динамическим контентом", PLAYWRIGHT_AVAILABLE),
            ("💻 Выполнение кода", True),
            ("📊 Экспорт в Excel", EXCEL_AVAILABLE),
            ("🧮 Научные вычисления", SKLEARN_AVAILABLE),
            ("🧠 LLM анализ намерений", True),
            ("📋 Умное планирование", True)
        ]
        
        for tool_name, available in tools_status:
            if available:
                st.success(f"✅ {tool_name}")
            else:
                st.error(f"❌ {tool_name}")
        
        st.markdown("---")
        st.info(
            "**🚀 Новые возможности v77 Enhanced+:**\n"
            "• Глубокий LLM-анализ намерений пользователя\n"
            "• Умное планирование с метарассуждениями\n"
            "• Я-конструкции в системных промптах\n"
            "• Улучшенная оценка рисков и адаптивность\n"
            "• Контекстуальное понимание временных запросов\n"
            "• Продвинутая работа с финансовыми данными\n"
            "• Автоматическое следование критериям успеха\n"
            "• Метакогнитивный самоанализ выполнения\n"
            "• Повышенная точность и надежность результатов"
        )

    # Основной интерфейс
    st.markdown("### 💬 Интерфейс запросов")
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        query = st.text_area(
            "Введите ваш запрос:",
            height=120,
            placeholder="Примеры запросов:\n• Найди последние новости об искусственном интеллекте\n• Какая ключевая ставка ЦБ РФ сегодня? Создай отчет в Excel\n• Что произошло на этой неделе в мире технологий?\n• Проанализируй текущую ситуацию на рынке криптовалют\n• Какие события запланированы на сегодня в России?\n• Сравни курсы валют и экспортируй данные",
            help="Агент автоматически проанализирует ваши намерения и создаст оптимальный план решения"
        )
    
    with col2:
        st.markdown("**🎯 Примеры задач:**")
        example_queries = [
            "Последние новости ИИ",
            "Ключевая ставка ЦБ в Excel",
            "События недели в России",
            "Анализ криптовалют",
            "Курсы валют сегодня",
            "Технологические тренды"
        ]
        
        for i, example in enumerate(example_queries):
            if st.button(f"🔍 {example}", key=f"example_{i}", use_container_width=True):
                query = example
                st.rerun()

    # Кнопка выполнения
    execute_button = st.button(
        "🚀 Выполнить задачу", 
        type="primary", 
        use_container_width=True,
        disabled=not (client_id and client_secret and query)
    )

    if execute_button:
        try:
            # Инициализируем агента
            client = GigaChatClient(client_id, client_secret, model=model_name)
            agent = SmartAgent(client)
            
            # Прогресс выполнения
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            with st.spinner("🤖 Агент выполняет глубокий анализ и умное планирование..."):
                status_text.text("🧠 Анализ намерений с помощью LLM...")
                progress_bar.progress(10)
                
                status_text.text("📋 Создание умного плана выполнения...")
                progress_bar.progress(20)
                
                # Обрабатываем запрос
                result = agent.process_query(query, max_iterations)
                progress_bar.progress(100)
                status_text.text("✅ Задача выполнена с использованием ИИ-анализа!")

            st.markdown("---")
            
            # Результаты
            st.subheader("📋 Результат")
            
            if result['success'] and result['final_answer']:
                st.success("**Ответ агента:**")
                st.write(result['final_answer'])
            else:
                st.error("Агент не смог найти ответ на задачу")

            # Расширенные метрики выполнения
            st.markdown("---")
            st.subheader("📊 Расширенные метрики выполнения")
            
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            
            with col1:
                st.metric(
                    "🔧 Инструментов",
                    result['tools_used'],
                    f"Успешно: {result['successful_tools']}"
                )
            
            with col2:
                st.metric(
                    "🔄 Итераций",
                    result['total_iterations'],
                    f"Лимит: {result.get('iteration_limit', max_iterations)}"
                )
            
            with col3:
                st.metric(
                    "🎯 Качество",
                    f"{result['quality_score']:.1%}",
                    "Общая оценка"
                )
            
            with col4:
                st.metric(
                    "📋 Следование плану",
                    f"{result.get('plan_completion_percent', 0):.0f}%",
                    f"Согласование: {result.get('plan_adherence_percent', 0):.0f}%"
                )
            
            with col5:
                llm_status = "🧠 LLM" if result.get('llm_analysis_used', False) else "📝 Базовый"
                st.metric(
                    "🔍 Анализ намерений",
                    llm_status,
                    f"Уверенность: {result['context'].confidence_score:.1%}"
                )
            
            with col6:
                st.metric(
                    "📊 Excel поддержка",
                    "✅ Работает" if result.get('excel_support', False) else "❌ Недоступен",
                    f"Движок: {EXCEL_ENGINE if EXCEL_AVAILABLE else 'Нет'}"
                )

            if result.get('evaluation'):
                st.markdown("**🔎 Оценка качества ответа:**")
                st.json(result['evaluation'])

            # Анализ намерений и планирование
            st.markdown("---")
            st.subheader("🧠 Интеллектуальный анализ и планирование")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**🎯 Анализ намерений пользователя:**")
                context = result['context']
                analysis_info = {
                    "Основное намерение": context.intent,
                    "Цель пользователя": context.user_goal,
                    "Сложность задачи": context.complexity,
                    "Предметная область": context.domain,
                    "Срочность": context.urgency,
                    "Временной контекст": context.temporal_context,
                    "Ожидаемые источники": context.expected_sources,
                    "Уверенность анализа": f"{context.confidence_score:.1%}",
                    "Использован LLM": "Да" if result.get('llm_analysis_used') else "Нет"
                }
                
                for key, value in analysis_info.items():
                    st.text(f"{key}: {value}")

                if result.get('analysis_reasoning'):
                    st.markdown("**💭 Обоснование анализа:**")
                    st.text_area("", result['analysis_reasoning'], height=100, disabled=True)

                optimization_meta = result.get('query_optimization')
                if optimization_meta:
                    st.markdown("**🛠 Оптимизация запроса:**")
                    st.text(f"Стратегия: {optimization_meta.get('strategy', 'no_rewrite')}")
                    if optimization_meta.get('changed'):
                        st.text(f"Исходный: {optimization_meta.get('original_query', '')}")
                        st.text(f"Рабочий: {optimization_meta.get('optimized_query', '')}")
                    feature_summary = optimization_meta.get('feature_summary')
                    if feature_summary:
                        st.caption(f"Признаки: {feature_summary}")

            with col2:
                st.markdown("**📋 Умное планирование задачи:**")
                plan = result['plan']
                plan_info = {
                    "Количество шагов": len(plan.steps),
                    "Оценочное время": f"{plan.estimated_time:.1f} сек",
                    "Уверенность в плане": f"{plan.confidence:.1%}",
                    "Уровень адаптивности": plan.adaptability_level
                }
                
                for key, value in plan_info.items():
                    st.text(f"{key}: {value}")
                
                if plan.reasoning:
                    st.markdown("**🎯 Логика планирования:**")
                    st.text_area("", plan.reasoning, height=100, disabled=True, key="plan_reasoning")

            # Сравнение плана и выполнения
            if result.get('planned_tools') and result.get('executed_tools'):
                st.markdown("---")
                st.subheader("📋 Выполнение плана vs. Реальность")

                plan_progress = result.get('plan_progress', {})
                if plan_progress:
                    progress_summary = f"{plan_progress.get('completed_steps', 0)}/{plan_progress.get('total_steps', 0)} шагов завершено"
                    next_tool = plan_progress.get('next_step', {}).get('tool') if plan_progress.get('next_step') else None
                    if next_tool:
                        progress_summary += f", следующий инструмент: {next_tool}"
                    st.info(f"📈 Прогресс плана: {progress_summary}")

                    if plan_progress.get('recent_notes'):
                        with st.expander("🧠 Последние заметки плана", expanded=False):
                            for note in plan_progress['recent_notes']:
                                st.write(f"- {note}")

                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown("**📝 Запланировано:**")
                    for i, tool in enumerate(result['planned_tools'], 1):
                        if tool in result['executed_tools']:
                            st.success(f"{i}. ✅ {tool}")
                        else:
                            st.error(f"{i}. ❌ {tool}")
                
                with col2:
                    st.markdown("**⚡ Выполнено:**")
                    for i, tool in enumerate(result['executed_tools'], 1):
                        log_entry = next((log for log in result['execution_log'] if log['function'] == tool), {})
                        planned_marker = "📋" if log_entry.get('planned', False) else "🔄"
                        st.success(f"{i}. {planned_marker} {tool}")
                
                with col3:
                    st.markdown("**🎯 Критерии успеха:**")
                    success_criteria = plan.success_criteria
                    if success_criteria:
                        for i, criterion in enumerate(success_criteria, 1):
                            st.info(f"{i}. {criterion}")
                    else:
                        st.info("Стандартные критерии")

            # Детали контекста
            with st.expander("🧠 Детальный анализ контекста", expanded=False):
                context = result['context']
                
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Определенные потребности:**")
                    needs = {
                        "Поиск информации": context.requires_search,
                        "Работа с браузером": context.requires_browser,
                        "Вычисления": context.requires_computation,
                        "Экспорт в Excel": context.requires_excel
                    }
                    st.json(needs)
                
                with col2:
                    st.markdown("**Извлеченные ключевые слова:**")
                    if context.keywords:
                        st.write(", ".join(context.keywords))
                    else:
                        st.write("Ключевые слова не выделены")
                
                if context.meta_analysis:
                    st.markdown("**📊 Метаанализ:**")
                    st.json(context.meta_analysis)

            # План выполнения с рисками
            with st.expander("📋 Детальный план с оценкой рисков", expanded=False):
                plan = result['plan']
                
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Запланированные шаги:**")
                    for i, step in enumerate(plan.steps, 1):
                        st.markdown(f"{i}. **{step['tool']}**")
                        if 'description' in step:
                            st.markdown(f"   📝 {step['description']}")
                        if 'priority' in step:
                            st.markdown(f"   ⚡ Приоритет: {step['priority']}")
                        if 'expected_outcome' in step:
                            st.markdown(f"   🎯 Ожидаемый результат: {step['expected_outcome']}")
                
                with col2:
                    st.markdown("**🚨 Оценка рисков:**")
                    if plan.risk_assessment:
                        st.json(plan.risk_assessment)
                    else:
                        st.info("Анализ рисков не проводился")
                
                if plan.fallback_plan:
                    st.markdown("**🔄 Резервный план:**")
                    for i, step in enumerate(plan.fallback_plan, 1):
                        st.markdown(f"{i}. {step.get('tool', 'Неизвестный инструмент')}")

            # Детальный лог выполнения
            with st.expander("📜 Детальный лог выполнения", expanded=False):
                if result['execution_log']:
                    plan_total_steps = result.get('plan_progress', {}).get('total_steps', len(result.get('planned_tools', [])))
                    for i, log_entry in enumerate(result['execution_log'], 1):
                        func_result = log_entry['result']
                        
                        # Заголовок шага
                        success_emoji = "✅" if func_result.success else "❌"
                        planned_emoji = "📋" if log_entry.get('planned', False) else "🔄"
                        st.markdown(f"### {success_emoji} {planned_emoji} Шаг {i}: {log_entry['function']}")
                        st.caption(f"⏰ Время: {log_entry.get('timestamp', 'Не указано')}")
                        
                        # Информация о выполнении
                        exec_col1, exec_col2 = st.columns(2)
                        
                        with exec_col1:
                            # Аргументы
                            if log_entry['arguments']:
                                st.markdown("**📝 Аргументы:**")
                                st.json(log_entry['arguments'])
                        
                        with exec_col2:
                            # Результат
                            if func_result.success:
                                st.success(f"✅ Успешно (уверенность: {func_result.confidence:.1%})")
                                if func_result.execution_time:
                                    st.caption(f"⏱️ Время выполнения: {func_result.execution_time:.2f}с")
                            else:
                                st.error(f"❌ Ошибка: {func_result.error}")
                        
                        # Данные результата
                        if func_result.data:
                            data_preview = str(func_result.data)[:300]
                            if len(str(func_result.data)) > 300:
                                data_preview += "..."
                            st.text_area(f"📊 Результат:", data_preview, height=100, disabled=True, key=f"result_{i}")
                        
                        # Метаданные
                        if func_result.metadata:
                            st.markdown("**🔍 Метаданные:**")
                            st.json(func_result.metadata)

                        if log_entry.get('plan_note'):
                            st.caption(f"📈 План: {log_entry['plan_note']}")

                        if log_entry.get('plan_state'):
                            completed = log_entry['plan_state'].get('completed_steps', 0)
                            pending = log_entry['plan_state'].get('pending_steps', 0)
                            next_tool = log_entry['plan_state'].get('next_tool')
                            plan_state_text = f"{completed}/{plan_total_steps} шагов выполнено, осталось {pending}"
                            if next_tool:
                                plan_state_text += f", следующий инструмент: {next_tool}"
                            st.caption(f"🧭 Прогресс плана: {plan_state_text}")

                        st.markdown("---")
                else:
                    st.info("Функции не вызывались")

        except Exception as e:
            st.error(f"⚠️ Ошибка выполнения: {str(e)}")
            logger.error("Ошибка выполнения задачи", exc_info=True)


if __name__ == "__main__":
    main()