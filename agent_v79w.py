import asyncio
import json
import re
import time
import base64
import uuid
import urllib3
import threading
import hashlib
import math
import random
import string
import logging
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import streamlit as st
from urllib.parse import urlparse, urljoin
import trafilatura
from datetime import datetime, date
from io import BytesIO
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from enum import Enum
from typing import List, Dict, Any, Optional, Tuple, Union, Callable, Set
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor

# Third-party imports
try:
    import numpy as np
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    logging.warning("scikit-learn не установлен. Семантический поиск будет недоступен.")

try:
    import nltk
    from nltk.tokenize import sent_tokenize, word_tokenize
    NLTK_AVAILABLE = True
except ImportError:
    NLTK_AVAILABLE = False
    logging.warning("NLTK не установлен. Обработка текста будет ограничена.")

try:
    from bs4 import BeautifulSoup, NavigableString, Tag
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False
    logging.warning("BeautifulSoup не установлен. Парсинг HTML будет ограничен.")

try:
    import html2text
    HTML2TEXT_AVAILABLE = True
except ImportError:
    HTML2TEXT_AVAILABLE = False
    logging.warning("html2text не установлен. Конвертация HTML в markdown недоступна.")

try:
    from ddgs import DDGS
    DDGS_AVAILABLE = True
except ImportError:
    DDGS_AVAILABLE = False
    logging.warning("duckduckgo-search не установлен. Веб-поиск будет недоступен.")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    logging.warning("Playwright не установлен. Браузер будет недоступен.")

# Excel support
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    try:
        import xlsxwriter
        EXCEL_AVAILABLE = True
        EXCEL_ENGINE = 'xlsxwriter'
    except ImportError:
        EXCEL_AVAILABLE = False
        logging.warning("Поддержка Excel недоступна. Установите: pip install openpyxl или xlsxwriter")

if EXCEL_AVAILABLE and 'openpyxl' in locals():
    EXCEL_ENGINE = 'openpyxl'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# CURRENT DATE PLACEHOLDER - Update this with the actual current date
CURRENT_DATE = datetime.now()
CURRENT_DATE_STR = CURRENT_DATE.strftime("%Y-%m-%d")
CURRENT_DATE_FORMATTED = CURRENT_DATE.strftime("%A, %B %d, %Y")

# Download NLTK data if available
if NLTK_AVAILABLE:
    try:
        nltk.data.find('tokenizers/punkt')
    except LookupError:
        try:
            nltk.download('punkt')
        except:
            NLTK_AVAILABLE = False


@dataclass
class TaskContext:
    """Контекст задачи с информацией о намерениях пользователя."""
    query: str
    intent: str
    requires_search: bool = False
    requires_browser: bool = False
    requires_computation: bool = False
    requires_excel: bool = False
    complexity: str = "simple"  # simple, medium, complex
    domain: str = "general"
    keywords: List[str] = field(default_factory=list)
    timestamp: datetime = field(default_factory=datetime.now)
    urgency: str = "normal"  # low, normal, high, critical
    expected_sources: int = 3
    temporal_context: str = "current"  # historical, current, future
    user_goal: str = "information"  # information, action, analysis, export
    confidence_score: float = 0.8
    meta_analysis: Dict[str, Any] = field(default_factory=dict)
    

@dataclass
class ExecutionPlan:
    """План выполнения задачи."""
    steps: List[Dict[str, Any]] = field(default_factory=list)
    estimated_time: float = 0.0
    confidence: float = 0.8
    fallback_plan: Optional[List[Dict[str, Any]]] = None
    reasoning: str = ""
    risk_assessment: Dict[str, str] = field(default_factory=dict)
    success_criteria: List[str] = field(default_factory=list)
    adaptability_level: str = "medium"  # low, medium, high
    

@dataclass
class ToolResult:
    """Результат выполнения инструмента."""
    tool_name: str
    success: bool
    data: Any
    metadata: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None
    execution_time: float = 0.0
    confidence: float = 0.8


@dataclass
class TodoItem:
    """Элемент списка задач для планирования."""
    content: str
    status: str = "pending"
    created_at: datetime = field(default_factory=datetime.now)
    updated_at: datetime = field(default_factory=datetime.now)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "content": self.content,
            "status": self.status,
            "created_at": self.created_at.isoformat(),
            "updated_at": self.updated_at.isoformat(),
        }


class TodoManager:
    """Управляет списком задач todo для планирования."""

    ALLOWED_STATUSES: Set[str] = {"pending", "in_progress", "completed"}

    def __init__(self):
        self.todos: List[TodoItem] = []

    def write_todos(self, todos_payload: List[Dict[str, Any]]) -> ToolResult:
        start_time = time.time()

        if not isinstance(todos_payload, list):
            return ToolResult(
                tool_name="write_todos",
                success=False,
                data=None,
                error="Аргумент 'todos' должен быть списком задач.",
            )

        if not todos_payload:
            self.todos = []
            execution_time = time.time() - start_time
            return ToolResult(
                tool_name="write_todos",
                success=True,
                data=[],
                metadata={"total_tasks": 0, "cleared": True},
                execution_time=execution_time,
                confidence=0.9,
            )

        new_todos: List[TodoItem] = []
        in_progress_present = False

        for raw_item in todos_payload:
            if not isinstance(raw_item, dict) or "content" not in raw_item:
                return ToolResult(
                    tool_name="write_todos",
                    success=False,
                    data=None,
                    error="Каждая задача должна быть объектом с полем 'content'.",
                )

            content = str(raw_item.get("content", "")).strip()
            if not content:
                return ToolResult(
                    tool_name="write_todos",
                    success=False,
                    data=None,
                    error="Описание задачи не может быть пустым.",
                )

            status = raw_item.get("status", "pending")
            if status not in self.ALLOWED_STATUSES:
                return ToolResult(
                    tool_name="write_todos",
                    success=False,
                    data=None,
                    error="Недопустимый статус задачи. Используйте pending, in_progress или completed.",
                )

            if status == "in_progress":
                in_progress_present = True

            todo_item = TodoItem(content=content, status=status)
            new_todos.append(todo_item)

        if not in_progress_present:
            # Автоматически отмечаем первую задачу как in_progress согласно рекомендациям
            new_todos[0].status = "in_progress"
            new_todos[0].updated_at = datetime.now()

        self.todos = new_todos
        execution_time = time.time() - start_time

        return ToolResult(
            tool_name="write_todos",
            success=True,
            data=[todo.to_dict() for todo in self.todos],
            metadata={
                "total_tasks": len(self.todos),
                "in_progress": sum(1 for todo in self.todos if todo.status == "in_progress"),
                "completed": sum(1 for todo in self.todos if todo.status == "completed"),
            },
            execution_time=execution_time,
            confidence=0.9,
        )

    def get_todos(self) -> List[Dict[str, Any]]:
        return [todo.to_dict() for todo in self.todos]


class VirtualFileSystem:
    """Простая виртуальная файловая система в памяти агента."""

    def __init__(self):
        self.files: Dict[str, str] = {}

    def list_files(self) -> List[str]:
        return sorted(self.files.keys())

    def read_file(self, file_path: str, offset: Union[int, str] = 0, limit: Union[int, str] = 2000) -> ToolResult:
        start_time = time.time()

        if not file_path:
            return ToolResult(
                tool_name="read_file",
                success=False,
                data=None,
                error="Необходимо указать имя файла для чтения.",
            )

        file_path = str(file_path)

        if file_path not in self.files:
            return ToolResult(
                tool_name="read_file",
                success=False,
                data=None,
                error=f"Файл '{file_path}' не найден.",
            )

        content = self.files[file_path]
        lines = content.splitlines()

        try:
            offset = int(offset)
            limit = int(limit)
        except (TypeError, ValueError):
            return ToolResult(
                tool_name="read_file",
                success=False,
                data=None,
                error="Параметры offset и limit должны быть целыми числами.",
            )

        if offset < 0:
            return ToolResult(
                tool_name="read_file",
                success=False,
                data=None,
                error="Параметр offset не может быть отрицательным.",
            )

        if offset >= len(lines) and lines:
            return ToolResult(
                tool_name="read_file",
                success=False,
                data=None,
                error="Offset превышает количество строк в файле.",
            )

        limit = max(limit, 1)
        start_index = offset
        end_index = start_index + limit
        selected_lines = lines[start_index:end_index] if lines else []

        if not selected_lines and not lines:
            formatted_content = "[Файл пуст]"
        else:
            total_lines = len(lines)
            width = max(len(str(total_lines)), 4)
            formatted_lines = [
                f"{(idx + 1):>{width}}⟶{line}"
                for idx, line in enumerate(selected_lines, start=start_index)
            ]
            formatted_content = "\n".join(formatted_lines)

        execution_time = time.time() - start_time
        return ToolResult(
            tool_name="read_file",
            success=True,
            data=formatted_content,
            metadata={
                "file_path": file_path,
                "total_lines": len(lines),
                "offset": offset,
                "limit": limit,
            },
            execution_time=execution_time,
            confidence=0.9,
        )

    def write_file(self, file_path: str, content: str) -> ToolResult:
        start_time = time.time()

        if not file_path:
            return ToolResult(
                tool_name="write_file",
                success=False,
                data=None,
                error="Имя файла не может быть пустым.",
            )

        file_path = str(file_path)

        if not isinstance(content, str):
            return ToolResult(
                tool_name="write_file",
                success=False,
                data=None,
                error="Содержимое файла должно быть строкой.",
            )

        overwritten = file_path in self.files
        self.files[file_path] = content
        execution_time = time.time() - start_time

        return ToolResult(
            tool_name="write_file",
            success=True,
            data=f"Файл '{file_path}' {'перезаписан' if overwritten else 'создан'}.",
            metadata={"file_path": file_path, "overwritten": overwritten},
            execution_time=execution_time,
            confidence=0.9,
        )

    def edit_file(
        self,
        file_path: str,
        old_string: str,
        new_string: str,
        replace_all: bool = False,
    ) -> ToolResult:
        start_time = time.time()

        if not file_path:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error="Имя файла не может быть пустым.",
            )

        file_path = str(file_path)

        if file_path not in self.files:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error=f"Файл '{file_path}' не найден.",
            )

        content = self.files[file_path]

        if old_string is None:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error="Параметр old_string не может быть None.",
            )

        old_string = str(old_string)

        if not old_string:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error="Параметр old_string не может быть пустым.",
            )

        if old_string not in content:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error="Строка для замены не найдена в файле.",
            )

        if new_string is None:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error="Параметр new_string не может быть None.",
            )

        new_string = str(new_string)

        occurrences = content.count(old_string)
        if occurrences > 1 and not replace_all:
            return ToolResult(
                tool_name="edit_file",
                success=False,
                data=None,
                error="Строка встречается несколько раз. Уточните контекст или установите replace_all=True.",
            )

        if replace_all:
            new_content = content.replace(old_string, new_string)
            replaced_count = occurrences
        else:
            new_content = content.replace(old_string, new_string, 1)
            replaced_count = 1

        self.files[file_path] = new_content
        execution_time = time.time() - start_time

        return ToolResult(
            tool_name="edit_file",
            success=True,
            data=f"Обновлен файл '{file_path}'. Заменено вхождений: {replaced_count}.",
            metadata={
                "file_path": file_path,
                "replace_all": replace_all,
                "occurrences_replaced": replaced_count,
            },
            execution_time=execution_time,
            confidence=0.85,
        )

class AdvancedIntentAnalyzer:
    """Продвинутый анализатор намерений пользователя с метарассуждениями."""
    
    def __init__(self, gigachat_client):
        self.client = gigachat_client
        self.search_indicators = [
            'найди', 'поищи', 'что такое', 'кто такой', 'где находится',
            'когда происходит', 'как работает', 'почему', 'узнай',
            'информация', 'данные', 'статистика', 'новости', 'последние',
            'сегодня', 'вчера', 'на этой неделе', 'текущий', 'актуальный'
        ]
        
        self.browser_indicators = [
            'сайт', 'страница', 'url', 'ссылка', 'перейди на', 'открой',
            'скачай', 'зарегистрируйся', 'заполни форму', 'нажми', 'интерактивно'
        ]
        
        self.computation_indicators = [
            'посчитай', 'вычисли', 'рассчитай', 'сколько', 'формула',
            'график', 'диаграмма', 'анализ', 'сравни', 'проанализируй'
        ]
        
        self.excel_indicators = [
            'excel', 'таблица', 'экспорт', 'выгрузи', 'сохрани в',
            'создай файл', 'отчет', 'xlsx', 'csv'
        ]
    
    def analyze_with_llm(self, query: str) -> TaskContext:
        """Анализирует запрос с помощью LLM для глубокого понимания намерений."""
        
        analysis_prompt = f"""Я - продвинутый аналитик задач. Мне нужно проанализировать запрос пользователя и понять его истинные намерения.

ТЕКУЩАЯ ДАТА: {CURRENT_DATE_FORMATTED}

ЗАПРОС ПОЛЬЗОВАТЕЛЯ: "{query}"

Я проведу глубокий анализ этого запроса:

1. ОСНОВНЫЕ НАМЕРЕНИЯ:
   - Что пользователь действительно хочет получить?
   - Какие скрытые потребности могут быть за этим запросом?
   - Какова конечная цель пользователя?

2. ВРЕМЕННОЙ КОНТЕКСТ:
   - Нужна ли актуальная информация на сегодняшний день?
   - Запрос касается прошлого, настоящего или будущего?
   - Есть ли временные ограничения?

3. НЕОБХОДИМЫЕ ИНСТРУМЕНТЫ:
   - Нужен ли веб-поиск для получения свежей информации?
   - Требуется ли интерактивная работа с браузером?
   - Нужны ли вычисления или анализ данных?
   - Следует ли экспортировать результат в Excel?

4. СЛОЖНОСТЬ И ОБЛАСТЬ:
   - Насколько сложна задача? (простая/средняя/сложная)
   - К какой предметной области относится?
   - Сколько этапов потребуется для решения?

5. КРИТЕРИИ УСПЕХА:
   - Как я пойму, что задача решена правильно?
   - Какой формат ответа будет наиболее полезен?

Я отвечу в формате JSON:
{{
    "intent": "основное намерение",
    "user_goal": "information/action/analysis/export",
    "requires_search": true/false,
    "requires_browser": true/false,
    "requires_computation": true/false,
    "requires_excel": true/false,
    "complexity": "simple/medium/complex",
    "domain": "область знаний",
    "urgency": "low/normal/high/critical",
    "temporal_context": "historical/current/future",
    "expected_sources": число_источников,
    "keywords": ["ключевое слово 1", "ключевое слово 2"],
    "success_criteria": ["критерий 1", "критерий 2"],
    "reasoning": "подробное объяснение анализа",
    "confidence_score": 0.0-1.0
}}"""

        try:
            messages = [
                {"role": "system", "content": "Ты - эксперт по анализу пользовательских намерений. Отвечай только в формате JSON."},
                {"role": "user", "content": analysis_prompt}
            ]
            
            response = self.client.chat(
                messages=messages,
                temperature=0.1,
                max_tokens=2048
            )
            
            if response and 'choices' in response:
                content = response['choices'][0]['message']['content']
                
                # Извлекаем JSON из ответа
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    analysis_data = json.loads(json_match.group())
                    
                    return TaskContext(
                        query=query,
                        intent=analysis_data.get('intent', 'general'),
                        user_goal=analysis_data.get('user_goal', 'information'),
                        requires_search=analysis_data.get('requires_search', False),
                        requires_browser=analysis_data.get('requires_browser', False),
                        requires_computation=analysis_data.get('requires_computation', False),
                        requires_excel=analysis_data.get('requires_excel', False),
                        complexity=analysis_data.get('complexity', 'simple'),
                        domain=analysis_data.get('domain', 'general'),
                        urgency=analysis_data.get('urgency', 'normal'),
                        temporal_context=analysis_data.get('temporal_context', 'current'),
                        expected_sources=analysis_data.get('expected_sources', 3),
                        keywords=analysis_data.get('keywords', []),
                        confidence_score=analysis_data.get('confidence_score', 0.8),
                        meta_analysis={
                            'success_criteria': analysis_data.get('success_criteria', []),
                            'reasoning': analysis_data.get('reasoning', ''),
                            'llm_analysis': True
                        },
                        timestamp=CURRENT_DATE
                    )
        
        except Exception as e:
            logger.warning(f"LLM анализ не удался, используем базовый: {e}")
        
        # Fallback на базовый анализ
        return self._basic_analysis(query)
    
    def _basic_analysis(self, query: str) -> TaskContext:
        """Базовый анализ намерений (fallback)."""
        query_lower = query.lower()
        
        # Определяем основные намерения
        requires_search = any(indicator in query_lower for indicator in self.search_indicators)
        requires_browser = any(indicator in query_lower for indicator in self.browser_indicators)
        requires_computation = any(indicator in query_lower for indicator in self.computation_indicators)
        requires_excel = any(indicator in query_lower for indicator in self.excel_indicators)
        
        # Определяем сложность
        complexity = "simple"
        if any(word in query_lower for word in ['анализ', 'сравни', 'исследуй']):
            complexity = "medium"
        if any(word in query_lower for word in ['детально', 'глубокий', 'стратегия', 'план']):
            complexity = "complex"
        
        # Определяем основное намерение
        intent = "general"
        if requires_search:
            intent = "search"
        elif requires_browser:
            intent = "web_interaction"
        elif requires_computation:
            intent = "computation"
        elif requires_excel:
            intent = "excel_export"
        
        # Извлекаем ключевые слова
        keywords = self._extract_keywords(query)
        
        # Определяем домен
        domain = self._detect_domain(query_lower, keywords)
        
        return TaskContext(
            query=query,
            intent=intent,
            requires_search=requires_search,
            requires_browser=requires_browser,
            requires_computation=requires_computation,
            requires_excel=requires_excel,
            complexity=complexity,
            domain=domain,
            keywords=keywords,
            timestamp=CURRENT_DATE,
            meta_analysis={'llm_analysis': False}
        )
    
    def _extract_keywords(self, text: str) -> List[str]:
        """Извлекает ключевые слова из текста."""
        words = re.findall(r'\b[а-яёa-z]{3,}\b', text.lower())
        stop_words = {
            'что', 'как', 'где', 'когда', 'почему', 'который', 'какой',
            'это', 'для', 'или', 'при', 'под', 'над', 'без', 'про'
        }
        keywords = [word for word in words if word not in stop_words]
        return list(set(keywords))[:10]
    
    def _detect_domain(self, query: str, keywords: List[str]) -> str:
        """Определяет предметную область запроса."""
        domains = {
            'technology': ['технология', 'компьютер', 'программирование', 'ai', 'ии'],
            'science': ['наука', 'исследование', 'эксперимент', 'теория'],
            'business': ['бизнес', 'компания', 'экономика', 'финансы'],
            'health': ['здоровье', 'медицина', 'лечение', 'болезнь'],
            'education': ['образование', 'учеба', 'университет', 'курс'],
            'finance': ['банк', 'ставка', 'процент', 'кредит', 'цб', 'рефинансирование']
        }
        
        for domain, domain_keywords in domains.items():
            if any(kw in query or kw in keywords for kw in domain_keywords):
                return domain
        
        return 'general'


class AdvancedTaskPlanner:
    """Продвинутый планировщик задач с метарассуждениями."""
    
    def __init__(self, gigachat_client):
        self.client = gigachat_client
        self.base_templates = {
            'search': [
                {'tool': 'web_search', 'priority': 1, 'description': 'Поиск информации'},
                {'tool': 'web_parse', 'priority': 2, 'description': 'Извлечение контента'},
                {'tool': 'analyze_results', 'priority': 3, 'description': 'Анализ результатов'}
            ],
            'web_interaction': [
                {'tool': 'browser_navigate', 'priority': 1, 'description': 'Переход на сайт'},
                {'tool': 'wait_dynamic_content', 'priority': 2, 'description': 'Ожидание загрузки'},
                {'tool': 'browser_extract', 'priority': 3, 'description': 'Извлечение данных'}
            ],
            'computation': [
                {'tool': 'web_search', 'priority': 1, 'description': 'Поиск данных'},
                {'tool': 'code_execute', 'priority': 2, 'description': 'Вычисления'},
                {'tool': 'analyze_results', 'priority': 3, 'description': 'Анализ результатов'}
            ],
            'excel_export': [
                {'tool': 'web_search', 'priority': 1, 'description': 'Сбор данных'},
                {'tool': 'code_execute', 'priority': 2, 'description': 'Обработка'},
                {'tool': 'excel_export', 'priority': 3, 'description': 'Экспорт в Excel'}
            ]
        }
    
    def create_smart_plan(self, context: TaskContext) -> ExecutionPlan:
        """Создает умный план выполнения с использованием LLM."""
        
        planning_prompt = f"""Я - опытный планировщик задач. Мне нужно создать оптимальный план выполнения для следующей задачи:

КОНТЕКСТ ЗАДАЧИ:
- Запрос: "{context.query}"
- Намерение: {context.intent}
- Цель пользователя: {context.user_goal}
- Сложность: {context.complexity}
- Область: {context.domain}
- Срочность: {context.urgency}
- Временной контекст: {context.temporal_context}
- Ключевые слова: {', '.join(context.keywords)}

ДОСТУПНЫЕ ИНСТРУМЕНТЫ:
1. web_search - поиск в интернете
2. web_parse - извлечение контента со страниц
3. browser_navigate - переход на сайт в браузере
4. browser_extract - извлечение данных из браузера
5. browser_click - клик по элементам
6. wait_dynamic_content - ожидание загрузки динамического контента
7. code_execute - выполнение Python кода
8. excel_export - экспорт в Excel

ТЕКУЩАЯ ДАТА: {CURRENT_DATE_FORMATTED}

Я проанализирую задачу и создам оптимальный план:

1. АНАЛИЗ ЗАДАЧИ:
   - Какие данные нужно получить?
   - Какие инструменты будут наиболее эффективны?
   - В каком порядке их лучше использовать?
   - Какие могут быть проблемы и как их избежать?

2. СТРАТЕГИЯ ВЫПОЛНЕНИЯ:
   - Начинать с самых надежных источников
   - Использовать браузер для динамических сайтов
   - Применять вычисления для анализа данных
   - Экспортировать результаты если требуется

3. ОЦЕНКА РИСКОВ:
   - Что может пойти не так?
   - Какие альтернативы подготовить?

Я отвечу в формате JSON:
{{
    "steps": [
        {{
            "tool": "название_инструмента",
            "priority": число,
            "description": "описание шага",
            "parameters": {{"param": "value"}},
            "expected_outcome": "ожидаемый результат",
            "fallback": "альтернативный инструмент"
        }}
    ],
    "reasoning": "подробное объяснение логики плана",
    "estimated_time": время_в_секундах,
    "confidence": 0.0-1.0,
    "risk_assessment": {{
        "high_risk": "описание высокого риска",
        "medium_risk": "описание среднего риска",
        "mitigation": "стратегии снижения рисков"
    }},
    "success_criteria": ["критерий 1", "критерий 2"],
    "adaptability_level": "low/medium/high"
}}"""

        try:
            messages = [
                {"role": "system", "content": "Ты - эксперт по планированию задач. Создавай эффективные планы в формате JSON."},
                {"role": "user", "content": planning_prompt}
            ]
            
            response = self.client.chat(
                messages=messages,
                temperature=0.2,
                max_tokens=2048
            )
            
            if response and 'choices' in response:
                content = response['choices'][0]['message']['content']
                
                # Извлекаем JSON из ответа
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    plan_data = json.loads(json_match.group())
                    
                    # Адаптируем шаги под контекст
                    adapted_steps = []
                    for step in plan_data.get('steps', []):
                        adapted_step = self._adapt_step_to_context(step, context)
                        adapted_steps.append(adapted_step)
                    
                    return ExecutionPlan(
                        steps=adapted_steps,
                        reasoning=plan_data.get('reasoning', ''),
                        estimated_time=plan_data.get('estimated_time', 30.0),
                        confidence=plan_data.get('confidence', 0.8),
                        risk_assessment=plan_data.get('risk_assessment', {}),
                        success_criteria=plan_data.get('success_criteria', []),
                        adaptability_level=plan_data.get('adaptability_level', 'medium'),
                        fallback_plan=self._create_fallback_plan(context)
                    )
        
        except Exception as e:
            logger.warning(f"LLM планирование не удалось, используем базовое: {e}")
        
        # Fallback на базовое планирование
        return self._create_basic_plan(context)
    
    def _adapt_step_to_context(self, step: Dict, context: TaskContext) -> Dict:
        """Адаптирует шаг под конкретный контекст задачи."""
        adapted = step.copy()
        
        # Добавляем специфичные для задачи параметры
        if step['tool'] == 'web_search':
            # Улучшаем поисковый запрос
            if context.temporal_context == 'current':
                adapted['query'] = f"{context.query} {CURRENT_DATE_STR}"
            elif any(kw in context.keywords for kw in ['ставка', 'цб', 'банк']):
                adapted['query'] = f"{context.query} ЦБ РФ сегодня"
            else:
                adapted['query'] = context.query
            
            adapted['keywords'] = context.keywords
        
        elif step['tool'] == 'browser_navigate':
            # Если нужна актуальная информация, используем специальные сайты
            if context.domain == 'finance' and any(kw in context.keywords for kw in ['ставка', 'цб']):
                adapted['url'] = 'https://www.cbr.ru/'
        
        return adapted
    
    def _create_basic_plan(self, context: TaskContext) -> ExecutionPlan:
        """Создает базовый план (fallback)."""
        steps = []
        
        # Выбираем базовый шаблон
        if context.intent in self.base_templates:
            base_steps = self.base_templates[context.intent].copy()
        else:
            base_steps = self.base_templates['search'].copy()
        
        # Адаптируем план под контекст
        for step in base_steps:
            adapted_step = self._adapt_step_to_context(step, context)
            steps.append(adapted_step)
        
        # Добавляем Excel экспорт если требуется
        if context.requires_excel:
            steps.append({
                'tool': 'excel_export',
                'priority': 10,
                'description': 'Экспорт результатов в Excel'
            })
        
        estimated_time = len(steps) * 5.0
        
        return ExecutionPlan(
            steps=steps,
            estimated_time=estimated_time,
            confidence=0.7,
            reasoning="Базовый план на основе шаблонов",
            fallback_plan=self._create_fallback_plan(context)
        )
    
    def _create_fallback_plan(self, context: TaskContext) -> List[Dict]:
        """Создает резервный план."""
        return [
            {'tool': 'web_search', 'query': context.query, 'priority': 1},
            {'tool': 'analyze_results', 'priority': 2}
        ]


class GigaChatClient:
    """Клиент для работы с GigaChat API."""

    def __init__(self, client_id: str, client_secret: str, verify_ssl: bool = False, 
                 model: str = "GigaChat-2-Max"):
        self.client_id = client_id
        self.client_secret = client_secret
        self.verify_ssl = verify_ssl
        self.model = model
        self.access_token = None
        self.token_expires_at = 0
        self.base_url = "https://gigachat.devices.sberbank.ru/api/v1"
        self.auth_url = "https://ngw.devices.sberbank.ru:9443/api/v2/oauth"
        self._lock = threading.Lock()
        self.session = self._create_session()

    def _create_session(self) -> requests.Session:
        """Создает requests сессию с retry стратегией."""
        session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        return session

    def _get_token(self):
        """Получает access token для GigaChat API."""
        with self._lock:
            if self.access_token and time.time() < self.token_expires_at - 60:
                return self.access_token

            credentials = base64.b64encode(f"{self.client_id}:{self.client_secret}".encode()).decode()
            headers = {
                'Authorization': f'Basic {credentials}',
                'RqUID': str(uuid.uuid4()),
                'Content-Type': 'application/x-www-form-urlencoded'
            }
            data = {'scope': 'GIGACHAT_API_CORP'}

            try:
                response = self.session.post(self.auth_url, headers=headers, data=data, 
                                           verify=self.verify_ssl, timeout=20)
                response.raise_for_status()
                token_data = response.json()
                self.access_token = token_data['access_token']
                if 'expires_at' in token_data:
                    self.token_expires_at = int(token_data['expires_at']) // 1000
                else:
                    self.token_expires_at = int(time.time() + int(token_data.get('expires_in', 1800)))
                logger.info("Получен новый GigaChat access token")
                return self.access_token
            except requests.exceptions.RequestException as e:
                logger.error(f"Ошибка получения GigaChat токена: {e}")
                raise

    def chat(self, messages: List[Dict], functions: Optional[List[Dict]] = None, 
             temperature: float = 0.3, max_tokens: int = 4096, 
             function_call: Union[str, Dict] = "auto"):
        """Отправляет запрос в GigaChat API."""
        token = self._get_token()
        url = f"{self.base_url}/chat/completions"
        headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
        data = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": max_tokens,
            "stream": False
        }
        
        if functions:
            data["functions"] = functions
            data["function_call"] = function_call

        try:
            response = self.session.post(url, headers=headers, json=data, 
                                       verify=self.verify_ssl, timeout=120)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка GigaChat запроса: {e}")
            if hasattr(e.response, 'text'):
                logger.error(f"Содержимое ответа: {e.response.text}")
            raise


class WebSearchTool:
    """Инструмент веб-поиска."""
    
    def __init__(self):
        self.available = DDGS_AVAILABLE
        if not self.available:
            logger.warning("DuckDuckGo поиск недоступен")
    
    def search(self, query: str, max_results: int = 5, region: str = 'wt-wt') -> ToolResult:
        """Выполняет веб-поиск."""
        start_time = time.time()
        
        if not self.available:
            return ToolResult(
                tool_name="web_search",
                success=False,
                data=None,
                error="DuckDuckGo search недоступен. Установите: pip install duckduckgo-search"
            )
        
        try:
            # Добавляем контекст текущей даты для временных запросов
            if any(time_word in query.lower() for time_word in ['сегодня', 'вчера', 'на этой неделе', 'текущий', 'актуальный', 'today', 'yesterday', 'current']):
                query = f"{query} {CURRENT_DATE_STR}"
                logger.info(f"Добавлена текущая дата к запросу: {CURRENT_DATE_STR}")
            
            logger.info(f"Выполняется поиск: '{query}'")
            
            with DDGS() as ddgs:
                results = list(ddgs.text(
                    query, 
                    region=region, 
                    safesearch='off', 
                    timelimit='y', 
                    max_results=max_results
                ))
            
            if not results:
                return ToolResult(
                    tool_name="web_search",
                    success=False,
                    data=None,
                    error="Результаты поиска не найдены"
                )
            
            # Форматируем результаты
            formatted_results = []
            for i, result in enumerate(results, 1):
                formatted_result = {
                    "rank": i,
                    "title": result.get("title", ""),
                    "link": result.get("href", ""),
                    "snippet": result.get("body", ""),
                    "relevance_score": 1.0 - (i * 0.1),
                    "search_date": CURRENT_DATE_STR
                }
                formatted_results.append(formatted_result)
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="web_search",
                success=True,
                data=formatted_results,
                metadata={
                    'query': query,
                    'results_count': len(formatted_results),
                    'execution_time': execution_time,
                    'source': 'DuckDuckGo',
                    'search_date': CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=0.9 if len(formatted_results) >= 3 else 0.6
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка веб-поиска: {e}")
            return ToolResult(
                tool_name="web_search",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )


class WebParsingTool:
    """Инструмент парсинга веб-страниц."""
    
    def __init__(self):
        self.available = trafilatura is not None
        if not self.available:
            logger.warning("Trafilatura недоступен")
    
    def parse(self, url: str, extract_focus: str = None) -> ToolResult:
        """Парсит содержимое веб-страницы."""
        start_time = time.time()
        
        if not self.available:
            return ToolResult(
                tool_name="web_parse",
                success=False,
                data=None,
                error="Trafilatura недоступен. Установите: pip install trafilatura"
            )
        
        try:
            logger.info(f"Парсинг URL: {url}")
            
            # Загружаем страницу
            downloaded = trafilatura.fetch_url(url)
            if not downloaded:
                return ToolResult(
                    tool_name="web_parse",
                    success=False,
                    data=None,
                    error=f"Не удалось загрузить контент с {url}"
                )
            
            # Извлекаем содержимое
            content = trafilatura.extract(
                downloaded, 
                include_links=True, 
                include_tables=True,
                include_comments=False,
                include_formatting=True,
                deduplicate=True
            )
            
            if not content:
                # Пробуем BeautifulSoup как fallback
                if BS4_AVAILABLE:
                    soup = BeautifulSoup(downloaded, 'html.parser')
                    content = soup.get_text(separator='\n', strip=True)
                else:
                    return ToolResult(
                        tool_name="web_parse",
                        success=False,
                        data=None,
                        error="Не удалось извлечь контент"
                    )
            
            # Анализируем контент
            content_analysis = {
                'word_count': len(content.split()),
                'has_tables': '|' in content or 'table' in downloaded.lower(),
                'has_lists': '•' in content or '1.' in content or '- ' in content,
                'estimated_read_time': len(content.split()) / 200,  # minutes
                'parsed_date': CURRENT_DATE_STR
            }
            
            # Если указан фокус извлечения, выделяем релевантные части
            if extract_focus and NLTK_AVAILABLE:
                try:
                    sentences = sent_tokenize(content)
                    relevant_sentences = []
                    focus_words = set(extract_focus.lower().split())
                    
                    for sentence in sentences:
                        sentence_words = set(sentence.lower().split())
                        overlap = len(focus_words & sentence_words)
                        if overlap > 0:
                            relevant_sentences.append((overlap, sentence))
                    
                    if relevant_sentences:
                        relevant_sentences.sort(key=lambda x: x[0], reverse=True)
                        top_sentences = [s[1] for s in relevant_sentences[:5]]
                        focused_content = "\n".join(top_sentences)
                        content = f"[РЕЛЕВАНТНЫЕ ФРАГМЕНТЫ]\n{focused_content}\n\n[ПОЛНЫЙ КОНТЕНТ]\n{content[:1000]}..."
                except:
                    pass  # Если не удалось, используем весь контент
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="web_parse",
                success=True,
                data=content,
                metadata={
                    'url': url,
                    'content_length': len(content),
                    'content_analysis': content_analysis,
                    'execution_time': execution_time,
                    'extract_focus': extract_focus
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка парсинга: {e}")
            return ToolResult(
                tool_name="web_parse",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )


class BrowserTool:
    """Улучшенный инструмент для работы с браузером."""
    
    def __init__(self):
        self.available = PLAYWRIGHT_AVAILABLE
        self.browser = None
        self.context = None
        self.page = None
        self.current_url = None
        
        if not self.available:
            logger.warning("Playwright недоступен")
    
    def start_session(self, headless: bool = True) -> ToolResult:
        """Запускает сессию браузера."""
        if not self.available:
            return ToolResult(
                tool_name="browser_start",
                success=False,
                data=None,
                error="Playwright недоступен. Установите: pip install playwright && playwright install chromium"
            )
        
        try:
            if self.browser:
                self.close_session()
            
            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.launch(
                headless=headless,
                args=[
                    '--disable-blink-features=AutomationControlled',
                    '--no-sandbox',
                    '--disable-dev-shm-usage',
                    '--disable-web-security',
                    '--disable-features=VizDisplayCompositor'
                ]
            )
            
            self.context = self.browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                java_script_enabled=True,
                extra_http_headers={
                    'Accept-Language': 'ru-RU,ru;q=0.9,en;q=0.8'
                }
            )
            
            self.page = self.context.new_page()
            self.page.set_default_timeout(30000)
            
            logger.info("Браузер запущен")
            return ToolResult(
                tool_name="browser_start",
                success=True,
                data="Браузер успешно запущен",
                confidence=0.9
            )
            
        except Exception as e:
            logger.error(f"Ошибка запуска браузера: {e}")
            return ToolResult(
                tool_name="browser_start",
                success=False,
                data=None,
                error=str(e)
            )
    
    def navigate(self, url: str) -> ToolResult:
        """Переходит на указанный URL с улучшенной обработкой динамического контента."""
        start_time = time.time()
        
        if not self.page:
            start_result = self.start_session()
            if not start_result.success:
                return start_result
        
        try:
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
            
            logger.info(f"Переход на: {url}")
            
            response = self.page.goto(url, wait_until="domcontentloaded", timeout=30000)
            
            # Ждем загрузки динамического контента
            try:
                self.page.wait_for_load_state("networkidle", timeout=10000)
            except:
                # Если networkidle не сработал, ждем немного
                time.sleep(3)
            
            # Дополнительное ожидание для JavaScript
            try:
                self.page.wait_for_timeout(2000)
                
                # Пытаемся дождаться появления основного контента
                common_selectors = [
                    'main', 'article', '.content', '#content', 
                    '.main', '[role="main"]', 'body'
                ]
                
                for selector in common_selectors:
                    try:
                        self.page.wait_for_selector(selector, timeout=5000)
                        break
                    except:
                        continue
                        
            except Exception as e:
                logger.debug(f"Дополнительное ожидание не удалось: {e}")
            
            self.current_url = self.page.url
            title = self.page.title() or "Без названия"
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="browser_navigate",
                success=True,
                data=f"Успешно перешли на {url}",
                metadata={
                    "url": self.current_url,
                    "title": title,
                    "status": response.status if response else None,
                    "execution_time": execution_time,
                    "navigation_date": CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=0.9
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка навигации: {e}")
            return ToolResult(
                tool_name="browser_navigate",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )
    
    def extract_content(self, selector: str = None, wait_for_element: bool = True) -> ToolResult:
        """Извлекает контент со страницы с улучшенной обработкой динамического контента."""
        start_time = time.time()
        
        if not self.page:
            return ToolResult(
                tool_name="browser_extract",
                success=False,
                data=None,
                error="Браузер не запущен"
            )
        
        try:
            # Если указан селектор, ждем его появления
            if selector and wait_for_element:
                try:
                    self.page.wait_for_selector(selector, timeout=10000)
                except:
                    logger.warning(f"Селектор {selector} не найден в течение 10 секунд")
            
            if selector:
                # Извлекаем по селектору
                elements = self.page.query_selector_all(selector)
                if elements:
                    content = "\n".join([elem.text_content() for elem in elements if elem.text_content()])
                else:
                    content = f"Элементы с селектором '{selector}' не найдены или пусты"
            else:
                # Извлекаем весь видимый текст с улучшенной обработкой
                try:
                    # Удаляем скрытые элементы перед извлечением
                    self.page.evaluate("""
                        () => {
                            const hiddenElements = document.querySelectorAll('[style*="display: none"], [style*="visibility: hidden"], .hidden');
                            hiddenElements.forEach(el => el.remove());
                        }
                    """)
                    
                    content = self.page.evaluate("() => document.body.innerText")
                    
                    # Очищаем контент от лишних пробелов и переносов
                    content = re.sub(r'\n\s*\n', '\n', content)
                    content = re.sub(r' +', ' ', content)
                    
                except Exception as e:
                    logger.warning(f"Не удалось извлечь контент через JavaScript: {e}")
                    content = self.page.text_content('body') or "Контент не найден"
            
            # Дополнительно пытаемся извлечь структурированные данные
            structured_data = self._extract_structured_data()
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="browser_extract",
                success=True,
                data=content,
                metadata={
                    'selector': selector,
                    'content_length': len(content),
                    'execution_time': execution_time,
                    'url': self.current_url,
                    'extraction_date': CURRENT_DATE_STR,
                    'structured_data': structured_data
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка извлечения контента: {e}")
            return ToolResult(
                tool_name="browser_extract",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )
    
    def _extract_structured_data(self) -> Dict[str, Any]:
        """Извлекает структурированные данные со страницы."""
        try:
            # Извлекаем таблицы
            tables = self.page.evaluate("""
                () => {
                    const tables = Array.from(document.querySelectorAll('table'));
                    return tables.map(table => {
                        const rows = Array.from(table.querySelectorAll('tr'));
                        return rows.map(row => 
                            Array.from(row.querySelectorAll('td, th')).map(cell => cell.innerText.trim())
                        );
                    });
                }
            """)
            
            # Извлекаем списки
            lists = self.page.evaluate("""
                () => {
                    const lists = Array.from(document.querySelectorAll('ul, ol'));
                    return lists.map(list => 
                        Array.from(list.querySelectorAll('li')).map(item => item.innerText.trim())
                    );
                }
            """)
            
            return {
                'tables': tables,
                'lists': lists,
                'has_tables': len(tables) > 0,
                'has_lists': len(lists) > 0
            }
        except:
            return {}
    
    def wait_for_dynamic_content(self, timeout: int = 10000) -> ToolResult:
        """Ждет загрузки динамического контента."""
        try:
            # Комбинация методов ожидания
            strategies = [
                lambda: self.page.wait_for_load_state("networkidle", timeout=timeout),
                lambda: self.page.wait_for_timeout(3000),
                lambda: self.page.wait_for_function("() => document.readyState === 'complete'", timeout=timeout)
            ]
            
            for strategy in strategies:
                try:
                    strategy()
                    break
                except:
                    continue
            
            return ToolResult(
                tool_name="wait_dynamic_content",
                success=True,
                data="Динамический контент загружен",
                confidence=0.8
            )
            
        except Exception as e:
            return ToolResult(
                tool_name="wait_dynamic_content",
                success=False,
                data=None,
                error=str(e)
            )
    
    def click_element(self, selector: str) -> ToolResult:
        """Кликает по элементу."""
        start_time = time.time()
        
        if not self.page:
            return ToolResult(
                tool_name="browser_click",
                success=False,
                data=None,
                error="Браузер не запущен"
            )
        
        try:
            logger.info(f"Клик по элементу: {selector}")
            
            # Пробуем разные стратегии поиска элемента
            strategies = [
                lambda: self.page.click(selector, timeout=5000),
                lambda: self.page.click(f"#{selector}", timeout=5000),
                lambda: self.page.click(f".{selector}", timeout=5000),
                lambda: self.page.click(f"text={selector}", timeout=5000),
            ]
            
            success = False
            for strategy in strategies:
                try:
                    strategy()
                    success = True
                    break
                except:
                    continue
            
            if not success:
                return ToolResult(
                    tool_name="browser_click",
                    success=False,
                    data=None,
                    error=f"Не удалось найти элемент: {selector}"
                )
            
            # Ждем изменений на странице
            self.page.wait_for_load_state("networkidle", timeout=3000)
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="browser_click",
                success=True,
                data=f"Успешно кликнули по {selector}",
                metadata={
                    'selector': selector,
                    'execution_time': execution_time,
                    'new_url': self.page.url
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка клика: {e}")
            return ToolResult(
                tool_name="browser_click",
                success=False,
                data=None,
                error=str(e),
                execution_time=execution_time
            )
    
    def close_session(self) -> ToolResult:
        """Закрывает сессию браузера."""
        try:
            if self.page:
                self.page.close()
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if hasattr(self, 'playwright'):
                self.playwright.stop()
            
            self.page = None
            self.context = None
            self.browser = None
            self.current_url = None
            
            logger.info("Браузер закрыт")
            return ToolResult(
                tool_name="browser_close",
                success=True,
                data="Браузер закрыт",
                confidence=0.9
            )
            
        except Exception as e:
            logger.error(f"Ошибка закрытия браузера: {e}")
            return ToolResult(
                tool_name="browser_close",
                success=False,
                data=None,
                error=str(e)
            )


class ExcelExporter:
    """Класс для экспорта данных в Excel."""
    
    def __init__(self):
        self.available = EXCEL_AVAILABLE
    
    def export_to_excel(self, data: Any, filename: str = None, sheet_name: str = "Данные") -> ToolResult:
        """Экспортирует данные в Excel файл."""
        if not self.available:
            return ToolResult(
                tool_name="excel_export",
                success=False,
                data=None,
                error="Excel поддержка недоступна. Установите: pip install openpyxl"
            )
        
        try:
            if filename is None:
                filename = f"export_{CURRENT_DATE_STR}_{int(time.time())}.xlsx"
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            if EXCEL_ENGINE == 'openpyxl':
                return self._export_with_openpyxl(data, filename, sheet_name)
            else:
                return self._export_with_xlsxwriter(data, filename, sheet_name)
                
        except Exception as e:
            return ToolResult(
                tool_name="excel_export",
                success=False,
                data=None,
                error=str(e)
            )
    
    def _export_with_openpyxl(self, data: Any, filename: str, sheet_name: str) -> ToolResult:
        """Экспорт с использованием openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Настройка стилей
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        if isinstance(data, dict):
            # Экспорт словаря
            row = 1
            for key, value in data.items():
                ws.cell(row=row, column=1, value=str(key))
                ws.cell(row=row, column=2, value=str(value))
                row += 1
            
            # Заголовки
            ws.cell(row=1, column=1, value="Ключ")
            ws.cell(row=1, column=2, value="Значение")
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2).font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            
        elif isinstance(data, list):
            # Экспорт списка
            for row_idx, item in enumerate(data, 1):
                if isinstance(item, dict):
                    # Если это список словарей
                    if row_idx == 1:
                        # Создаем заголовки
                        for col_idx, key in enumerate(item.keys(), 1):
                            cell = ws.cell(row=1, column=col_idx, value=str(key))
                            cell.font = header_font
                            cell.fill = header_fill
                        row_idx = 2
                    
                    for col_idx, value in enumerate(item.values(), 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value))
                else:
                    # Простой список
                    ws.cell(row=row_idx, column=1, value=str(item))
        
        else:
            # Простое значение
            ws.cell(row=1, column=1, value="Данные")
            ws.cell(row=1, column=1).font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=2, column=1, value=str(data))
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем информацию о создании
        info_sheet = wb.create_sheet("Информация")
        info_sheet.cell(row=1, column=1, value="Дата создания")
        info_sheet.cell(row=1, column=2, value=CURRENT_DATE_FORMATTED)
        info_sheet.cell(row=2, column=1, value="Время создания")
        info_sheet.cell(row=2, column=2, value=datetime.now().strftime("%H:%M:%S"))
        
        wb.save(filename)
        
        return ToolResult(
            tool_name="excel_export",
            success=True,
            data=f"Данные экспортированы в {filename}",
            metadata={
                'filename': filename,
                'sheet_name': sheet_name,
                'rows_exported': len(data) if isinstance(data, (list, dict)) else 1,
                'export_date': CURRENT_DATE_STR
            },
            confidence=0.9
        )
    
    def _export_with_xlsxwriter(self, data: Any, filename: str, sheet_name: str) -> ToolResult:
        """Экспорт с использованием xlsxwriter."""
        import xlsxwriter
        
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet(sheet_name)
        
        # Стили
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092'
        })
        
        if isinstance(data, dict):
            worksheet.write(0, 0, "Ключ", header_format)
            worksheet.write(0, 1, "Значение", header_format)
            
            for row, (key, value) in enumerate(data.items(), 1):
                worksheet.write(row, 0, str(key))
                worksheet.write(row, 1, str(value))
                
        elif isinstance(data, list):
            for row, item in enumerate(data):
                if isinstance(item, dict):
                    if row == 0:
                        for col, key in enumerate(item.keys()):
                            worksheet.write(0, col, str(key), header_format)
                    for col, value in enumerate(item.values()):
                        worksheet.write(row + 1, col, str(value))
                else:
                    worksheet.write(row, 0, str(item))
        else:
            worksheet.write(0, 0, "Данные", header_format)
            worksheet.write(1, 0, str(data))
        
        workbook.close()
        
        return ToolResult(
            tool_name="excel_export",
            success=True,
            data=f"Данные экспортированы в {filename}",
            metadata={
                'filename': filename,
                'sheet_name': sheet_name,
                'export_date': CURRENT_DATE_STR
            },
            confidence=0.9
        )


class CodeExecutor:
    """Улучшенный исполнитель Python кода с поддержкой Excel."""
    
    def __init__(self):
        self.globals = {
            '__builtins__': __builtins__,
            'json': json,
            'math': math,
            'random': random,
            're': re,
            'time': time,
            'datetime': datetime,
            'CURRENT_DATE': CURRENT_DATE,
            'CURRENT_DATE_STR': CURRENT_DATE_STR
        }
        
        # Добавляем numpy если доступен
        if SKLEARN_AVAILABLE:
            import numpy as np
            self.globals['np'] = np
        
        # Добавляем Excel поддержку
        if EXCEL_AVAILABLE:
            self.excel_exporter = ExcelExporter()
            self.globals['excel_export'] = self.excel_exporter.export_to_excel
    
    def execute(self, code: str) -> ToolResult:
        """Выполняет Python код с поддержкой Excel операций."""
        start_time = time.time()
        
        try:
            # Создаем изолированное пространство имен
            local_namespace = {}
            
            # Добавляем специальные функции
            local_namespace['save_to_excel'] = self._save_to_excel
            local_namespace['create_excel_report'] = self._create_excel_report
            
            # Перехватываем вывод
            import io
            from contextlib import redirect_stdout, redirect_stderr
            
            stdout_buffer = io.StringIO()
            stderr_buffer = io.StringIO()
            
            with redirect_stdout(stdout_buffer), redirect_stderr(stderr_buffer):
                exec(code, self.globals, local_namespace)
            
            stdout = stdout_buffer.getvalue()
            stderr = stderr_buffer.getvalue()
            
            # Извлекаем результат
            result_data = local_namespace.get('result', stdout if stdout else "Код выполнен успешно")
            
            # Проверяем, были ли созданы Excel файлы
            excel_files = local_namespace.get('excel_files', [])
            
            execution_time = time.time() - start_time
            
            return ToolResult(
                tool_name="code_execute",
                success=True,
                data=result_data,
                metadata={
                    'execution_time': execution_time,
                    'code_length': len(code),
                    'output': stdout,
                    'errors': stderr,
                    'excel_files': excel_files,
                    'execution_date': CURRENT_DATE_STR
                },
                execution_time=execution_time,
                confidence=0.8
            )
            
        except Exception as e:
            execution_time = time.time() - start_time
            logger.error(f"Ошибка выполнения кода: {e}")
            
            return ToolResult(
                tool_name="code_execute",
                success=False,
                data=None,
                error=str(e),
                metadata={
                    'execution_time': execution_time,
                    'code_length': len(code)
                },
                execution_time=execution_time
            )
    
    def _save_to_excel(self, data, filename=None, sheet_name="Данные"):
        """Вспомогательная функция для сохранения в Excel."""
        if EXCEL_AVAILABLE:
            result = self.excel_exporter.export_to_excel(data, filename, sheet_name)
            return result.data if result.success else f"Ошибка: {result.error}"
        else:
            return "Excel поддержка недоступна"
    
    def _create_excel_report(self, title, data_dict, filename=None):
        """Создает отчет в Excel с заданным форматом."""
        if not EXCEL_AVAILABLE:
            return "Excel поддержка недоступна"
        
        try:
            if filename is None:
                filename = f"report_{CURRENT_DATE_STR}.xlsx"
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчет"
            
            # Заголовок отчета
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(size=16, bold=True)
            
            # Дата создания
            ws.cell(row=2, column=1, value=f"Дата создания: {CURRENT_DATE_FORMATTED}")
            
            # Данные
            row = 4
            for section, data in data_dict.items():
                ws.cell(row=row, column=1, value=section)
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                if isinstance(data, dict):
                    for key, value in data.items():
                        ws.cell(row=row, column=1, value=f"  {key}")
                        ws.cell(row=row, column=2, value=str(value))
                        row += 1
                elif isinstance(data, list):
                    for item in data:
                        ws.cell(row=row, column=1, value=f"  • {item}")
                        row += 1
                else:
                    ws.cell(row=row, column=1, value=f"  {data}")
                    row += 1
                
                row += 1  # Пустая строка между секциями
            
            wb.save(filename)
            return f"Отчет сохранен в {filename}"
            
        except Exception as e:
            return f"Ошибка создания отчета: {e}"


class SmartAgent:
    """Умный агент для решения задач пользователя с улучшенным анализом и планированием."""
    
    def __init__(self, gigachat_client: GigaChatClient):
        self.client = gigachat_client
        self.intent_analyzer = AdvancedIntentAnalyzer(gigachat_client)
        self.task_planner = AdvancedTaskPlanner(gigachat_client)

        # Инициализируем инструменты
        self.web_search = WebSearchTool()
        self.web_parser = WebParsingTool()
        self.browser = BrowserTool()
        self.code_executor = CodeExecutor()
        self.excel_exporter = ExcelExporter()
        self.todo_manager = TodoManager()
        self.virtual_fs = VirtualFileSystem()

        # История выполнения
        self.execution_history = []
        
    def get_available_functions(self) -> List[Dict]:
        """Возвращает список доступных функций."""
        functions = []
        
        if self.web_search.available:
            functions.append({
                "name": "web_search",
                "description": "Выполняет поиск в интернете по запросу",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {
                            "type": "string",
                            "description": "Поисковый запрос"
                        },
                        "max_results": {
                            "type": "integer",
                            "description": "Максимальное количество результатов",
                            "default": 5
                        }
                    },
                    "required": ["query"]
                }
            })
        
        if self.web_parser.available:
            functions.append({
                "name": "web_parse",
                "description": "Извлекает содержимое веб-страницы",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {
                            "type": "string",
                            "description": "URL страницы для парсинга"
                        },
                        "extract_focus": {
                            "type": "string",
                            "description": "Фокус для извлечения (ключевые слова)"
                        }
                    },
                    "required": ["url"]
                }
            })
        
        if self.browser.available:
            functions.extend([
                {
                    "name": "browser_navigate",
                    "description": "Переходит на веб-страницу в браузере с поддержкой динамического контента",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "url": {
                                "type": "string",
                                "description": "URL для перехода"
                            }
                        },
                        "required": ["url"]
                    }
                },
                {
                    "name": "browser_extract",
                    "description": "Извлекает контент со страницы в браузере с ожиданием динамического контента",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "selector": {
                                "type": "string",
                                "description": "CSS селектор для извлечения (опционально)"
                            },
                            "wait_for_element": {
                                "type": "boolean",
                                "description": "Ждать появления элемента",
                                "default": True
                            }
                        }
                    }
                },
                {
                    "name": "browser_click",
                    "description": "Кликает по элементу на странице",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "selector": {
                                "type": "string",
                                "description": "Селектор элемента для клика"
                            }
                        },
                        "required": ["selector"]
                    }
                },
                {
                    "name": "wait_dynamic_content",
                    "description": "Ждет загрузки динамического контента на странице",
                    "parameters": {
                        "type": "object",
                        "properties": {
                            "timeout": {
                                "type": "integer",
                                "description": "Таймаут ожидания в миллисекундах",
                                "default": 10000
                            }
                        }
                    }
                }
            ])
        
        functions.append({
            "name": "code_execute",
            "description": "Выполняет Python код для вычислений, анализа и экспорта в Excel",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Python код для выполнения. Доступны функции save_to_excel() и create_excel_report()"
                    }
                },
                "required": ["code"]
            }
        })
        
        if self.excel_exporter.available:
            functions.append({
                "name": "excel_export",
                "description": "Экспортирует данные в Excel файл",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "data": {
                            "type": "string",
                            "description": "Данные для экспорта в JSON формате"
                        },
                        "filename": {
                            "type": "string",
                            "description": "Имя файла (опционально)"
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Название листа",
                            "default": "Данные"
                        }
                    },
                    "required": ["data"]
                }
            })

        # Инструменты виртуальной файловой системы
        functions.extend([
            {
                "name": "ls",
                "description": "Возвращает список файлов в виртуальной файловой системе",
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
            },
            {
                "name": "read_file",
                "description": "Читает файл с выводом номеров строк (как cat -n)",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Имя файла для чтения",
                        },
                        "offset": {
                            "type": "integer",
                            "description": "С какой строки начать вывод (0 по умолчанию)",
                            "default": 0,
                        },
                        "limit": {
                            "type": "integer",
                            "description": "Сколько строк прочитать (2000 по умолчанию)",
                            "default": 2000,
                        },
                    },
                    "required": ["file_path"],
                },
            },
            {
                "name": "write_file",
                "description": "Создает новый файл или полностью перезаписывает существующий",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Имя файла",
                        },
                        "content": {
                            "type": "string",
                            "description": "Полное содержимое файла",
                        },
                    },
                    "required": ["file_path", "content"],
                },
            },
            {
                "name": "edit_file",
                "description": "Точечно изменяет существующий файл, заменяя одну строку на другую",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "file_path": {
                            "type": "string",
                            "description": "Имя файла",
                        },
                        "old_string": {
                            "type": "string",
                            "description": "Текст, который нужно заменить (без номеров строк)",
                        },
                        "new_string": {
                            "type": "string",
                            "description": "Новый текст",
                        },
                        "replace_all": {
                            "type": "boolean",
                            "description": "Заменить все вхождения строки",
                            "default": False,
                        },
                    },
                    "required": ["file_path", "old_string", "new_string"],
                },
            },
        ])

        # Инструмент планирования задач
        functions.append(
            {
                "name": "write_todos",
                "description": "Создает или обновляет структурированный todo-список задач",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "todos": {
                            "type": "array",
                            "description": "Полный список задач с указанием статусов",
                            "items": {
                                "type": "object",
                                "properties": {
                                    "content": {
                                        "type": "string",
                                        "description": "Текст задачи",
                                    },
                                    "status": {
                                        "type": "string",
                                        "enum": ["pending", "in_progress", "completed"],
                                        "description": "Статус задачи",
                                    },
                                },
                                "required": ["content"],
                            },
                        }
                    },
                    "required": ["todos"],
                },
            }
        )

        functions.append({
            "name": "finish_task",
            "description": "Завершает выполнение задачи с финальным ответом",
            "parameters": {
                "type": "object",
                "properties": {
                    "answer": {
                        "type": "string",
                        "description": "Финальный ответ пользователю"
                    }
                },
                "required": ["answer"]
            }
        })
        
        return functions
    
    def execute_function(self, function_name: str, arguments: Dict) -> ToolResult:
        """Выполняет функцию."""
        try:
            if function_name == "web_search":
                return self.web_search.search(
                    query=arguments.get("query"),
                    max_results=arguments.get("max_results", 5)
                )
            
            elif function_name == "web_parse":
                return self.web_parser.parse(
                    url=arguments.get("url"),
                    extract_focus=arguments.get("extract_focus")
                )
            
            elif function_name == "browser_navigate":
                return self.browser.navigate(url=arguments.get("url"))
            
            elif function_name == "browser_extract":
                return self.browser.extract_content(
                    selector=arguments.get("selector"),
                    wait_for_element=arguments.get("wait_for_element", True)
                )
            
            elif function_name == "browser_click":
                return self.browser.click_element(selector=arguments.get("selector"))
            
            elif function_name == "wait_dynamic_content":
                return self.browser.wait_for_dynamic_content(
                    timeout=arguments.get("timeout", 10000)
                )
            
            elif function_name == "code_execute":
                return self.code_executor.execute(code=arguments.get("code"))
            
            elif function_name == "excel_export":
                try:
                    data = json.loads(arguments.get("data"))
                except:
                    data = arguments.get("data")

                return self.excel_exporter.export_to_excel(
                    data=data,
                    filename=arguments.get("filename"),
                    sheet_name=arguments.get("sheet_name", "Данные")
                )

            elif function_name == "ls":
                start_time = time.time()
                files = self.virtual_fs.list_files()
                return ToolResult(
                    tool_name="ls",
                    success=True,
                    data=files,
                    metadata={"total_files": len(files)},
                    execution_time=time.time() - start_time,
                    confidence=0.95,
                )

            elif function_name == "read_file":
                return self.virtual_fs.read_file(
                    file_path=arguments.get("file_path", ""),
                    offset=arguments.get("offset", 0),
                    limit=arguments.get("limit", 2000),
                )

            elif function_name == "write_file":
                return self.virtual_fs.write_file(
                    file_path=arguments.get("file_path", ""),
                    content=arguments.get("content", ""),
                )

            elif function_name == "edit_file":
                replace_all = arguments.get("replace_all", False)
                if isinstance(replace_all, str):
                    replace_all = replace_all.lower() == "true"

                return self.virtual_fs.edit_file(
                    file_path=arguments.get("file_path", ""),
                    old_string=arguments.get("old_string", ""),
                    new_string=arguments.get("new_string", ""),
                    replace_all=replace_all,
                )

            elif function_name == "write_todos":
                return self.todo_manager.write_todos(arguments.get("todos", []))

            elif function_name == "finish_task":
                return ToolResult(
                    tool_name="finish_task",
                    success=True,
                    data=arguments.get("answer"),
                    confidence=1.0
                )
            
            else:
                return ToolResult(
                    tool_name=function_name,
                    success=False,
                    data=None,
                    error=f"Неизвестная функция: {function_name}"
                )
                
        except Exception as e:
            return ToolResult(
                tool_name=function_name,
                success=False,
                data=None,
                error=str(e)
            )
    
    def _format_plan_steps(self, steps):
        """Форматирует шаги плана в читаемый вид."""
        formatted_steps = []
        for i, step in enumerate(steps, 1):
            description = step.get('description', step['tool'])
            priority = step.get('priority', 'не указан')
            formatted_steps.append(f"{i}. {description} (приоритет: {priority})")
        return "\n".join(formatted_steps)
    
    def build_enhanced_system_prompt(self, context: TaskContext, plan: ExecutionPlan) -> str:
        """Создает улучшенный системный промпт с метаролью и я-конструкциями."""
        
        excel_info = ""
        if EXCEL_AVAILABLE:
            excel_info = """
- Я могу экспортировать данные в Excel используя excel_export или code_execute с save_to_excel()
- У меня есть функция save_to_excel(data, filename, sheet_name) в code_execute
- Я могу создать форматированный отчет через create_excel_report(title, data_dict, filename)"""

        tools_status = f"""
МОИ ДОСТУПНЫЕ ИНСТРУМЕНТЫ:
🔍 Веб-поиск: {'✅ Работает' if self.web_search.available else '❌ Недоступен'}
📄 Парсинг страниц: {'✅ Работает' if self.web_parser.available else '❌ Недоступен'}
🌐 Браузер: {'✅ Работает' if self.browser.available else '❌ Недоступен'}
💻 Выполнение кода: ✅ Работает
📊 Excel экспорт: {'✅ Работает' if EXCEL_AVAILABLE else '❌ Недоступен'}
📁 Виртуальная ФС: ✅ Всегда доступна
📝 Планирование через todo: ✅ Всегда доступно"""

        plan_reasoning = plan.reasoning if plan.reasoning else "План создан на основе базовых шаблонов"

        # Исправляем строку с форматированием
        plan_steps_formatted = '\n'.join([f'{i}. {step.get("description", step["tool"])} (приоритет: {step.get("priority", "не указан")})' for i, step in enumerate(plan.steps, 1)])

        fs_guidelines = """РАБОТА С ВИРТУАЛЬНОЙ ФАЙЛОВОЙ СИСТЕМОЙ:
- Перед чтением или редактированием обязательно вызывай ls() для просмотра файлов
- Всегда используй read_file() перед edit_file() и не включай номера строк в old_string
- Применяй write_file() для новых файлов или полной перезаписи, а для точечных правок — только edit_file()
- Если строка встречается несколько раз, уточняй контекст или используй replace_all=True"""

        planning_guidelines = """ПЛАНИРОВАНИЕ С ПОМОЩЬЮ TODO:
- Задействуй write_todos при сложных задачах (3+ шагов) или по прямой просьбе пользователя
- Держи хотя бы одну задачу в статусе in_progress и отмечай завершенные как completed
- Своевременно удаляй или обновляй устаревшие пункты и избегай чрезмерно мелких задач"""

        return f"""Я - продвинутый интеллектуальный агент X-Master v77 Enhanced. Моя роль - эффективно решать задачи пользователей, используя доступные инструменты и глубокий анализ.

МОЯ ТЕКУЩАЯ СИТУАЦИЯ:
📅 Сегодня: {CURRENT_DATE_FORMATTED}
🎯 Задача пользователя: "{context.query}"
🧠 Мой анализ намерений: {context.intent} (уверенность: {context.confidence_score:.1%})
📋 Цель пользователя: {context.user_goal}
⚡ Сложность: {context.complexity}
🌍 Область: {context.domain}
⏰ Срочность: {context.urgency}
🕐 Временной контекст: {context.temporal_context}

{tools_status}

МОЙ ПЛАН ДЕЙСТВИЙ:
{plan_reasoning}

Конкретные шаги:
{plan_steps_formatted}

{fs_guidelines}

{planning_guidelines}

Критерии успеха: {', '.join(plan.success_criteria) if plan.success_criteria else 'Полный и точный ответ пользователю'}

МОИ ПРИНЦИПЫ РАБОТЫ:
1. Я всегда начинаю с глубокого анализа поставленной задачи
2. Я использую инструменты последовательно и обдуманно
3. Я проверяю результаты каждого шага перед переходом к следующему
4. Для динамических сайтов я обязательно жду загрузки контента
5. Я синтезирую информацию из разных источников для полного ответа
6. Я учитываю текущую дату при поиске актуальной информации
7. Для финансовых данных я использую браузер для работы с официальными сайтами{excel_info}
8. Я ОБЯЗАТЕЛЬНО завершаю каждую задачу вызовом finish_task с исчерпывающим ответом

МОИ МЕТАКОГНИТИВНЫЕ СПОСОБНОСТИ:
- Я анализирую свои действия и корректирую план при необходимости
- Я понимаю ограничения своих инструментов и адаптируюсь
- Я оцениваю качество получаемых данных и ищу дополнительные источники
- Я помню контекст всего диалога и использую накопленную информацию

СТРАТЕГИЯ РАБОТЫ С ВРЕМЕННЫМИ ДАННЫМИ:
Поскольку сегодня {CURRENT_DATE_FORMATTED}, я:
- Добавляю текущую дату к поисковым запросам об актуальных событиях
- Использую браузер для получения свежих данных с официальных сайтов
- Указываю дату получения информации в своих ответах
- Проверяю актуальность найденной информации

Теперь я приступаю к выполнению задачи с полной концентрацией и профессионализмом!"""
    
    def process_query(self, query: str, max_iterations: int = 15) -> Dict[str, Any]:
        """Обрабатывает запрос пользователя с улучшенным анализом."""
        logger.info(f"Начинаю обработку запроса: {query}")
        logger.info(f"Текущая дата для контекста: {CURRENT_DATE_FORMATTED}")
        
        # Глубокий анализ намерений с помощью LLM
        context = self.intent_analyzer.analyze_with_llm(query)
        logger.info(f"Результат анализа намерений: {context}")
        
        # Создаем умный план выполнения
        plan = self.task_planner.create_smart_plan(context)
        logger.info(f"Создан план: {[step['tool'] for step in plan.steps]}")
        logger.info(f"Обоснование плана: {plan.reasoning}")
        
        # Инициализируем диалог с улучшенным промптом
        messages = [
            {"role": "system", "content": self.build_enhanced_system_prompt(context, plan)},
            {"role": "user", "content": query}
        ]
        
        execution_log = []
        final_answer = None
        
        for iteration in range(max_iterations):
            try:
                logger.info(f"Итерация {iteration + 1}/{max_iterations}")
                
                # Отправляем запрос в GigaChat
                response = self.client.chat(
                    messages=messages,
                    functions=self.get_available_functions(),
                    temperature=0.2,  # Снижаем температуру для более точного следования плану
                    function_call="auto"
                )
                
                if not response or 'choices' not in response:
                    logger.error("Некорректный ответ от GigaChat")
                    break
                
                choice = response['choices'][0]
                message = choice['message']
                
                # Добавляем сообщение в диалог
                messages.append({
                    "role": message['role'],
                    "content": message.get('content', ''),
                    "function_call": message.get('function_call')
                })
                
                # Обрабатываем вызов функции
                if 'function_call' in message:
                    func_call = message['function_call']
                    func_name = func_call.get('name')
                    
                    # Парсим аргументы
                    try:
                        if isinstance(func_call.get('arguments'), str):
                            func_args = json.loads(func_call['arguments'])
                        else:
                            func_args = func_call.get('arguments', {})
                    except json.JSONDecodeError:
                        func_args = {}
                    
                    logger.info(f"Выполняется функция: {func_name} с аргументами: {func_args}")
                    
                    # Выполняем функцию
                    result = self.execute_function(func_name, func_args)
                    execution_log.append({
                        'function': func_name,
                        'arguments': func_args,
                        'result': result,
                        'iteration': iteration + 1,
                        'timestamp': datetime.now().isoformat(),
                        'planned': func_name in [step['tool'] for step in plan.steps]
                    })
                    
                    # Проверяем на завершение задачи
                    if func_name == "finish_task" and result.success:
                        final_answer = result.data
                        break
                    
                    # Добавляем результат в диалог с дополнительным контекстом
                    function_response = {
                        "role": "function",
                        "name": func_name,
                        "content": json.dumps({
                            "success": result.success,
                            "data": str(result.data)[:2000] if result.data else None,  # Увеличиваем лимит
                            "error": result.error,
                            "metadata": result.metadata,
                            "confidence": result.confidence,
                            "execution_time": result.execution_time
                        }, ensure_ascii=False)
                    }
                    messages.append(function_response)
                
                else:
                    # Модель ответила без вызова функции
                    content = message.get('content', '')
                    if content and not final_answer:
                        # Принудительно вызываем finish_task
                        result = self.execute_function("finish_task", {"answer": content})
                        execution_log.append({
                            'function': 'finish_task',
                            'arguments': {"answer": content},
                            'result': result,
                            'iteration': iteration + 1,
                            'timestamp': datetime.now().isoformat(),
                            'planned': True
                        })
                        final_answer = content
                        break
                
            except Exception as e:
                logger.error(f"Ошибка в итерации {iteration + 1}: {e}")
                break
        
        # Закрываем браузер если был открыт
        try:
            if self.browser.page:
                self.browser.close_session()
        except:
            pass
        
        # Анализируем выполнение
        successful_tools = sum(1 for log in execution_log if log['result'].success)
        total_tools = len(execution_log)
        
        # Получаем списки инструментов
        planned_tools = [step['tool'] for step in plan.steps]
        executed_tools = [log['function'] for log in execution_log]
        
        # Анализируем следование плану
        plan_adherence = sum(1 for log in execution_log if log.get('planned', False))
        plan_adherence_percent = (plan_adherence / max(total_tools, 1)) * 100
        
        # Оцениваем качество выполнения
        quality_score = 0.0
        if final_answer:
            quality_score += 0.4  # Есть финальный ответ
        if successful_tools > 0:
            quality_score += 0.3 * (successful_tools / max(total_tools, 1))  # Успешность инструментов
        if plan_adherence_percent > 50:
            quality_score += 0.3  # Следование плану
        
        return {
            'success': final_answer is not None,
            'final_answer': final_answer or "Не удалось получить ответ",
            'context': context,
            'plan': plan,
            'execution_log': execution_log,
            'total_iterations': iteration + 1,
            'tools_used': total_tools,
            'successful_tools': successful_tools,
            'confidence': quality_score,
            'query_date': CURRENT_DATE_STR,
            'planned_tools': planned_tools,
            'executed_tools': executed_tools,
            'plan_adherence_percent': plan_adherence_percent,
            'plan_followed': plan_adherence_percent > 70,
            'excel_support': EXCEL_AVAILABLE,
            'llm_analysis_used': context.meta_analysis.get('llm_analysis', False),
            'analysis_reasoning': context.meta_analysis.get('reasoning', ''),
            'plan_reasoning': plan.reasoning,
            'quality_score': quality_score,
            'risk_assessment': plan.risk_assessment,
            'todos': self.todo_manager.get_todos(),
            'virtual_files': self.virtual_fs.list_files(),
        }


def main():
    """Основное Streamlit приложение."""
    st.set_page_config(
        page_title="Умный X-Master Agent v77 Enhanced+",
        page_icon="🤖",
        layout="wide"
    )

    st.title("🤖 Умный X-Master Agent v77 Enhanced+")
    st.markdown(f"**Интеллектуальный агент с продвинутым анализом намерений и умным планированием задач**")
    st.caption(f"📅 Текущая дата: {CURRENT_DATE_FORMATTED}")

    # Боковая панель конфигурации
    with st.sidebar:
        st.header("⚙️ Настройки")
        
        # Отображаем текущую дату в сайдбаре
        st.info(f"📅 **Сегодня:** {CURRENT_DATE_FORMATTED}")

        client_id = st.text_input("GigaChat Client ID", type="password")
        client_secret = st.text_input("GigaChat Client Secret", type="password")

        st.markdown("---")

        model_name = st.selectbox(
            "Модель GigaChat",
            ["GigaChat-2-Max", "GigaChat", "GigaChat-2-Pro"],
            index=0
        )
        
        max_iterations = st.slider(
            "Максимум итераций",
            min_value=5,
            max_value=20,
            value=15,
            help="Максимальное количество шагов для решения задачи"
        )
        
        st.markdown("---")
        st.subheader("🛠️ Доступные инструменты")
        
        # Показываем статус инструментов
        tools_status = [
            ("🔍 Веб-поиск", DDGS_AVAILABLE),
            ("📄 Парсинг веб-страниц", trafilatura is not None),
            ("🌐 Браузер с динамическим контентом", PLAYWRIGHT_AVAILABLE),
            ("💻 Выполнение кода", True),
            ("📊 Экспорт в Excel", EXCEL_AVAILABLE),
            ("🧮 Научные вычисления", SKLEARN_AVAILABLE),
            ("🧠 LLM анализ намерений", True),
            ("📋 Умное планирование", True),
            ("📁 Виртуальная файловая система", True),
            ("📝 Todo-план", True),
        ]
        
        for tool_name, available in tools_status:
            if available:
                st.success(f"✅ {tool_name}")
            else:
                st.error(f"❌ {tool_name}")
        
        st.markdown("---")
        st.info(
            "**🚀 Новые возможности v77 Enhanced+:**\n"
            "• Глубокий LLM-анализ намерений пользователя\n"
            "• Умное планирование с метарассуждениями\n"
            "• Я-конструкции в системных промптах\n"
            "• Улучшенная оценка рисков и адаптивность\n"
            "• Контекстуальное понимание временных запросов\n"
            "• Продвинутая работа с финансовыми данными\n"
            "• Автоматическое следование критериям успеха\n"
            "• Метакогнитивный самоанализ выполнения\n"
            "• Повышенная точность и надежность результатов"
        )

    # Основной интерфейс
    st.markdown("### 💬 Интерфейс запросов")
    
    col1, col2 = st.columns([3, 1])
    
    with col1:
        query = st.text_area(
            "Введите ваш запрос:",
            height=120,
            placeholder="Примеры запросов:\n• Найди последние новости об искусственном интеллекте\n• Какая ключевая ставка ЦБ РФ сегодня? Создай отчет в Excel\n• Что произошло на этой неделе в мире технологий?\n• Проанализируй текущую ситуацию на рынке криптовалют\n• Какие события запланированы на сегодня в России?\n• Сравни курсы валют и экспортируй данные",
            help="Агент автоматически проанализирует ваши намерения и создаст оптимальный план решения"
        )
    
    with col2:
        st.markdown("**🎯 Примеры задач:**")
        example_queries = [
            "Последние новости ИИ",
            "Ключевая ставка ЦБ в Excel",
            "События недели в России",
            "Анализ криптовалют",
            "Курсы валют сегодня",
            "Технологические тренды"
        ]
        
        for i, example in enumerate(example_queries):
            if st.button(f"🔍 {example}", key=f"example_{i}", use_container_width=True):
                query = example
                st.rerun()

    # Кнопка выполнения
    execute_button = st.button(
        "🚀 Выполнить задачу", 
        type="primary", 
        use_container_width=True,
        disabled=not (client_id and client_secret and query)
    )

    if execute_button:
        try:
            # Инициализируем агента
            client = GigaChatClient(client_id, client_secret, model=model_name)
            agent = SmartAgent(client)
            
            # Прогресс выполнения
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            with st.spinner("🤖 Агент выполняет глубокий анализ и умное планирование..."):
                status_text.text("🧠 Анализ намерений с помощью LLM...")
                progress_bar.progress(10)
                
                status_text.text("📋 Создание умного плана выполнения...")
                progress_bar.progress(20)
                
                # Обрабатываем запрос
                result = agent.process_query(query, max_iterations)
                progress_bar.progress(100)
                status_text.text("✅ Задача выполнена с использованием ИИ-анализа!")

            st.markdown("---")
            
            # Результаты
            st.subheader("📋 Результат")
            
            if result['success'] and result['final_answer']:
                st.success("**Ответ агента:**")
                st.write(result['final_answer'])
            else:
                st.error("Агент не смог найти ответ на задачу")

            # Расширенные метрики выполнения
            st.markdown("---")
            st.subheader("📊 Расширенные метрики выполнения")
            
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            
            with col1:
                st.metric(
                    "🔧 Инструментов",
                    result['tools_used'],
                    f"Успешно: {result['successful_tools']}"
                )
            
            with col2:
                st.metric(
                    "🔄 Итераций",
                    result['total_iterations'],
                    f"Лимит: {max_iterations}"
                )
            
            with col3:
                st.metric(
                    "🎯 Качество",
                    f"{result['quality_score']:.1%}",
                    "Общая оценка"
                )
            
            with col4:
                st.metric(
                    "📋 Следование плану",
                    f"{result.get('plan_adherence_percent', 0):.0f}%",
                    "✅ Хорошо" if result.get('plan_followed', False) else "⚠️ Частично"
                )
            
            with col5:
                llm_status = "🧠 LLM" if result.get('llm_analysis_used', False) else "📝 Базовый"
                st.metric(
                    "🔍 Анализ намерений",
                    llm_status,
                    f"Уверенность: {result['context'].confidence_score:.1%}"
                )
            
            with col6:
                st.metric(
                    "📊 Excel поддержка",
                    "✅ Работает" if result.get('excel_support', False) else "❌ Недоступен",
                    f"Движок: {EXCEL_ENGINE if EXCEL_AVAILABLE else 'Нет'}"
                )

            # Анализ намерений и планирование
            st.markdown("---")
            st.subheader("🧠 Интеллектуальный анализ и планирование")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**🎯 Анализ намерений пользователя:**")
                context = result['context']
                analysis_info = {
                    "Основное намерение": context.intent,
                    "Цель пользователя": context.user_goal,
                    "Сложность задачи": context.complexity,
                    "Предметная область": context.domain,
                    "Срочность": context.urgency,
                    "Временной контекст": context.temporal_context,
                    "Ожидаемые источники": context.expected_sources,
                    "Уверенность анализа": f"{context.confidence_score:.1%}",
                    "Использован LLM": "Да" if result.get('llm_analysis_used') else "Нет"
                }
                
                for key, value in analysis_info.items():
                    st.text(f"{key}: {value}")
                
                if result.get('analysis_reasoning'):
                    st.markdown("**💭 Обоснование анализа:**")
                    st.text_area("", result['analysis_reasoning'], height=100, disabled=True)
            
            with col2:
                st.markdown("**📋 Умное планирование задачи:**")
                plan = result['plan']
                plan_info = {
                    "Количество шагов": len(plan.steps),
                    "Оценочное время": f"{plan.estimated_time:.1f} сек",
                    "Уверенность в плане": f"{plan.confidence:.1%}",
                    "Уровень адаптивности": plan.adaptability_level
                }
                
                for key, value in plan_info.items():
                    st.text(f"{key}: {value}")
                
                if plan.reasoning:
                    st.markdown("**🎯 Логика планирования:**")
                    st.text_area("", plan.reasoning, height=100, disabled=True, key="plan_reasoning")

            # Сравнение плана и выполнения
            if result.get('planned_tools') and result.get('executed_tools'):
                st.markdown("---")
                st.subheader("📋 Выполнение плана vs. Реальность")
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown("**📝 Запланировано:**")
                    for i, tool in enumerate(result['planned_tools'], 1):
                        if tool in result['executed_tools']:
                            st.success(f"{i}. ✅ {tool}")
                        else:
                            st.error(f"{i}. ❌ {tool}")
                
                with col2:
                    st.markdown("**⚡ Выполнено:**")
                    for i, tool in enumerate(result['executed_tools'], 1):
                        log_entry = next((log for log in result['execution_log'] if log['function'] == tool), {})
                        planned_marker = "📋" if log_entry.get('planned', False) else "🔄"
                        st.success(f"{i}. {planned_marker} {tool}")
                
                with col3:
                    st.markdown("**🎯 Критерии успеха:**")
                    success_criteria = plan.success_criteria
                    if success_criteria:
                        for i, criterion in enumerate(success_criteria, 1):
                            st.info(f"{i}. {criterion}")
                    else:
                        st.info("Стандартные критерии")

            # Детали контекста
            with st.expander("🧠 Детальный анализ контекста", expanded=False):
                context = result['context']

                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**Определенные потребности:**")
                    needs = {
                        "Поиск информации": context.requires_search,
                        "Работа с браузером": context.requires_browser,
                        "Вычисления": context.requires_computation,
                        "Экспорт в Excel": context.requires_excel
                    }
                    st.json(needs)

                with col2:
                    st.markdown("**Извлеченные ключевые слова:**")
                    if context.keywords:
                        st.write(", ".join(context.keywords))
                    else:
                        st.write("Ключевые слова не выделены")

                if context.meta_analysis:
                    st.markdown("**📊 Метаанализ:**")
                    st.json(context.meta_analysis)

            with st.expander("📝 Текущий todo-план", expanded=False):
                todos = result.get('todos', [])
                if todos:
                    for i, todo in enumerate(todos, 1):
                        status = todo.get('status', 'pending')
                        st.write(f"{i}. [{status}] {todo.get('content', '')}")
                else:
                    st.info("Список задач пуст")

            with st.expander("📁 Виртуальная файловая система", expanded=False):
                files = result.get('virtual_files', [])
                if files:
                    st.write("\n".join(files))
                else:
                    st.info("Файлы еще не созданы")

            # План выполнения с рисками
            with st.expander("📋 Детальный план с оценкой рисков", expanded=False):
                plan = result['plan']

                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Запланированные шаги:**")
                    for i, step in enumerate(plan.steps, 1):
                        st.markdown(f"{i}. **{step['tool']}**")
                        if 'description' in step:
                            st.markdown(f"   📝 {step['description']}")
                        if 'priority' in step:
                            st.markdown(f"   ⚡ Приоритет: {step['priority']}")
                        if 'expected_outcome' in step:
                            st.markdown(f"   🎯 Ожидаемый результат: {step['expected_outcome']}")
                
                with col2:
                    st.markdown("**🚨 Оценка рисков:**")
                    if plan.risk_assessment:
                        st.json(plan.risk_assessment)
                    else:
                        st.info("Анализ рисков не проводился")
                
                if plan.fallback_plan:
                    st.markdown("**🔄 Резервный план:**")
                    for i, step in enumerate(plan.fallback_plan, 1):
                        st.markdown(f"{i}. {step.get('tool', 'Неизвестный инструмент')}")

            # Детальный лог выполнения
            with st.expander("📜 Детальный лог выполнения", expanded=False):
                if result['execution_log']:
                    for i, log_entry in enumerate(result['execution_log'], 1):
                        func_result = log_entry['result']
                        
                        # Заголовок шага
                        success_emoji = "✅" if func_result.success else "❌"
                        planned_emoji = "📋" if log_entry.get('planned', False) else "🔄"
                        st.markdown(f"### {success_emoji} {planned_emoji} Шаг {i}: {log_entry['function']}")
                        st.caption(f"⏰ Время: {log_entry.get('timestamp', 'Не указано')}")
                        
                        # Информация о выполнении
                        exec_col1, exec_col2 = st.columns(2)
                        
                        with exec_col1:
                            # Аргументы
                            if log_entry['arguments']:
                                st.markdown("**📝 Аргументы:**")
                                st.json(log_entry['arguments'])
                        
                        with exec_col2:
                            # Результат
                            if func_result.success:
                                st.success(f"✅ Успешно (уверенность: {func_result.confidence:.1%})")
                                if func_result.execution_time:
                                    st.caption(f"⏱️ Время выполнения: {func_result.execution_time:.2f}с")
                            else:
                                st.error(f"❌ Ошибка: {func_result.error}")
                        
                        # Данные результата
                        if func_result.data:
                            data_preview = str(func_result.data)[:300]
                            if len(str(func_result.data)) > 300:
                                data_preview += "..."
                            st.text_area(f"📊 Результат:", data_preview, height=100, disabled=True, key=f"result_{i}")
                        
                        # Метаданные
                        if func_result.metadata:
                            st.markdown("**🔍 Метаданные:**")
                            st.json(func_result.metadata)
                        
                        st.markdown("---")
                else:
                    st.info("Функции не вызывались")

        except Exception as e:
            st.error(f"⚠️ Ошибка выполнения: {str(e)}")
            logger.error("Ошибка выполнения задачи", exc_info=True)


if __name__ == "__main__":
    main()